#!/usr/bin/env python3
"""Regression tests for the fault p10 model-results deliverable."""
from __future__ import annotations

import csv
import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook


TRACK_DIR = Path(__file__).resolve().parent


def _load_module():
    spec = importlib.util.spec_from_file_location(
        "fault_p10_model_results", TRACK_DIR / "fault_p10_model_results.py"
    )
    if spec is None or spec.loader is None:
        raise ImportError("cannot load fault_p10_model_results.py")
    module = importlib.util.module_from_spec(spec)
    sys.modules["fault_p10_model_results"] = module
    spec.loader.exec_module(module)
    return module


fault_p10_model_results = _load_module()


class FaultP10ModelResultsTests(unittest.TestCase):
    def test_run_emits_single_sheet_workbook_and_manifests(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory) / "p10_model_results"
            report = fault_p10_model_results.run(root)
            workbook = root / fault_p10_model_results.WORKBOOK_NAME
            figures_manifest = root / fault_p10_model_results.FIGURES_MANIFEST_NAME
            tables_manifest = root / fault_p10_model_results.TABLES_MANIFEST_NAME
            audit_report = root / fault_p10_model_results.AUDIT_REPORT_NAME
            figure = root / fault_p10_model_results.FIGURE_NAME

            self.assertTrue(workbook.is_file())
            self.assertTrue(figures_manifest.is_file())
            self.assertTrue(tables_manifest.is_file())
            self.assertTrue(audit_report.is_file())
            self.assertTrue(figure.is_file())
            self.assertGreater(figure.stat().st_size, 1000)
            with Image.open(figure) as opened:
                opened.verify()

            book = load_workbook(workbook)
            self.assertEqual(book.sheetnames, [fault_p10_model_results.WORKBOOK_SHEET_NAME])
            sheet = book[fault_p10_model_results.WORKBOOK_SHEET_NAME]
            header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(header, fault_p10_model_results.HEADER)
            self.assertEqual(len(header), 26)
            self.assertEqual(sheet.max_column, 26)
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 8)

            rows_by_status = {row[19]: [] for row in rows}
            for row in rows:
                rows_by_status[row[19]].append(row)
                evidence_path = TRACK_DIR / row[20]
                self.assertTrue(evidence_path.is_file())
                self.assertIn(row[21], (None, ""))
                self.assertEqual(row[22], report["source_commit"])
                self.assertIn(row[5], (True, False))
            self.assertEqual(sorted(rows_by_status), ["data_blocked", "reference_only"])
            self.assertEqual(len(rows_by_status["reference_only"]), 7)
            self.assertEqual(len(rows_by_status["data_blocked"]), 1)
            blocked = rows_by_status["data_blocked"][0]
            self.assertEqual(blocked[3], "blueyo0/SAM-Med3D:sam_med3d_turbo.pth")
            self.assertIsNone(blocked[13])
            self.assertEqual(blocked[19], "data_blocked")
            self.assertEqual(blocked[24], "none; preserved fail-closed gate and evidence-only SAM-Med3D audit")

            with figures_manifest.open(encoding="utf-8") as handle:
                figure_rows = list(csv.DictReader(handle))
            self.assertEqual(len(figure_rows), 6)
            self.assertTrue(any(row["kind"] == "generated_chart" for row in figure_rows))
            self.assertTrue(
                any(
                    row["path"] == str(figure)
                    for row in figure_rows
                )
            )

            with tables_manifest.open(encoding="utf-8") as handle:
                table_rows = list(csv.DictReader(handle))
            self.assertEqual(len(table_rows), 8)
            self.assertTrue(any(row["path"] == str(workbook) for row in table_rows))
            self.assertTrue(any(row["path"] == str(fault_p10_model_results.P9_SAMMED3D_SUMMARY.relative_to(TRACK_DIR)) for row in table_rows))

            report_text = audit_report.read_text(encoding="utf-8")
            self.assertTrue(report_text.startswith("# Fault P10 model-results audit"))
            self.assertIn("## Conclusion", report_text)
            self.assertIn("## Evidence summary", report_text)
            self.assertIn("## Preserved scientific conclusion", report_text)
            self.assertNotIn("{", report_text.splitlines()[0])

    def test_source_requires_present_evidence_files(self) -> None:
        self.assertEqual(fault_p10_model_results.PRIMARY_METRIC, "average_precision")
        self.assertTrue(fault_p10_model_results.AUDITED_V2_BASELINE.is_file())
        self.assertTrue(fault_p10_model_results.P9_SAMMED3D_SUMMARY.is_file())


if __name__ == "__main__":
    unittest.main()
