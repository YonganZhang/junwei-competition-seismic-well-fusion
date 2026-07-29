from __future__ import annotations

import csv
import json
import subprocess
import sys
import unittest
from pathlib import Path

TRACK_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = TRACK_DIR.parents[2]
for root in (str(PROJECT_ROOT), str(TRACK_DIR)):
    if root not in sys.path:
        sys.path.insert(0, root)

import p10_model_results


OUTPUT_DIR = TRACK_DIR / "_outputs" / "p10_model_results"


class P10ModelResultsTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.outputs = {
            "track_model_metrics.xlsx": OUTPUT_DIR / "track_model_metrics.xlsx",
            "figures_manifest.csv": OUTPUT_DIR / "figures_manifest.csv",
            "tables_manifest.csv": OUTPUT_DIR / "tables_manifest.csv",
            "audit_report.md": OUTPUT_DIR / "audit_report.md",
            "before_after_primary_metric.png": OUTPUT_DIR / "before_after_primary_metric.png",
        }

    def test_expected_artifacts_exist(self) -> None:
        expected = {
            "track_model_metrics.xlsx",
            "figures_manifest.csv",
            "tables_manifest.csv",
            "audit_report.md",
            "before_after_primary_metric.png",
        }
        self.assertEqual(set(self.outputs), expected)
        for path in self.outputs.values():
            self.assertTrue(path.exists(), path)
            self.assertGreater(path.stat().st_size, 0, path)

    def test_workbook_reopens_with_single_metrics_sheet(self) -> None:
        script = (
            "from openpyxl import load_workbook; "
            f"wb = load_workbook(r'{self.outputs['track_model_metrics.xlsx']}', read_only=True, data_only=True); "
            "sheet = wb['模型指标']; "
            "header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]; "
            "import json; print(json.dumps({'sheetnames': wb.sheetnames, 'header': header}))"
        )
        completed = subprocess.run(["python3", "-c", script], check=True, capture_output=True, text=True)
        payload = json.loads(completed.stdout)
        self.assertEqual(payload["sheetnames"], ["模型指标"])
        header = payload["header"]
        self.assertEqual(header[0:6], ["track", "dataset", "task_type", "model_name", "model_family", "is_foundation_model"])
        self.assertIn("checkpoint_path", header)
        self.assertIn("evidence_path", header)

    def test_manifest_paths_exist(self) -> None:
        for manifest_name in ("figures_manifest.csv", "tables_manifest.csv"):
            with self.outputs[manifest_name].open(newline="", encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
            self.assertTrue(rows)
            for row in rows:
                path = TRACK_DIR.parents[2] / row["path"]
                self.assertTrue(path.exists(), path)

    def test_real_repair_audits_replace_the_false_environment_blocker(self) -> None:
        text = self.outputs["audit_report.md"].read_text(encoding="utf-8")
        self.assertIn("atom-sam2-py310", text)
        self.assertIn("Foundation gain", text)
        self.assertIn("non_beneficial", text)

        summary_paths = [
            TRACK_DIR / "_outputs" / "p10_sam2_repair_audit" / task / "summary.json"
            for task in ("facies_f3", "facies_penobscot")
        ]
        for summary_path in summary_paths:
            self.assertTrue(summary_path.exists(), summary_path)
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
            self.assertEqual(summary["decision"]["state"], "NON_BENEFICIAL")
            self.assertFalse(summary["evaluation"]["frozen_test_accessed"])
            self.assertTrue(summary["model"]["real_pretrained_weights_loaded"])
            comparison = summary["comparison"]
            self.assertGreater(
                comparison["pretrained_adapter_macro_fold_miou"],
                comparison["random_init_control_macro_fold_miou"],
            )
            self.assertLess(
                comparison["gated_residual_repair_macro_fold_miou"],
                comparison["pretrained_adapter_macro_fold_miou"],
            )

        with self.outputs["tables_manifest.csv"].open(newline="", encoding="utf-8") as handle:
            paths = {row["path"] for row in csv.DictReader(handle)}
        self.assertNotIn(
            "_pipelines/02_task_datasets/facies/_outputs/p10_model_results/"
            "p10_sam2_repair_audit/repair_blocker.json",
            paths,
        )
        for summary_path in summary_paths:
            self.assertIn(str(summary_path.relative_to(PROJECT_ROOT)), paths)


if __name__ == "__main__":
    unittest.main()
