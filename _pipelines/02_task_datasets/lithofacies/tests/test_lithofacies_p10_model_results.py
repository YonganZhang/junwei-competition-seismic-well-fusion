from __future__ import annotations

import os
import subprocess
import sys
import tempfile
from pathlib import Path
import unittest


TRACK_DIR = Path(__file__).resolve().parents[1]
WORKTREE_ROOT = TRACK_DIR.parents[2]
WORKTREE_PARENT = WORKTREE_ROOT.parent
SCRIPT = TRACK_DIR / "lithofacies_p10_model_results.py"
STAGE3_RESULTS = TRACK_DIR / "_outputs" / "p5_stage3" / "p5_stage3_results.jsonl"
STAGE3_LEADERBOARD = TRACK_DIR / "_outputs" / "p5_stage3" / "p5_stage3_gm09_p_leaderboard.json"
DEVELOPMENT_BATCH = WORKTREE_PARENT / "p5-stage3-lithofacies" / "_pipelines" / "02_task_datasets" / "lithofacies" / "_outputs" / "p5_stage3" / "runtime" / "development_logo4.npz"
SNAPSHOT = Path(
    "/mnt/data/yongan-admin-2/.cache/huggingface/hub/models--AutonLab--MOMENT-1-base/snapshots/5e44b0ea26376a176360f87831124e018f876d96"
)


def _bootstrap_openpyxl() -> None:
    candidates = [
        Path.home() / ".local" / "lib" / "python3.11" / "site-packages",
        Path.home() / ".local" / "lib" / "python3.10" / "site-packages",
    ]
    for candidate in candidates:
        if (candidate / "openpyxl").exists():
            candidate_str = str(candidate)
            if candidate_str not in sys.path:
                sys.path.append(candidate_str)
            return


class LithofaciesP10ModelResultsTest(unittest.TestCase):
    def test_direct_cli_build_and_verify(self) -> None:
        _bootstrap_openpyxl()
        from openpyxl import load_workbook

        self.assertTrue(SCRIPT.exists())
        self.assertTrue(STAGE3_RESULTS.exists())
        self.assertTrue(STAGE3_LEADERBOARD.exists())
        self.assertTrue(DEVELOPMENT_BATCH.exists())
        self.assertTrue((SNAPSHOT / "model.safetensors").exists())

        with tempfile.TemporaryDirectory() as tmp:
            output_dir = Path(tmp) / "p10_model_results"
            env = os.environ.copy()
            env["CUDA_VISIBLE_DEVICES"] = "5"
            build_cmd = [
                sys.executable,
                str(SCRIPT),
                "build",
                "--development-batch",
                str(DEVELOPMENT_BATCH),
                "--stage3-results",
                str(STAGE3_RESULTS),
                "--stage3-leaderboard",
                str(STAGE3_LEADERBOARD),
                "--snapshot",
                str(SNAPSHOT),
                "--output-dir",
                str(output_dir),
                "--device",
                "cuda:0",
                "--fold-ids",
                "0",
                "--seeds",
                "1867973658",
            ]
            subprocess.run(
                build_cmd,
                check=True,
                cwd=WORKTREE_ROOT,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            verify_cmd = [
                sys.executable,
                str(SCRIPT),
                "verify",
                "--output-dir",
                str(output_dir),
            ]
            subprocess.run(
                verify_cmd,
                check=True,
                cwd=WORKTREE_ROOT,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )

            workbook = output_dir / "track_model_metrics.xlsx"
            wb = load_workbook(workbook, read_only=True)
            try:
                self.assertEqual(wb.sheetnames, ["model_metrics"])
            finally:
                wb.close()

            self.assertTrue((output_dir / "audit_report.md").exists())
            self.assertTrue((output_dir / "before_after_primary_metric.png").exists())


if __name__ == "__main__":
    unittest.main()
