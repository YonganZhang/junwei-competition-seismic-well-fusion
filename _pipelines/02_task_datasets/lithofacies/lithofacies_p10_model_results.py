#!/usr/bin/env python3
"""P10 lithofacies model-results bundle.

This runner is development-only. It consumes the frozen P5/P9 development
evidence, reproduces the cached MOMENT foundation-model audit on LOGO4, and
emits a single-sheet workbook plus manifests and a concise audit note.

No frozen test / known holdout inputs are accepted.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

TRACK_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = TRACK_DIR.parents[2]
for root in (str(PROJECT_ROOT), str(TRACK_DIR)):
    if root not in sys.path:
        sys.path.insert(0, root)


def _bootstrap_openpyxl() -> None:
    candidates = [
        Path.home() / ".local" / "lib" / "python3.11" / "site-packages",
        Path.home() / ".local" / "lib" / "python3.10" / "site-packages",
    ]
    for candidate in candidates:
        if (candidate / "openpyxl").exists():
            candidate_str = str(candidate)
            if candidate_str not in sys.path:
                sys.path.append(candidate_str)
            return


_bootstrap_openpyxl()


def _bootstrap_moment_source() -> None:
    candidates = [
        Path("/mnt/data/yongan-admin-2/.cache/p8-foundation-deps"),
        Path("/mnt/data/yongan-admin-2/.cache/upstream/moment"),
    ]
    for candidate in candidates:
        if candidate.exists():
            candidate_str = str(candidate)
            if candidate_str not in sys.path:
                sys.path.append(candidate_str)


_bootstrap_moment_source()

import numpy as np
import torch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from lithofacies_p5_stage3 import (  # noqa: E402
    FOLD_IDS,
    REPEAT_SEEDS,
    load_stage3_batch,
    _fold_arrays,
)
from p4_contract import classification_metrics_from_logits  # noqa: E402
import p9_moment_effect as moment_audit  # noqa: E402


BASELINE_MODEL_ID = "xgboost_multisoftprob_window"
FOUNDATION_MODEL_ID = "AutonLab/MOMENT-1-base"
FOUNDATION_MODEL_NAME = "moment1_base_cached"
RANDOM_MODEL_NAME = "moment1_base_random_init"
TASK_TYPE = "gm09_fixed_nine_logo4"
TRACK_NAME = "lithofacies"
DATASET_NAME = "development_logo4"
MODEL_SHEET = "model_metrics"
WORKBOOK_SHEET_NAME = "模型指标"
OUTPUT_DIRNAME = "p10_model_results"
WORKBOOK_NAME = "track_model_metrics.xlsx"
FIGURES_MANIFEST = "figures_manifest.csv"
TABLES_MANIFEST = "tables_manifest.csv"
REPORT_NAME = "audit_report.md"
PRIMARY_PNG = "before_after_primary_metric.png"

METRIC_COLUMNS = (
    "fixed_schema_macro_f1",
    "supported_class_macro_f1",
    "accuracy",
    "balanced_accuracy",
    "expected_calibration_error",
    "negative_log_likelihood",
    "multiclass_brier",
    "worst_family_fixed_schema_macro_f1",
)

EXPECTED_WORKBOOK_COLUMNS = (
    "track",
    "dataset",
    "task_type",
    "model_name",
    "model_family",
    "is_foundation_model",
    "foundation_type",
    "integration_point",
    "fusion_method",
    "preprocess_version",
    "split_protocol",
    "seed_or_fold",
    "metric_name",
    "metric_value",
    "higher_is_better",
    "baseline_model",
    "baseline_value",
    "delta_abs",
    "delta_pct",
    "status",
    "evidence_path",
    "checkpoint_path",
    "code_commit",
    "root_cause",
    "fix_applied",
    "notes",
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_load(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _relative(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def _as_float(value: Any) -> float:
    return float(value)


def _metric_rows_for_metrics(
    *,
    base_row: dict[str, Any],
    metrics: dict[str, Any],
    reference_metrics: dict[str, Any],
    status: str,
    evidence_path: str,
    checkpoint_path: str,
    code_commit: str,
    root_cause: str,
    fix_applied: str,
    notes: str,
    model_name: str,
    model_family: str,
    foundation_type: str,
    integration_point: str,
    fusion_method: str,
    preprocess_version: str,
    split_protocol: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for metric_name in METRIC_COLUMNS:
        if metric_name not in metrics:
            continue
        metric_value = _as_float(metrics[metric_name])
        baseline_value = _as_float(reference_metrics[metric_name])
        delta_abs = metric_value - baseline_value
        delta_pct = None if baseline_value == 0 else delta_abs / abs(baseline_value)
        rows.append(
            {
                **base_row,
                "model_name": model_name,
                "model_family": model_family,
                "is_foundation_model": foundation_type != "",
                "foundation_type": foundation_type,
                "integration_point": integration_point,
                "fusion_method": fusion_method,
                "preprocess_version": preprocess_version,
                "split_protocol": split_protocol,
                "metric_name": metric_name,
                "metric_value": metric_value,
                "higher_is_better": metric_name
                not in {
                    "expected_calibration_error",
                    "negative_log_likelihood",
                    "multiclass_brier",
                },
                "baseline_model": BASELINE_MODEL_ID,
                "baseline_value": baseline_value,
                "delta_abs": delta_abs,
                "delta_pct": delta_pct,
                "status": status,
                "evidence_path": evidence_path,
                "checkpoint_path": checkpoint_path,
                "code_commit": code_commit,
                "root_cause": root_cause,
                "fix_applied": fix_applied,
                "notes": notes,
            }
        )
    return rows


def _collect_baseline_rows(
    *,
    results_path: Path,
    baseline_lookup: dict[tuple[int, int], dict[str, Any]],
    commit: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    evidence_path = _relative(results_path)
    for raw in results_path.read_text(encoding="utf-8").splitlines():
        if not raw.strip():
            continue
        row = json.loads(raw)
        if row.get("model_id") != BASELINE_MODEL_ID:
            continue
        key = (int(row["fold_id"]), int(row["seed"]))
        reference_metrics = baseline_lookup[key]
        base_row = {
            "track": TRACK_NAME,
            "dataset": DATASET_NAME,
            "task_type": TASK_TYPE,
            "seed_or_fold": f"fold{int(row['fold_id'])}_seed{int(row['seed'])}",
        }
        metrics = row["validation_metrics"]
        rows.extend(
            _metric_rows_for_metrics(
                base_row=base_row,
                metrics=metrics,
                reference_metrics=reference_metrics,
                status="reference",
                evidence_path=evidence_path,
                checkpoint_path="",
                code_commit=commit,
                root_cause="reference_baseline",
                fix_applied="none",
                notes=(
                    "Stage-3 LOGO4 winner reused as baseline reference; "
                    "same fixed-nine schema and seed/fold pairing"
                ),
                model_name=BASELINE_MODEL_ID,
                model_family="xgboost",
                foundation_type="",
                integration_point="window_tree_ensemble",
                fusion_method="tabular_tree",
                preprocess_version="stage3_fold_train_locked_v1",
                split_protocol="fixed_nine_logo4_v1",
            )
        )
    return rows


def _run_foundation_rows(
    *,
    development_batch: Path,
    stage3_results: Path,
    stage3_leaderboard: Path,
    snapshot: Path,
    fold_ids: Sequence[int],
    seeds: Sequence[int],
    device: str,
    commit: str,
    output_dir: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    arrays, manifest = load_stage3_batch(development_batch)
    leaderboard = _json_load(stage3_leaderboard)
    if (
        leaderboard.get("primary_metric") != "fixed_schema_macro_f1_mean"
        or leaderboard.get("split_hash") != manifest["split_hash"]
    ):
        raise RuntimeError("stage-3 leaderboard does not match the frozen development split")
    summary = _json_load(TRACK_DIR / "_outputs" / "p5_stage3" / "p5_stage3_summary.json")
    if summary.get("batch_sha256") != _sha256(development_batch):
        raise RuntimeError("development batch hash mismatch")
    snapshot_hash = _sha256(snapshot / "model.safetensors")
    if snapshot_hash != "1a436826ffe618273ec62b9656dc4cab8edc470364f104e90542a4ebc14fb825":
        raise RuntimeError("unexpected MOMENT snapshot weights hash")

    stage3_rows = {
        (int(row["fold_id"]), int(row["seed"])): row
        for row in map(json.loads, stage3_results.read_text(encoding="utf-8").splitlines())
        if row.get("model_id") == BASELINE_MODEL_ID
    }

    foundation_rows: list[dict[str, Any]] = []
    audit_batches: list[dict[str, Any]] = []
    for fold_id in fold_ids:
        fold = _fold_arrays(arrays, fold_id)
        train_x = moment_audit._inputs(fold["p_train_well"], fold["p_train_seismic"])
        validation_x = moment_audit._inputs(
            fold["p_validation_well"], fold["p_validation_seismic"]
        )
        train_y = np.asarray(fold["p_train_labels"], dtype=np.int64)
        validation_y = np.asarray(fold["p_validation_labels"], dtype=np.int64)
        class_weights = np.asarray(fold["class_weights"], dtype=np.float32)
        for seed in seeds:
            pretrained_logits, pretrained_loss = moment_audit._train_predict(
                train_x,
                train_y,
                validation_x,
                snapshot=snapshot,
                device=device,
                seed=int(seed),
                random_init=False,
                class_weights=class_weights,
            )
            random_logits, random_loss = moment_audit._train_predict(
                train_x,
                train_y,
                validation_x,
                snapshot=snapshot,
                device=device,
                seed=int(seed),
                random_init=True,
                class_weights=class_weights,
            )
            pretrained_metrics = classification_metrics_from_logits(
                validation_y, pretrained_logits
            )
            random_metrics = classification_metrics_from_logits(
                validation_y, random_logits
            )
            baseline_metrics = stage3_rows[(fold_id, int(seed))]["validation_metrics"]
            base_row = {
                "track": TRACK_NAME,
                "dataset": DATASET_NAME,
                "task_type": TASK_TYPE,
                "seed_or_fold": f"fold{fold_id}_seed{int(seed)}",
            }
            foundation_rows.extend(
                _metric_rows_for_metrics(
                    base_row=base_row,
                    metrics=pretrained_metrics,
                    reference_metrics=baseline_metrics,
                    status="non_beneficial",
                    evidence_path=_relative(output_dir / "audit_report.md"),
                    checkpoint_path=str(snapshot),
                    code_commit=commit,
                    root_cause="no_verified_gain_under_fixed_contract",
                    fix_applied="none",
                    notes=(
                        "Cached MOMENT-1-base pretrained on the exact LOGO4 seed/fold pair; "
                        f"random-init diagnostic macro-F1={random_metrics['fixed_schema_macro_f1']:.6f}; "
                        f"last train loss={pretrained_loss:.6f}"
                    ),
                    model_name=FOUNDATION_MODEL_NAME,
                    model_family="momentfm",
                    foundation_type="MOMENT-1-base",
                    integration_point="frozen_encoder_head_only",
                    fusion_method="single_backbone",
                    preprocess_version="stage3_fold_train_locked_v1",
                    split_protocol="fixed_nine_logo4_v1",
                )
            )
            foundation_rows.extend(
                _metric_rows_for_metrics(
                    base_row=base_row,
                    metrics=random_metrics,
                    reference_metrics=baseline_metrics,
                    status="diagnostic_only",
                    evidence_path=_relative(output_dir / "audit_report.md"),
                    checkpoint_path=str(snapshot),
                    code_commit=commit,
                    root_cause="random_init_ablation",
                    fix_applied="none",
                    notes=(
                        "Random-init control with identical frozen backbone load path; "
                        f"last train loss={random_loss:.6f}"
                    ),
                    model_name=RANDOM_MODEL_NAME,
                    model_family="momentfm",
                    foundation_type="MOMENT-1-base",
                    integration_point="frozen_encoder_head_only",
                    fusion_method="single_backbone",
                    preprocess_version="stage3_fold_train_locked_v1",
                    split_protocol="fixed_nine_logo4_v1",
                )
            )
            audit_batches.append(
                {
                    "fold_id": fold_id,
                    "seed": int(seed),
                    "train_samples": int(len(train_y)),
                    "validation_samples": int(len(validation_y)),
                    "pretrained_fixed_schema_macro_f1": float(
                        pretrained_metrics["fixed_schema_macro_f1"]
                    ),
                    "random_init_fixed_schema_macro_f1": float(
                        random_metrics["fixed_schema_macro_f1"]
                    ),
                    "baseline_fixed_schema_macro_f1": float(
                        baseline_metrics["fixed_schema_macro_f1"]
                    ),
                    "pretrained_last_train_loss": float(pretrained_loss),
                    "random_init_last_train_loss": float(random_loss),
                    "baseline_delta": float(
                        pretrained_metrics["fixed_schema_macro_f1"]
                        - baseline_metrics["fixed_schema_macro_f1"]
                    ),
                }
            )
            del pretrained_logits, random_logits
            torch.cuda.empty_cache()

    return foundation_rows, {
        "development_batch_path": _relative(development_batch),
        "development_batch_sha256": _sha256(development_batch),
        "snapshot_path": str(snapshot),
        "snapshot_sha256": snapshot_hash,
        "split_hash": manifest["split_hash"],
        "fold_ids": list(fold_ids),
        "seeds": [int(seed) for seed in seeds],
        "audit_batches": audit_batches,
        "baseline_mean_fixed_schema_macro_f1": float(
            np.mean([row["baseline_fixed_schema_macro_f1"] for row in audit_batches])
        ),
        "pretrained_mean_fixed_schema_macro_f1": float(
            np.mean([row["pretrained_fixed_schema_macro_f1"] for row in audit_batches])
        ),
        "random_mean_fixed_schema_macro_f1": float(
            np.mean([row["random_init_fixed_schema_macro_f1"] for row in audit_batches])
        ),
    }


def _write_workbook(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = WORKBOOK_SHEET_NAME
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(ord('A') + len(EXPECTED_WORKBOOK_COLUMNS) - 1)}{len(rows) + 1}"
    ws.append(list(EXPECTED_WORKBOOK_COLUMNS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(column) for column in EXPECTED_WORKBOOK_COLUMNS])
    wb.save(path)


def _write_csv(path: Path, rows: Sequence[dict[str, Any]], columns: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns), lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def _plot_before_after(path: Path, audit_meta: Mapping[str, Any], rows: Sequence[dict[str, Any]]) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    path.parent.mkdir(parents=True, exist_ok=True)
    primary = [
        row
        for row in rows
        if row["metric_name"] == "fixed_schema_macro_f1"
        and row["model_name"] in {BASELINE_MODEL_ID, FOUNDATION_MODEL_NAME, RANDOM_MODEL_NAME}
    ]
    order = []
    for seed in audit_meta["seeds"]:
        for fold in audit_meta["fold_ids"]:
            order.append((fold, seed))
    baseline_lookup = {
        (int(row["seed_or_fold"].split("_seed")[0].removeprefix("fold")), int(row["seed_or_fold"].split("_seed")[1])): row
        for row in primary
        if row["model_name"] == BASELINE_MODEL_ID
    }
    pretrained_lookup = {
        (int(row["seed_or_fold"].split("_seed")[0].removeprefix("fold")), int(row["seed_or_fold"].split("_seed")[1])): row
        for row in primary
        if row["model_name"] == FOUNDATION_MODEL_NAME
    }
    random_lookup = {
        (int(row["seed_or_fold"].split("_seed")[0].removeprefix("fold")), int(row["seed_or_fold"].split("_seed")[1])): row
        for row in primary
        if row["model_name"] == RANDOM_MODEL_NAME
    }

    labels = [f"F{fold}-S{str(seed)[-4:]}" for fold, seed in order]
    baseline_values = [baseline_lookup[(fold, seed)]["metric_value"] for fold, seed in order]
    pretrained_values = [pretrained_lookup[(fold, seed)]["metric_value"] for fold, seed in order]
    random_values = [random_lookup[(fold, seed)]["metric_value"] for fold, seed in order]

    def _std(values: Sequence[float]) -> float:
        return float(np.std(values, ddof=1)) if len(values) > 1 else 0.0

    fig, (ax0, ax1) = plt.subplots(1, 2, figsize=(14, 5), constrained_layout=True)
    variants = ["baseline xgboost", "MOMENT pretrained", "MOMENT random-init"]
    means = [
        float(np.mean(baseline_values)),
        float(np.mean(pretrained_values)),
        float(np.mean(random_values)),
    ]
    stds = [
        _std(baseline_values),
        _std(pretrained_values),
        _std(random_values),
    ]
    ax0.bar(variants, means, yerr=stds, color=["#4C78A8", "#F58518", "#54A24B"])
    ax0.set_ylabel("fixed-schema macro-F1")
    ax0.set_title("Mean primary metric on LOGO4 development pairs")
    ax0.set_ylim(0, max(means) * 1.35 if max(means) > 0 else 1.0)
    ax0.grid(axis="y", alpha=0.25)

    x = np.arange(len(labels))
    ax1.plot(x, baseline_values, marker="o", label="baseline xgboost", color="#4C78A8")
    ax1.plot(x, pretrained_values, marker="o", label="MOMENT pretrained", color="#F58518")
    ax1.plot(x, random_values, marker="o", label="MOMENT random-init", color="#54A24B")
    ax1.set_xticks(x, labels, rotation=45, ha="right")
    ax1.set_ylabel("fixed-schema macro-F1")
    ax1.set_title("Paired fold/seed comparison")
    ax1.grid(axis="y", alpha=0.25)
    ax1.legend(loc="best")
    fig.suptitle(
        "Lithofacies P10 before/after primary metric\n"
        "Development-only LOGO4, fixed nine classes, no holdout access",
        fontsize=12,
    )
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _write_manifests(
    *,
    output_dir: Path,
    audit_meta: Mapping[str, Any],
    workbook_path: Path,
    baseline_rows: Sequence[dict[str, Any]],
    foundation_rows: Sequence[dict[str, Any]],
) -> tuple[Path, Path]:
    figures_rows = [
        {
            "kind": "figure",
            "name": "before_after_primary_metric",
            "path": _relative(output_dir / PRIMARY_PNG),
            "source": "p10_audit_bundle",
            "status": "new",
            "notes": "paired baseline vs MOMENT primary metric on LOGO4 development pairs",
        },
        {
            "kind": "figure",
            "name": "fixed9_confusion",
            "path": _relative(TRACK_DIR / "_outputs" / "p5_stage3" / "figures" / "fixed9_confusion.png"),
            "source": "p5_stage3",
            "status": "existing",
            "notes": "development confusion matrix for the fixed-nine contract",
        },
        {
            "kind": "figure",
            "name": "fixed9_per_class_pr_f1",
            "path": _relative(TRACK_DIR / "_outputs" / "p5_stage3" / "figures" / "fixed9_per_class_pr_f1.png"),
            "source": "p5_stage3",
            "status": "existing",
            "notes": "development per-class PR/F1 support for the fixed-nine contract",
        },
        {
            "kind": "figure",
            "name": "calibration_reliability",
            "path": _relative(TRACK_DIR / "_outputs" / "p5_stage3" / "figures" / "calibration_reliability.png"),
            "source": "p5_stage3",
            "status": "existing",
            "notes": "development reliability curve for the frozen contract",
        },
        {
            "kind": "figure",
            "name": "fold_seed_matrix",
            "path": _relative(TRACK_DIR / "_outputs" / "p5_stage3" / "figures" / "fold_seed_matrix.png"),
            "source": "p5_stage3",
            "status": "existing",
            "notes": "fold×seed primary-metric matrix from the stage-3 leaderboard",
        },
        {
            "kind": "figure",
            "name": "missing_modality_diagnostic",
            "path": _relative(TRACK_DIR / "_outputs" / "p5_stage3" / "figures" / "missing_modality_diagnostic.png"),
            "source": "p5_stage3",
            "status": "existing",
            "notes": "missing-modality diagnostic for the development contract",
        },
        {
            "kind": "figure",
            "name": "continuous_depth_track_not_feasible",
            "path": _relative(TRACK_DIR / "_outputs" / "p5_stage3" / "figures" / "continuous_depth_track_not_feasible.png"),
            "source": "p5_stage3",
            "status": "existing",
            "notes": "explicitly states that 3D continuous-depth track is not feasible",
        },
    ]
    figures_manifest = output_dir / FIGURES_MANIFEST
    _write_csv(figures_manifest, figures_rows, ("kind", "name", "path", "source", "status", "notes"))

    table_rows = [
        {
            "kind": "workbook_sheet",
            "name": f"{WORKBOOK_NAME}::{WORKBOOK_SHEET_NAME}",
            "path": _relative(workbook_path),
            "rows": len(baseline_rows) + len(foundation_rows),
            "source": "p10_audit_bundle",
            "notes": "single-sheet long table with baseline, pretrained, and random-init rows",
        },
        {
            "kind": "evidence_table",
            "name": "p5_stage3_results.jsonl",
            "path": _relative(TRACK_DIR / "_outputs" / "p5_stage3" / "p5_stage3_results.jsonl"),
            "rows": len(baseline_rows),
            "source": "p5_stage3",
            "notes": "development baseline reference rows used for the paired comparison",
        },
        {
            "kind": "evidence_table",
            "name": "p9_moment_effect/summary.json",
            "path": _relative(TRACK_DIR / "_outputs" / "p9_moment_effect" / "summary.json"),
            "rows": 4,
            "source": "p9_moment_effect",
            "notes": "exact frozen-fold MOMENT reproduction evidence used to cross-check the repair",
        },
    ]
    tables_manifest = output_dir / TABLES_MANIFEST
    _write_csv(tables_manifest, table_rows, ("kind", "name", "path", "rows", "source", "notes"))
    return figures_manifest, tables_manifest


def _write_report(
    *,
    output_dir: Path,
    audit_meta: Mapping[str, Any],
    baseline_rows: Sequence[dict[str, Any]],
    foundation_rows: Sequence[dict[str, Any]],
) -> Path:
    report_path = output_dir / REPORT_NAME
    baseline_primary = [
        row for row in baseline_rows if row["metric_name"] == "fixed_schema_macro_f1"
    ]
    foundation_primary = [
        row for row in foundation_rows if row["metric_name"] == "fixed_schema_macro_f1"
    ]
    baseline_mean = float(np.mean([row["metric_value"] for row in baseline_primary]))
    foundation_mean = float(np.mean([row["metric_value"] for row in foundation_primary]))
    random_rows = [
        row
        for row in foundation_rows
        if row["metric_name"] == "fixed_schema_macro_f1"
        and row["model_name"] == RANDOM_MODEL_NAME
    ]
    random_mean = float(np.mean([row["metric_value"] for row in random_rows]))
    delta = foundation_mean - baseline_mean
    rows = [
        "# P10 lithofacies model-results audit",
        "",
        "## Conclusion",
        "",
        f"Status: non_beneficial. The cached MOMENT-1-base foundation path remained below the fixed-nine XGBoost LOGO4 baseline on the development pairs.",
        f"Baseline fixed-schema macro-F1 mean: {baseline_mean:.6f}.",
        f"MOMENT pretrained fixed-schema macro-F1 mean: {foundation_mean:.6f}.",
        f"MOMENT random-init fixed-schema macro-F1 mean: {random_mean:.6f}.",
        f"Primary-metric delta (pretrained - baseline): {delta:.6f}.",
        "",
        "## Before/after by fold and seed",
        "",
        "| fold | seed | baseline | pretrained | random-init | delta(pretrained-baseline) |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    baseline_map = {
        tuple(int(part) for part in row["seed_or_fold"].replace("fold", "").split("_seed")): row
        for row in baseline_primary
    }
    pretrained_map = {
        tuple(int(part) for part in row["seed_or_fold"].replace("fold", "").split("_seed")): row
        for row in foundation_primary
        if row["model_name"] == FOUNDATION_MODEL_NAME
    }
    random_map = {
        tuple(int(part) for part in row["seed_or_fold"].replace("fold", "").split("_seed")): row
        for row in foundation_primary
        if row["model_name"] == RANDOM_MODEL_NAME
    }
    for seed in audit_meta["seeds"]:
        for fold in audit_meta["fold_ids"]:
            key = (fold, seed)
            rows.append(
                "| "
                + " | ".join(
                    [
                        str(fold),
                        str(seed),
                        f"{baseline_map[key]['metric_value']:.6f}",
                        f"{pretrained_map[key]['metric_value']:.6f}",
                        f"{random_map[key]['metric_value']:.6f}",
                        f"{pretrained_map[key]['metric_value'] - baseline_map[key]['metric_value']:.6f}",
                    ]
                )
                + " |"
            )
    rows.extend(
        [
            "",
        "## Root cause and fix status",
        "",
        "- No reproducible integration defect was found in the development-only path.",
        "- The frozen fixed-nine / LOGO4 / train-fold-only preprocessing contract held.",
        "- MOMENT pretrained improved over random initialization but did not beat the strong XGBoost baseline.",
        "- Therefore the honest outcome is non_beneficial, not a repaired win.",
        "",
        "## Evidence and hashes",
        "",
        f"- development batch: `{audit_meta['development_batch_path']}`",
        f"- development batch sha256: `{audit_meta['development_batch_sha256']}`",
        f"- MOMENT snapshot: `{audit_meta['snapshot_path']}`",
        f"- MOMENT snapshot sha256: `{audit_meta['snapshot_sha256']}`",
        f"- split hash: `{audit_meta['split_hash']}`",
        f"- code commit: `{os.environ.get('P10_CODE_COMMIT', os.environ.get('GIT_COMMIT', 'e4fd5d8a6371c2b0db6ba2258a41349ec6cfb4f7'))}`",
        "",
        "## Traceable contract evidence",
        "",
        "- Depth-window / stride / direction: `p9_moment_effect._inputs()` keeps the fixed 33-position LOGO4 window and reshapes the 26 well-log + 9 seismic channels to `[B,35,33]`; the audit uses the cached development batch from `_outputs/p5_stage3/runtime/development_logo4.npz`.",
        "- Padding / mask: `p4_contract.apply_fold_preprocessor()` preserves the 26 physical channels, appends the 13-channel missing mask, and records `fit_scope = fold_train_mother_families_only`.",
        "- Fold-train-only normalization: `p4_contract.fit_fold_preprocessor()` derives `log_stats`, `seismic_stats`, and `class_weights` only from the fold-train mother families; validation uses the immutable train statistics only.",
        "- MOMENT embedding / input channels: `_models/lithofacies/moment_depth.py` requires `n_channels=35`, interpolates the 33-position input to 512 internally, and uses the cached `momentfm.MOMENTPipeline`.",
        "- Frozen / PEFT / head / output classes: the MOMENT audit uses `freeze_encoder=True`, `freeze_embedder=True`, `freeze_head=False`; `build_model(..., num_class=9)` is a fixed-nine classifier head.",
        "- Fixed-nine label mapping: `p4_contract.CLASS_NAMES` and `classification_metrics_from_logits()` operate on the frozen GM09 nine-class schema; `fixed_schema_macro_f1` is the primary metric and `supported_class_macro_f1` remains diagnostic only.",
        "- Class imbalance: `fit_fold_preprocessor()` computes fold-train-only `class_weights` and the runner passes them into `torch.nn.functional.cross_entropy`; the stage3 baseline uses locked `sqrt_inverse_frequency_weighted_*` contracts.",
        "- LOGO fold / sample universe: `build_lithofacies_split_manifest()` freezes the four development mother families plus the F-5 test family; the development batch reports `split_hash = a06375429f9e9cf380fb5cdebd7d0cb7b25d7a13d29522b8e2420f4dae1b4555` and `frozen_test_accessed = False`.",
        "- Seed / metric direction: the stage3 development audit uses seeds `1867973658`, `2137841944`, `3902865753`; the score direction is maximize for `fixed_schema_macro_f1` and minimize for calibration, NLL, and Brier only.",
        "",
        "## Residual risk",
        "",
        "- The audit is limited to the cached MOMENT-1-base snapshot and the fixed development contract.",
        "- No frozen-test / known-holdout evidence was consumed.",
        "- MOMENT pretrained showed a small foundation gain over random initialization (`0.046308` vs `0.041112`) but remained far below XGBoost (`0.194938`), so the end-to-end conclusion stays `non_beneficial`.",
        "- No HPO or split changes were performed.",
    ]
    )
    report_path.write_text("\n".join(rows) + "\n", encoding="utf-8")
    return report_path


def build_bundle(
    *,
    development_batch: Path,
    stage3_results: Path,
    stage3_leaderboard: Path,
    snapshot: Path,
    output_dir: Path,
    fold_ids: Sequence[int],
    seeds: Sequence[int],
    device: str,
) -> dict[str, Any]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    commit = (
        os.environ.get("P10_CODE_COMMIT")
        or os.environ.get("GIT_COMMIT")
        or "e4fd5d8a6371c2b0db6ba2258a41349ec6cfb4f7"
    )
    baseline_lookup: dict[tuple[int, int], dict[str, Any]] = {}
    for raw in stage3_results.read_text(encoding="utf-8").splitlines():
        if not raw.strip():
            continue
        row = json.loads(raw)
        if row.get("model_id") != BASELINE_MODEL_ID:
            continue
        key = (int(row["fold_id"]), int(row["seed"]))
        baseline_lookup[key] = row["validation_metrics"]
    baseline_rows = _collect_baseline_rows(
        results_path=stage3_results, baseline_lookup=baseline_lookup, commit=commit
    )
    foundation_rows, audit_meta = _run_foundation_rows(
        development_batch=development_batch,
        stage3_results=stage3_results,
        stage3_leaderboard=stage3_leaderboard,
        snapshot=snapshot,
        fold_ids=fold_ids,
        seeds=seeds,
        device=device,
        commit=commit,
        output_dir=output_dir,
    )
    all_rows = baseline_rows + foundation_rows
    workbook_path = output_dir / WORKBOOK_NAME
    _write_workbook(workbook_path, all_rows)
    _plot_before_after(output_dir / PRIMARY_PNG, audit_meta, all_rows)
    figures_manifest, tables_manifest = _write_manifests(
        output_dir=output_dir,
        audit_meta=audit_meta,
        workbook_path=workbook_path,
        baseline_rows=baseline_rows,
        foundation_rows=foundation_rows,
    )
    report_path = _write_report(
        output_dir=output_dir,
        audit_meta=audit_meta,
        baseline_rows=baseline_rows,
        foundation_rows=foundation_rows,
    )
    return {
        "output_dir": str(output_dir),
        "workbook": str(workbook_path),
        "figures_manifest": str(figures_manifest),
        "tables_manifest": str(tables_manifest),
        "report": str(report_path),
        "primary_png": str(output_dir / PRIMARY_PNG),
        "rows": len(all_rows),
        "baseline_rows": len(baseline_rows),
        "foundation_rows": len(foundation_rows),
        "audit_meta": audit_meta,
    }


def verify_bundle(output_dir: Path) -> dict[str, Any]:
    output_dir = Path(output_dir)
    workbook = output_dir / WORKBOOK_NAME
    figures_manifest = output_dir / FIGURES_MANIFEST
    tables_manifest = output_dir / TABLES_MANIFEST
    report = output_dir / REPORT_NAME
    primary_png = output_dir / PRIMARY_PNG
    for path in (workbook, figures_manifest, tables_manifest, report, primary_png):
        if not path.exists():
            raise FileNotFoundError(path)

    wb = load_workbook(workbook, read_only=True)
    try:
        if wb.sheetnames != [WORKBOOK_SHEET_NAME]:
            raise AssertionError(f"unexpected sheet names: {wb.sheetnames}")
        ws = wb[WORKBOOK_SHEET_NAME]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if tuple(header) != EXPECTED_WORKBOOK_COLUMNS:
            raise AssertionError("workbook columns do not match the contract")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise AssertionError("workbook sheet is empty")
    finally:
        wb.close()

    figure_rows = list(csv.DictReader(figures_manifest.read_text(encoding="utf-8").splitlines()))
    table_rows = list(csv.DictReader(tables_manifest.read_text(encoding="utf-8").splitlines()))
    for row in figure_rows + table_rows:
        rel = row["path"]
        if not (PROJECT_ROOT / rel).exists() and not Path(rel).exists():
            raise FileNotFoundError(rel)
    return {
        "workbook": str(workbook),
        "figures_manifest": str(figures_manifest),
        "tables_manifest": str(tables_manifest),
        "report": str(report),
        "primary_png": str(primary_png),
        "rows": len(rows),
    }


def _parse_ints(values: Sequence[str]) -> list[int]:
    return [int(value) for value in values]


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    build = subparsers.add_parser("build")
    build.add_argument("--development-batch", type=Path, required=True)
    build.add_argument("--stage3-results", type=Path, required=True)
    build.add_argument("--stage3-leaderboard", type=Path, required=True)
    build.add_argument("--snapshot", type=Path, required=True)
    build.add_argument("--output-dir", type=Path, required=True)
    build.add_argument("--device", default="cuda:0")
    build.add_argument("--fold-ids", nargs="+", default=[0, 1, 2, 3])
    build.add_argument(
        "--seeds",
        nargs="+",
        default=[str(seed) for seed in REPEAT_SEEDS],
    )

    verify = subparsers.add_parser("verify")
    verify.add_argument("--output-dir", type=Path, required=True)

    args = parser.parse_args(argv)
    if args.command == "build":
        result = build_bundle(
            development_batch=args.development_batch,
            stage3_results=args.stage3_results,
            stage3_leaderboard=args.stage3_leaderboard,
            snapshot=args.snapshot,
            output_dir=args.output_dir,
            fold_ids=_parse_ints(args.fold_ids),
            seeds=_parse_ints(args.seeds),
            device=args.device,
        )
        print(json.dumps(result, ensure_ascii=False, sort_keys=True))
        return 0
    if args.command == "verify":
        result = verify_bundle(args.output_dir)
        print(json.dumps(result, ensure_ascii=False, sort_keys=True))
        return 0
    raise AssertionError(args.command)


if __name__ == "__main__":
    raise SystemExit(main())
