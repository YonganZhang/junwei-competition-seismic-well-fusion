from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook


HERE = Path(__file__).resolve().parents[1]
OUTPUT_DIR = HERE / "_outputs" / "p10_model_results"


@pytest.mark.integration
def test_p10_artifacts_are_single_sheet_and_evidence_backed() -> None:
    required = [
        OUTPUT_DIR / "track_model_metrics.xlsx",
        OUTPUT_DIR / "figures_manifest.csv",
        OUTPUT_DIR / "tables_manifest.csv",
        OUTPUT_DIR / "audit_report.md",
        OUTPUT_DIR / "before_after_primary_metric.png",
        OUTPUT_DIR / "tabicl_reproduction.jsonl",
    ]
    missing = [path.name for path in required if not path.is_file()]
    if missing:
        pytest.skip("P10 artifacts are absent: " + ", ".join(missing))

    workbook = load_workbook(OUTPUT_DIR / "track_model_metrics.xlsx", read_only=True)
    assert workbook.sheetnames == ["模型指标"]
    sheet = workbook["模型指标"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header[:4] == ["track", "dataset", "task_type", "model_name"]
    assert "evidence_path" in header
    assert "checkpoint_path" in header
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows
    evidence_idx = header.index("evidence_path")
    status_idx = header.index("status")
    assert all(Path(str(row[evidence_idx])).is_file() for row in rows)
    assert any("tabiclv2_regressor" in str(row[header.index("model_name")]) for row in rows)
    statuses = {str(row[status_idx]) for row in rows}
    assert "effect_supported_not_promoted" in statuses
    assert "evidence_only" in statuses
    assert "control" in statuses
    assert "data_blocked" not in statuses
    assert any("p9_tabicl_effect/summary.json" in str(row[evidence_idx]) for row in rows)

    index = {name: idx for idx, name in enumerate(header)}
    p9_rows = [
        row
        for row in rows
        if row[index["model_name"]] == "tabiclv2_regressor"
        and "p9_tabicl_effect/summary.json" in str(row[evidence_idx])
    ]
    for row in p9_rows:
        if row[index["seed_or_fold"]] == "macro":
            continue
        assert row[index["status"]] in {"evidence_only", "control"}
        assert row[index["baseline_value"]] is None
        assert row[index["delta_abs"]] is None
        assert row[index["delta_pct"]] is None

    expected_rmse_gain = {
        "PHIF_RMSE_physical": 0.048948485749996995,
        "KLOGH_RMSE_physical": 0.13565572419478755,
        "SW_RMSE_physical": 0.25185017996367165,
    }
    for metric_name, expected in expected_rmse_gain.items():
        matches = [
            row
            for row in p9_rows
            if row[index["seed_or_fold"]] == "macro"
            and row[index["metric_name"]] == metric_name
            and row[index["status"]] == "effect_supported_not_promoted"
        ]
        assert len(matches) == 1
        assert float(matches[0][index["delta_pct"]]) == pytest.approx(expected)


@pytest.mark.integration
def test_p10_manifests_index_existing_valid_tables_and_charts() -> None:
    figures = list(csv.DictReader((OUTPUT_DIR / "figures_manifest.csv").open(encoding="utf-8")))
    tables = list(csv.DictReader((OUTPUT_DIR / "tables_manifest.csv").open(encoding="utf-8")))
    assert figures
    assert tables
    assert any(row["figure_path"].endswith("before_after_primary_metric.png") for row in figures)
    assert any(row["table_path"].endswith("track_model_metrics.xlsx") for row in tables)
    assert all(Path(row["evidence_path"]).is_file() for row in tables)
    assert all(Path(row["figure_path"]).is_file() for row in figures)


@pytest.mark.integration
def test_tabicl_reproduction_jsonl_uses_only_local_cache_hash() -> None:
    path = OUTPUT_DIR / "tabicl_reproduction.jsonl"
    if not path.is_file():
        pytest.skip("P10 TabICL reproduction evidence absent")
    first = json.loads(path.read_text(encoding="utf-8").splitlines()[0])
    assert first["track"] == "property"
    assert first["model_name"] == "tabiclv2_regressor"
    assert first["checkpoint_path"] == "artifact_unavailable"
    assert len(first["checkpoint_sha256"]) == 64
    assert first["status"] == "beneficial"
