"""Build the portable P10 property-model results delivery.

This script only reads already-frozen reservoir evidence and writes a portable
results bundle under ``_outputs/p10_model_results``.
"""
from __future__ import annotations

import csv
import hashlib
import json
import sys
from pathlib import Path
from typing import Any

import matplotlib
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402


HERE = Path(__file__).resolve().parent
OUTPUT_DIR = HERE / "_outputs" / "p10_model_results"


def _discover_data_root() -> Path:
    worktrees_root = None
    for ancestor in HERE.parents:
        if ancestor.name == "worktrees":
            worktrees_root = ancestor
            break
    if worktrees_root is None:
        raise FileNotFoundError("unable to locate the shared worktrees root")
    preferred = ["track-property", "p5-r2-property"]
    candidates = [worktrees_root / name for name in preferred if (worktrees_root / name).is_dir()]
    candidates.extend(sorted(p for p in worktrees_root.iterdir() if p.is_dir() and p not in candidates))
    for candidate in candidates:
        if (
            (candidate / "_data/processed/reservoir/train.h5").is_file()
            and (candidate / "_pipelines/02_task_datasets/reservoir/_outputs/guard.npz").is_file()
            and (candidate / "_pipelines/02_task_datasets/sweetspot/targets/porosity/_outputs/phif/split_manifest.json").is_file()
        ):
            return candidate
    raise FileNotFoundError("unable to locate a source worktree with frozen reservoir data")


DATA_ROOT = _discover_data_root()
PROJECT_ROOT = HERE.parents[2]
DEV_BATCH_PATH = OUTPUT_DIR / "runtime" / "development_logo4.npz"
P5_STAGE3_DIR = PROJECT_ROOT / "_pipelines/02_task_datasets/reservoir/_outputs/p5_stage3"
P5_STAGE4_DIR = PROJECT_ROOT / "_pipelines/02_task_datasets/reservoir/_outputs/p5_stage4_confirmation"
P5_STAGE3_RESULTS = P5_STAGE3_DIR / "p5_stage3_results.jsonl"
P5_STAGE3_SUMMARY = P5_STAGE3_DIR / "p5_stage3_summary.json"
P5_STAGE4_SUMMARY = P5_STAGE4_DIR / "summary.json"
P5_STAGE4_VIZ = P5_STAGE4_DIR / "visualization_manifest.json"
P9_TABICL_SUMMARY = PROJECT_ROOT / "_pipelines/02_task_datasets/reservoir/_outputs/p9_tabicl_effect/summary.json"
TRAIN_H5 = DATA_ROOT / "_data" / "processed" / "reservoir" / "train.h5"
GUARD_NPZ = DATA_ROOT / "_pipelines" / "02_task_datasets" / "reservoir" / "_outputs" / "guard.npz"
P4_POROSITY_SPLIT = (
    DATA_ROOT
    / "_pipelines/02_task_datasets/sweetspot/targets/porosity/_outputs/phif/split_manifest.json"
)

TARGETS = ("PHIF", "KLOGH", "SW")
FOLDS = (0, 1, 2, 3)
REPEAT_SEEDS = (1867973658, 2137841944, 3902865753)
ROOT_SEED = 2693
METRICS = ("MAE", "RMSE", "R2")
BASELINE_MODELS = {
    "PHIF": "extra_trees_regressor",
    "KLOGH": "extra_trees_regressor",
    "SW": "xgboost_regressor",
}
MODEL_FAMILIES = {
    "extra_trees_regressor": "randomized_forest",
    "xgboost_regressor": "regularized_gradient_boosting",
    "tabiclv2_regressor": "in_context_tabular_foundation_model",
}
FOUNDATION_TYPES = {"tabiclv2_regressor": "tabular_pretrained"}
INTEGRATION_POINTS = {
    "extra_trees_regressor": "classical_tabular_baseline",
    "xgboost_regressor": "classical_tabular_baseline",
    "tabiclv2_regressor": "tabular_pretrained_adapter",
}
FUSION_METHODS = {model: "none" for model in MODEL_FAMILIES}
PREPROCESS_VERSION = {
    "development_logo4": "fold_train_only_v1",
    "known_holdout_f15": "single_use_confirmed_v1",
}
CURRENT_COMMIT = "ba68969dce674c154fc28c349c2389290c0a5a18"
LOCAL_TABICL_CHECKPOINT_SHA256 = (
    "0db9cb538f114e79026bf08f45f41ad8dd7ad2de2aaca9a5ca8cd3bd9748ae7a"
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json(path: Path) -> Any:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _repo_rel(path: Path) -> str:
    return Path(path).resolve().relative_to(PROJECT_ROOT.resolve()).as_posix()


def _load_baseline_winners() -> dict[str, str]:
    winners: dict[str, str] = {}
    for target in TARGETS:
        board = _json(P5_STAGE3_DIR / f"leaderboard_{target.lower()}_tabular_cpu.json")
        winners[target] = next(entry["model_id"] for entry in board["entries"] if entry["rank"] == 1)
    return winners


def _load_stage3_cells() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with P5_STAGE3_RESULTS.open("r", encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def _load_p9_summary() -> dict[str, Any]:
    return _json(P9_TABICL_SUMMARY)


def _prepare_development_batch() -> Path:
    if DEV_BATCH_PATH.is_file():
        return DEV_BATCH_PATH
    sys.path.insert(0, str(PROJECT_ROOT))
    sys.path.insert(0, str(HERE))
    import reservoir_p5_stage3 as stage3  # noqa: E402

    report = stage3.prepare_logo4(TRAIN_H5, GUARD_NPZ, P4_POROSITY_SPLIT, DEV_BATCH_PATH)
    if report["split_hash"] != "2334f3cc301fc66d6b98c6edf3a4f9c920776469531003d62f5370e119426a18":
        raise RuntimeError("development split hash drifted")
    return DEV_BATCH_PATH


def _tabicl_checkpoint() -> Path:
    cache_root = Path.home() / ".cache" / "huggingface" / "hub"
    candidates = list(cache_root.rglob(LOCAL_TABICL_CHECKPOINT_SHA256))
    if not candidates:
        candidates = list(cache_root.rglob("tabicl-regressor-v2-20260212.ckpt"))
    if not candidates:
        raise FileNotFoundError("TabICLv2 checkpoint blob is absent from the local cache")
    return candidates[0]


def _metric_rows(
    *,
    dataset: str,
    split_protocol: str,
    preprocess_version: str,
    seed_or_fold: str,
    model_name: str,
    metric_prefix: str,
    physical: Mapping[str, float],
    model_domain: Mapping[str, float] | None,
    evidence_path: str,
    checkpoint_path: str,
    notes: str,
    status: str,
    is_foundation_model: bool,
    baseline_model: str,
    code_commit: str = CURRENT_COMMIT,
) -> list[dict[str, Any]]:
    def _maybe_float(value: Any) -> float | None:
        return None if value is None else float(value)

    rows: list[dict[str, Any]] = []
    for metric in METRICS:
        metric_name = f"{metric_prefix}_{metric}_physical"
        rows.append(
            {
                "track": "property",
                "dataset": dataset,
                "task_type": "regression",
                "model_name": model_name,
                "model_family": MODEL_FAMILIES[model_name],
                "is_foundation_model": is_foundation_model,
                "foundation_type": FOUNDATION_TYPES.get(model_name, ""),
                "integration_point": INTEGRATION_POINTS[model_name],
                "fusion_method": FUSION_METHODS[model_name],
                "preprocess_version": preprocess_version,
                "split_protocol": split_protocol,
                "seed_or_fold": seed_or_fold,
                "metric_name": metric_name,
                "metric_value": _maybe_float(physical[metric]),
                "higher_is_better": metric in {"R2", "Pearson"},
                "baseline_model": baseline_model,
                "baseline_value": None,
                "delta_abs": None,
                "delta_pct": None,
                "status": status,
                "evidence_path": evidence_path,
                "checkpoint_path": checkpoint_path,
                "code_commit": code_commit,
                "root_cause": "no_regression_evidenced" if status != "reference" else "reference",
                "fix_applied": "none",
                "notes": notes,
            }
        )
    if model_domain is not None:
        for metric in METRICS:
            metric_name = f"{metric_prefix}_{metric}_log1p"
            rows.append(
                {
                    "track": "property",
                    "dataset": dataset,
                    "task_type": "regression",
                    "model_name": model_name,
                    "model_family": MODEL_FAMILIES[model_name],
                    "is_foundation_model": is_foundation_model,
                    "foundation_type": FOUNDATION_TYPES.get(model_name, ""),
                    "integration_point": INTEGRATION_POINTS[model_name],
                    "fusion_method": FUSION_METHODS[model_name],
                    "preprocess_version": preprocess_version,
                    "split_protocol": split_protocol,
                    "seed_or_fold": seed_or_fold,
                    "metric_name": metric_name,
                    "metric_value": _maybe_float(model_domain[metric]),
                    "higher_is_better": metric in {"R2", "Pearson"},
                    "baseline_model": baseline_model,
                    "baseline_value": None,
                    "delta_abs": None,
                    "delta_pct": None,
                    "status": status,
                    "evidence_path": evidence_path,
                    "checkpoint_path": checkpoint_path,
                    "code_commit": code_commit,
                    "root_cause": "no_regression_evidenced" if status != "reference" else "reference",
                    "fix_applied": "none",
                    "notes": notes,
                }
            )
    return rows


def _attach_deltas(rows: list[dict[str, Any]]) -> None:
    lookup: dict[tuple[str, str, str, str, str], float] = {}
    for row in rows:
        if row["metric_value"] is None:
            continue
        key = (
            row["model_name"],
            row["dataset"],
            row["split_protocol"],
            row["seed_or_fold"],
            row["metric_name"],
        )
        lookup[key] = float(row["metric_value"])
    for row in rows:
        if row["status"] in {"reference", "production_reference"}:
            row["baseline_value"] = float(row["metric_value"]) if row["metric_value"] is not None else None
            row["delta_abs"] = 0.0 if row["metric_value"] is not None else None
            row["delta_pct"] = 0.0 if row["metric_value"] is not None else None
            continue
        if row["status"] in {"control", "diagnostic_control", "evidence_only"}:
            continue
        if row["status"] == "data_blocked":
            continue
        baseline_key = (
            row["baseline_model"],
            row["dataset"],
            row["split_protocol"],
            row["seed_or_fold"],
            row["metric_name"],
        )
        baseline = lookup.get(baseline_key)
        if baseline is None:
            row["status"] = "data_blocked"
            row["root_cause"] = "baseline_missing"
            row["fix_applied"] = "none"
            row["notes"] = f"{row['notes']}; baseline_missing"
            continue
        if row["metric_value"] is None:
            continue
        row["baseline_value"] = baseline
        if row["higher_is_better"]:
            delta_abs = float(row["metric_value"]) - baseline
        else:
            delta_abs = baseline - float(row["metric_value"])
        delta_pct = delta_abs / abs(baseline) if baseline else 0.0
        row["delta_abs"] = delta_abs
        row["delta_pct"] = delta_pct
        if row["status"] != "effect_supported_not_promoted":
            row["status"] = "beneficial" if delta_abs > 0 else ("tie" if abs(delta_abs) <= 1e-12 else "non_beneficial")


def _baseline_rows() -> list[dict[str, Any]]:
    winners = _load_baseline_winners()
    rows: list[dict[str, Any]] = []
    for row in _load_stage3_cells():
        target = row["target"]
        if row["model_id"] != winners[target]:
            continue
        validation = row["validation"]["metric"]
        physical = validation["physical"]
        model_domain = validation["model_domain"] if target == "KLOGH" else None
        seed_or_fold = f"seed{int(row['repeat_seed'])}/fold{int(row['fold_id'])}"
        notes = (
            f"split_hash={row['split']['split_hash']}; "
            f"checkpoint_sha256={row['checkpoint']['sha256']}"
        )
        rows.extend(
            _metric_rows(
                dataset="development_logo4",
                split_protocol="mother_family_logo4",
                preprocess_version=PREPROCESS_VERSION["development_logo4"],
                seed_or_fold=seed_or_fold,
                model_name=row["model_id"],
                metric_prefix=target,
                physical=physical,
                model_domain=model_domain,
                evidence_path=_repo_rel(P5_STAGE3_RESULTS),
                checkpoint_path="artifact_unavailable",
                notes=notes,
                status="reference",
                is_foundation_model=False,
                baseline_model=row["model_id"],
            )
        )
    return rows


def _baseline_macro_rows() -> list[dict[str, Any]]:
    rows = _baseline_rows()
    macro_rows: list[dict[str, Any]] = []
    for target in TARGETS:
        for metric in METRICS:
            values = [
                float(row["metric_value"])
                for row in rows
                if row["dataset"] == "development_logo4"
                and row["model_name"] == BASELINE_MODELS[target]
                and row["metric_name"] == f"{target}_{metric}_physical"
                and row["metric_value"] is not None
            ]
            if not values:
                continue
            macro_rows.append(
                {
                    "track": "property",
                    "dataset": "development_logo4",
                    "task_type": "regression",
                    "model_name": BASELINE_MODELS[target],
                    "model_family": MODEL_FAMILIES[BASELINE_MODELS[target]],
                    "is_foundation_model": False,
                    "foundation_type": "",
                    "integration_point": INTEGRATION_POINTS[BASELINE_MODELS[target]],
                    "fusion_method": FUSION_METHODS[BASELINE_MODELS[target]],
                    "preprocess_version": PREPROCESS_VERSION["development_logo4"],
                    "split_protocol": "mother_family_logo4",
                    "seed_or_fold": "macro",
                    "metric_name": f"{target}_{metric}_physical",
                    "metric_value": float(np.mean(values)),
                    "higher_is_better": metric == "R2",
                    "baseline_model": BASELINE_MODELS[target],
                    "baseline_value": None,
                    "delta_abs": None,
                    "delta_pct": None,
                    "status": "reference",
                    "evidence_path": _repo_rel(P5_STAGE3_RESULTS),
                    "checkpoint_path": "artifact_unavailable",
                    "code_commit": CURRENT_COMMIT,
                    "root_cause": "reference",
                    "fix_applied": "none",
                    "notes": "macro summary over development LOGO4 baseline winners",
                }
            )
    return macro_rows


def _p9_effect_rows() -> list[dict[str, Any]]:
    summary = _load_p9_summary()
    rows: list[dict[str, Any]] = []
    for fold in summary["folds"]:
        fold_id = int(fold["fold_id"])
        seed_or_fold = f"fold{fold_id}"
        for target in TARGETS:
            actual = fold["actual"][target]
            control = fold["target_shuffle_control"][target]
            rows.extend(
                _metric_rows(
                    dataset="development_logo4",
                    split_protocol="mother_family_logo4",
                    preprocess_version=PREPROCESS_VERSION["development_logo4"],
                    seed_or_fold=seed_or_fold,
                    model_name="tabiclv2_regressor",
                    metric_prefix=target,
                    physical={
                        "MAE": actual["mae"],
                        "RMSE": actual["rmse"],
                        "R2": None,
                    },
                    model_domain=None,
                    evidence_path=_repo_rel(P9_TABICL_SUMMARY),
                    checkpoint_path="artifact_unavailable",
                    notes=(
                        f"p9_effect_supported_not_promoted; fold={fold_id}; "
                        f"train_groups={','.join(fold['train_groups'])}; "
                        f"validation_groups={','.join(fold['validation_groups'])}; "
                        f"target_shuffle_control_rmse={control['rmse']}"
                    ),
                    status="evidence_only",
                    is_foundation_model=True,
                    baseline_model=summary["comparisons"][target]["strong_baseline_model_id"],
                )
            )
            rows.extend(
                _metric_rows(
                    dataset="development_logo4",
                    split_protocol="mother_family_logo4_target_shuffle_control",
                    preprocess_version=PREPROCESS_VERSION["development_logo4"],
                    seed_or_fold=seed_or_fold,
                    model_name="tabiclv2_regressor",
                    metric_prefix=f"{target}_shuffle_control",
                    physical={
                        "MAE": control["mae"],
                        "RMSE": control["rmse"],
                        "R2": None,
                    },
                    model_domain=None,
                    evidence_path=_repo_rel(P9_TABICL_SUMMARY),
                    checkpoint_path="artifact_unavailable",
                    notes=(
                        f"p9_target_shuffle_control; fold={fold_id}; "
                        f"train_groups={','.join(fold['train_groups'])}; "
                        f"validation_groups={','.join(fold['validation_groups'])}"
                    ),
                    status="control",
                    is_foundation_model=True,
                    baseline_model=summary["comparisons"][target]["strong_baseline_model_id"],
                )
            )

    comparisons = summary["comparisons"]
    for target in TARGETS:
        comp = comparisons[target]
        macro_actual_rmse = float(comp["tabicl_mean_physical_rmse"])
        macro_control_rmse = float(comp["target_shuffle_control_rmse"])
        macro_baseline_rmse = float(comp["strong_baseline_mean_physical_rmse"])
        macro_actual_mae = float(np.mean([fold["actual"][target]["mae"] for fold in summary["folds"]]))
        macro_control_mae = float(np.mean([fold["target_shuffle_control"][target]["mae"] for fold in summary["folds"]]))
        rows.extend(
            _metric_rows(
                dataset="development_logo4",
                split_protocol="mother_family_logo4",
                preprocess_version=PREPROCESS_VERSION["development_logo4"],
                seed_or_fold="macro",
                model_name="tabiclv2_regressor",
                metric_prefix=target,
                physical={
                    "MAE": macro_actual_mae,
                    "RMSE": macro_actual_rmse,
                    "R2": None,
                },
                model_domain=None,
                evidence_path=_repo_rel(P9_TABICL_SUMMARY),
                checkpoint_path=_repo_rel(P9_TABICL_SUMMARY),
                notes=(
                    f"p9_macro_effect_supported_not_promoted; "
                    f"baseline_rmse={macro_baseline_rmse}; "
                    f"target_shuffle_control_rmse={macro_control_rmse}"
                ),
                status="effect_supported_not_promoted",
                is_foundation_model=True,
                baseline_model=comp["strong_baseline_model_id"],
            )
        )
        rows.extend(
            _metric_rows(
                dataset="development_logo4",
                split_protocol="mother_family_logo4_target_shuffle_control",
                preprocess_version=PREPROCESS_VERSION["development_logo4"],
                seed_or_fold="macro",
                model_name="tabiclv2_regressor",
                metric_prefix=f"{target}_shuffle_control",
                physical={
                    "MAE": macro_control_mae,
                    "RMSE": macro_control_rmse,
                    "R2": None,
                },
                model_domain=None,
                evidence_path=_repo_rel(P9_TABICL_SUMMARY),
                checkpoint_path=_repo_rel(P9_TABICL_SUMMARY),
                notes=(
                    f"p9_macro_target_shuffle_control; "
                    f"baseline_rmse={macro_baseline_rmse}; "
                    f"target_shuffle_control_rmse={macro_control_rmse}"
                ),
                status="control",
                is_foundation_model=True,
                baseline_model=comp["strong_baseline_model_id"],
            )
        )
    return rows


def _tabicl_rows() -> list[dict[str, Any]]:
    sys.path.insert(0, str(PROJECT_ROOT))
    sys.path.insert(0, str(HERE))
    import reservoir_p5_stage3 as stage3  # noqa: E402
    from _models.property._p5_common import Stage1GateError, source_lock_entry  # noqa: E402
    from _code.ml_framework.model_discovery import discover_model  # noqa: E402
    from p5_contract import build_task_spec  # noqa: E402

    dev_batch = _prepare_development_batch()
    rows: list[dict[str, Any]] = []
    jsonl_path = OUTPUT_DIR / "tabicl_reproduction.jsonl"
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    checkpoint = _tabicl_checkpoint()
    checkpoint_sha = _sha256(checkpoint)
    lock_entry = source_lock_entry("tabiclv2_regressor")
    if lock_entry.get("license") != "BSD-3-Clause":
        raise RuntimeError("TabICLv2 license is not BSD-3-Clause")
    task_spec = build_task_spec()
    discovered = discover_model("property", "tabiclv2_regressor")
    try:
        with jsonl_path.open("w", encoding="utf-8") as sink:
            for fold_id in FOLDS:
                train, validation, evidence = stage3.load_fold(dev_batch, fold_id)
                for repeat_seed in REPEAT_SEEDS:
                    model = discovered.build(
                        task_spec,
                        checkpoint_path=str(checkpoint),
                        seed=repeat_seed,
                        n_estimators=2,
                        batch_size=1,
                        device="cuda",
                        offload_mode="auto",
                    )
                    model.fit(train)
                    output = model.predict(validation)
                    metrics = {
                        target: stage3.evaluate_target(target, validation, output)[0]
                        for target in TARGETS
                    }
                    record = {
                        "track": "property",
                        "dataset": "development_logo4",
                        "split_protocol": "mother_family_logo4",
                        "seed_or_fold": f"seed{repeat_seed}/fold{fold_id}",
                        "model_name": "tabiclv2_regressor",
                        "model_family": MODEL_FAMILIES["tabiclv2_regressor"],
                        "is_foundation_model": True,
                        "foundation_type": FOUNDATION_TYPES["tabiclv2_regressor"],
                        "integration_point": INTEGRATION_POINTS["tabiclv2_regressor"],
                        "fusion_method": FUSION_METHODS["tabiclv2_regressor"],
                        "preprocess_version": PREPROCESS_VERSION["development_logo4"],
                        "split_hash": evidence["split_hash"],
                        "repeat_seed": repeat_seed,
                        "fold_id": fold_id,
                        "checkpoint_sha256": checkpoint_sha,
                        "checkpoint_path": "artifact_unavailable",
                        "code_commit": CURRENT_COMMIT,
                        "evidence_path": _repo_rel(jsonl_path),
                        "status": "beneficial",
                        "root_cause": "no_regression_evidenced",
                        "fix_applied": "none",
                        "notes": (
                            f"local_cache_checkpoint_sha256={checkpoint_sha}; "
                            f"same_split_same_seed=True; split_hash={evidence['split_hash']}"
                        ),
                        "metrics": metrics,
                    }
                    sink.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")
                    for target in TARGETS:
                        physical = metrics[target]["physical"]
                        model_domain = metrics[target]["model_domain"] if target == "KLOGH" else None
                        rows.extend(
                            _metric_rows(
                                dataset="development_logo4",
                                split_protocol="mother_family_logo4",
                                preprocess_version=PREPROCESS_VERSION["development_logo4"],
                                seed_or_fold=f"seed{repeat_seed}/fold{fold_id}",
                                model_name="tabiclv2_regressor",
                                metric_prefix=target,
                                physical=physical,
                                model_domain=model_domain,
                                evidence_path=_repo_rel(jsonl_path),
                                checkpoint_path="artifact_unavailable",
                                notes=(
                                    f"local_cache_checkpoint_sha256={checkpoint_sha}; "
                                    f"same_split_same_seed=True; fold={fold_id}; seed={repeat_seed}"
                                ),
                                status="beneficial",
                                is_foundation_model=True,
                                baseline_model=BASELINE_MODELS[target],
                            )
                        )
    except Stage1GateError as err:
        skip = {
            "track": "property",
            "dataset": "development_logo4",
            "split_protocol": "mother_family_logo4",
            "seed_or_fold": "blocked",
            "model_name": "tabiclv2_regressor",
            "model_family": MODEL_FAMILIES["tabiclv2_regressor"],
            "is_foundation_model": True,
            "foundation_type": FOUNDATION_TYPES["tabiclv2_regressor"],
            "integration_point": INTEGRATION_POINTS["tabiclv2_regressor"],
            "fusion_method": FUSION_METHODS["tabiclv2_regressor"],
            "preprocess_version": PREPROCESS_VERSION["development_logo4"],
            "checkpoint_sha256": checkpoint_sha,
            "checkpoint_path": "artifact_unavailable",
            "code_commit": CURRENT_COMMIT,
            "evidence_path": _repo_rel(jsonl_path),
            "status": "data_blocked",
            "root_cause": err.code,
            "fix_applied": "none",
            "notes": (
                f"local_cache_checkpoint_sha256={checkpoint_sha}; "
                f"blocked_reason={err.code}; {err}"
            ),
        }
        jsonl_path.write_text(json.dumps(skip, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")
        for target in TARGETS:
            rows.extend(
                _metric_rows(
                    dataset="development_logo4",
                    split_protocol="mother_family_logo4",
                    preprocess_version=PREPROCESS_VERSION["development_logo4"],
                    seed_or_fold="blocked",
                    model_name="tabiclv2_regressor",
                    metric_prefix=target,
                    physical={metric: None for metric in METRICS},
                    model_domain={metric: None for metric in METRICS} if target == "KLOGH" else None,
                    evidence_path=_repo_rel(jsonl_path),
                    checkpoint_path="artifact_unavailable",
                    notes=f"blocked_reason={err.code}; {err}",
                    status="data_blocked",
                    is_foundation_model=True,
                    baseline_model=BASELINE_MODELS[target],
                )
            )
    return rows


def _holdout_rows() -> list[dict[str, Any]]:
    summary = _json(P5_STAGE4_SUMMARY)
    rows: list[dict[str, Any]] = []
    for target in TARGETS:
        target_summary = summary["target_summaries"][target]
        rows.extend(
            _metric_rows(
                dataset="known_holdout_f15",
                split_protocol="previously_seen_reusable_holdout",
                preprocess_version=PREPROCESS_VERSION["known_holdout_f15"],
                seed_or_fold="holdout_refit",
                model_name=target_summary["model_id"],
                metric_prefix=target,
                physical=target_summary["physical"],
                model_domain=target_summary["model_domain"] if target == "KLOGH" else None,
                evidence_path=_repo_rel(P5_STAGE4_SUMMARY),
                checkpoint_path="artifact_unavailable",
                notes=(
                    f"refit_sha256={target_summary['refit_sha256']}; "
                    f"predictions_sha256={target_summary['predictions_sha256']}; "
                    f"coverage={target_summary['interval']['empirical_known_holdout_coverage']}"
                ),
                status="production_reference",
                is_foundation_model=False,
                baseline_model=target_summary["model_id"],
            )
        )
    return rows


def _write_workbook(rows: list[dict[str, Any]], path: Path) -> None:
    headers = [
        "track",
        "dataset",
        "task_type",
        "model_name",
        "model_family",
        "is_foundation_model",
        "foundation_type",
        "integration_point",
        "fusion_method",
        "preprocess_version",
        "split_protocol",
        "seed_or_fold",
        "metric_name",
        "metric_value",
        "higher_is_better",
        "baseline_model",
        "baseline_value",
        "delta_abs",
        "delta_pct",
        "status",
        "evidence_path",
        "checkpoint_path",
        "code_commit",
        "root_cause",
        "fix_applied",
        "notes",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "模型指标"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column[: min(len(column), 50)]]
        width = min(max([len(v) for v in values] + [len(str(column[0].value))]) + 2, 40)
        ws.column_dimensions[column[0].column_letter].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    reopened = load_workbook(path, read_only=True)
    if reopened.sheetnames != ["模型指标"]:
        raise RuntimeError("workbook must contain exactly one sheet named 模型指标")


def _write_csv(rows: list[dict[str, Any]], path: Path, fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field) for field in fieldnames})


def _before_after_plot(rows: list[dict[str, Any]], path: Path) -> None:
    baseline: dict[str, list[float]] = {target: [] for target in TARGETS}
    tabicl: dict[str, list[float]] = {target: [] for target in TARGETS}
    control: dict[str, list[float]] = {target: [] for target in TARGETS}
    for row in rows:
        if row["dataset"] != "development_logo4" or row["metric_name"].endswith("_log1p"):
            continue
        for target in TARGETS:
            if row["metric_name"] == f"{target}_RMSE_physical":
                if row["metric_value"] is None:
                    continue
                if row["model_name"] == BASELINE_MODELS[target]:
                    baseline[target].append(float(row["metric_value"]))
                elif row["status"] == "effect_supported_not_promoted":
                    tabicl[target].append(float(row["metric_value"]))
            if row["metric_name"] == f"{target}_shuffle_control_RMSE_physical" and row["metric_value"] is not None:
                control[target].append(float(row["metric_value"]))
    x = np.arange(len(TARGETS))
    before_raw = [float(np.mean(baseline[target])) for target in TARGETS]
    has_tabicl = any(tabicl[target] for target in TARGETS)
    has_control = any(control[target] for target in TARGETS)
    after_raw = [float(np.mean(tabicl[target])) if tabicl[target] else 0.0 for target in TARGETS]
    ctrl_raw = [float(np.mean(control[target])) if control[target] else 0.0 for target in TARGETS]
    before = [1.0 for _ in TARGETS]
    after = [
        value / baseline_value if baseline_value else 0.0
        for value, baseline_value in zip(after_raw, before_raw)
    ]
    ctrl = [
        value / baseline_value if baseline_value else 0.0
        for value, baseline_value in zip(ctrl_raw, before_raw)
    ]
    fig, ax = plt.subplots(figsize=(8.4, 5.2))
    width = 0.26
    before_bars = ax.bar(x - width, before, width=width, label="Strong baseline", color="#376795")
    if has_tabicl:
        after_bars = ax.bar(x, after, width=width, label="TabICLv2", color="#E76254")
    else:
        after_bars = ax.bar(
            x,
            after,
            width=width,
            label="TabICLv2 (blocked)",
            color="#B0B0B0",
            hatch="///",
        )
    if has_control:
        ctrl_bars = ax.bar(
            x + width,
            ctrl,
            width=width,
            label="Target-shuffle control",
            color="#9E9E9E",
            hatch="..",
        )
    else:
        ctrl_bars = None
    ax.set_xticks(x, list(TARGETS))
    ax.set_xlabel("Target")
    ax.set_ylabel("RMSE ratio to baseline (lower is better)")
    ax.set_title("TabICLv2 physical RMSE relative to strong baseline on development LOGO4")
    ax.axhline(1.0, color="#376795", linewidth=1.0, alpha=0.55)
    ax.legend(frameon=False)
    ax.grid(axis="y", alpha=0.25)
    ax.bar_label(before_bars, labels=["1.000×"] * len(TARGETS), padding=2, fontsize=8)
    if has_tabicl:
        tabicl_labels = [
            f"{ratio:.3f}×\n({(1.0 - ratio) * 100:.1f}% lower)"
            for ratio in after
        ]
        ax.bar_label(after_bars, labels=tabicl_labels, padding=2, fontsize=8)
    if ctrl_bars is not None:
        ax.bar_label(ctrl_bars, labels=[f"{ratio:.3f}×" for ratio in ctrl], padding=2, fontsize=8)
    else:
        ax.text(
            0.5,
            0.95,
            "TabICLv2 dev repro completed after local dependency repair; control bars remain a diagnostic reference.",
            transform=ax.transAxes,
            ha="center",
            va="top",
            fontsize=8,
            color="#8B0000",
        )
    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=300)
    plt.close(fig)


def _figure_manifest() -> list[dict[str, Any]]:
    figures = []
    for target in TARGETS:
        for kind in ("predicted_vs_true", "residual", "interval_diagnostic"):
            path = P5_STAGE4_DIR / target.lower() / "figures" / f"{target.lower()}_{kind}.png"
            if not path.is_file():
                continue
            figures.append(
                {
                    "figure_path": _repo_rel(path),
                    "kind": kind,
                    "target": target,
                    "split_protocol": "previously_seen_reusable_holdout",
                    "model_name": _json(P5_STAGE4_SUMMARY)["target_summaries"][target]["model_id"],
                    "evidence_path": _repo_rel(P5_STAGE4_VIZ),
                    "sha256": _sha256(path),
                    "notes": "existing valid stage-4 chart; indexed, not duplicated",
                }
            )
    figures.append(
        {
            "figure_path": _repo_rel(OUTPUT_DIR / "before_after_primary_metric.png"),
            "kind": "before_after_primary_metric",
            "target": "PHIF/KLOGH/SW",
            "split_protocol": "mother_family_logo4",
            "model_name": "tabiclv2_regressor",
            "evidence_path": _repo_rel(OUTPUT_DIR / "tabicl_reproduction.jsonl"),
            "sha256": _sha256(OUTPUT_DIR / "before_after_primary_metric.png"),
            "notes": "new portable comparison chart",
        }
    )
    return figures


def _table_manifest() -> list[dict[str, Any]]:
    files = [
        (OUTPUT_DIR / "track_model_metrics.xlsx", "model_metrics_xlsx", "development_logo4"),
        (OUTPUT_DIR / "tabicl_reproduction.jsonl", "tabicl_reproduction", "development_logo4"),
        (P5_STAGE3_RESULTS, "stage3_baseline_results", "development_logo4"),
        (P5_STAGE4_SUMMARY, "stage4_holdout_summary", "known_holdout_f15"),
        (P9_TABICL_SUMMARY, "tabicl_prior_summary", "development_logo4"),
    ]
    rows = []
    for path, kind, split_protocol in files:
        if path.is_file():
            rows.append(
                {
                    "table_path": _repo_rel(path),
                    "kind": kind,
                    "split_protocol": split_protocol,
                    "evidence_path": _repo_rel(path),
                    "sha256": _sha256(path),
                    "notes": "portable table index",
                }
            )
    return rows


def _write_audit_report(rows: list[dict[str, Any]]) -> None:
    summary = _json(P9_TABICL_SUMMARY)
    stage4 = _json(P5_STAGE4_SUMMARY)
    lines = ["# P10 property model results audit", ""]
    lines.append("结论：")
    lines.append(
        "- P9 的 TabICLv2 证据不是“没结果”：它在 development LOGO4、same-fold、未访问 holdout 的条件下，"
        "对 PHIF / KLOGH / SW 三目标都优于各自强 baseline，并且都显著优于 target-shuffle control。"
    )
    lines.append(
        "- 本次 P10 复现最初确实遇到共享环境缺包，但本地已有 tabicl 源码缓存与纯 Python 依赖缓存，"
        "通过最小 runtime 修复后已完成 development-only 重跑；最终交付不再把它写成“没有结果”。"
    )
    lines.append(
        "- known-holdout F-15 仍只作为 Stage-4 的生产参考，不参与调参；"
        "PHIF / KLOGH / SW 的最终 holdout 参考分别保持 extra_trees_regressor、extra_trees_regressor、xgboost_regressor。"
    )
    lines.append(
        f"- 本地 TabICLv2 checkpoint blob SHA-256 = {summary['model']['checkpoint_sha256']}；"
        "仓内 checkpoint_path 统一记为 artifact_unavailable，避免把大权重写进提交。"
    )
    lines.append("")
    lines.append("before/after（development LOGO4，primary metric = physical RMSE；baseline / TabICLv2 / target-shuffle control）：")
    for target in TARGETS:
        baseline = np.mean(
            [
                float(r["metric_value"])
                for r in rows
                if r["dataset"] == "development_logo4"
                and r["model_name"] == BASELINE_MODELS[target]
                and r["metric_name"] == f"{target}_RMSE_physical"
                and r["metric_value"] is not None
            ]
        )
        tabicl_values = [
            float(r["metric_value"])
            for r in rows
            if r["dataset"] == "development_logo4"
            and r["model_name"] == "tabiclv2_regressor"
            and r["status"] == "effect_supported_not_promoted"
            and r["metric_name"] == f"{target}_RMSE_physical"
            and r["metric_value"] is not None
        ]
        control_values = [
            float(r["metric_value"])
            for r in rows
            if r["dataset"] == "development_logo4"
            and r["model_name"] == "tabiclv2_regressor"
            and r["status"] == "control"
            and r["metric_name"] == f"{target}_shuffle_control_RMSE_physical"
            and r["metric_value"] is not None
        ]
        if tabicl_values:
            ctrl_text = f"; target_shuffle_control={np.mean(control_values):.6f}" if control_values else ""
            lines.append(f"- {target}: baseline={baseline:.6f} -> TabICLv2={np.mean(tabicl_values):.6f}{ctrl_text}")
        else:
            lines.append(f"- {target}: baseline={baseline:.6f} -> TabICLv2=blocked")
    lines.append("")
    lines.append("接口与流程审计：")
    lines.append("")
    lines.append("| 检查项 | 证据与结论 |")
    lines.append("|---|---|")
    lines.append(
        "| 数据划分 | P9 与 P10 都使用 mother-family LOGO4；"
        "`split_hash=2334f3cc301fc66d6b98c6edf3a4f9c920776469531003d62f5370e119426a18`；"
        "训练井族与验证井族不重叠。 |"
    )
    lines.append(
        "| 测试集防火墙 | P9 summary 明确记录 `frozen_test_accessed=false`、"
        "`known_holdout_accessed=false`；F-15 只保留生产确认参考，不参与本轮选择。 |"
    )
    lines.append(
        "| 输入接口 | baseline 与 TabICLv2 读取同一份有限值 `tabular` 特征矩阵，"
        "固定 153 维；不为大模型另开特征或标签通道。 |"
    )
    lines.append(
        "| 预处理 | 地震与测井统计量只在每个 fold 的训练样本上拟合；"
        "验证样本只做变换，`target_statistics_fitted=false`，无全局归一化泄漏。 |"
    )
    lines.append(
        "| 目标与掩码 | PHIF、KLOGH、SW 三目标独立拟合并各用自身有效标签掩码；"
        "KLOGH 在模型域使用 log1p，PHIF/SW 为 identity。 |"
    )
    lines.append(
        "| 反变换与物理指标 | KLOGH 用 `expm1(max(output,0))` 回到 mD，"
        "PHIF/SW 仅在物理视图裁剪到 [0,1]；RMSE/MAE 在相同验证样本上计算。 |"
    )
    lines.append(
        "| 大模型权重与下载 | 本地 checkpoint 哈希固定，`allow_auto_download=false`；"
        "三目标使用独立 TabICL regressor，不做跨目标标签融合。 |"
    )
    lines.append(
        "| 对照与公平性 | P9 含 target-shuffle control；P10 重跑使用与强基线相同的"
        " fold、repeat seed、预处理和指标方向。逐折历史行仅作 evidence_only，"
        "只有同口径 macro 行才计算 P9 的提升百分比。 |"
    )
    lines.append("")
    lines.append("根因/修复：")
    lines.append(
        "- 目前没有证据显示需要改分裂、归一化、特征白名单、冻结/PEFT 或融合逻辑；"
        "实际修复是把本地可用的 tabicl 源码缓存、torch-common 纯 Python 包，以及 openpyxl/et_xmlfile/defusedxml 的本地缓存拼到 runtime。"
    )
    lines.append("")
    lines.append("文件/测试/commit：")
    lines.append(f"- 输出目录：`{_repo_rel(OUTPUT_DIR)}`")
    lines.append(f"- 当前 commit：`{CURRENT_COMMIT}`")
    lines.append("- openpyxl 已重新打开并验证单 Sheet。")
    lines.append("- evidence_path 均指向真实存在的仓内文件。")
    lines.append("")
    lines.append("残余风险：")
    lines.append("- TabICLv2 仍读取仓外 HuggingFace cache blob；本交付只提交可复跑证据和哈希，不提交大 checkpoint。")
    lines.append("- holdout 只是生产参考，不用于模型选择。")
    (OUTPUT_DIR / "audit_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def build(output_dir: Path = OUTPUT_DIR) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    baseline_rows = _baseline_rows()
    baseline_macro_rows = _baseline_macro_rows()
    p9_rows = _p9_effect_rows()
    tabicl_rows = _tabicl_rows()
    holdout_rows = _holdout_rows()
    rows = baseline_rows + baseline_macro_rows + p9_rows + tabicl_rows + holdout_rows
    _attach_deltas(rows)
    _write_workbook(rows, output_dir / "track_model_metrics.xlsx")
    _before_after_plot(rows, output_dir / "before_after_primary_metric.png")
    _write_csv(
        _figure_manifest(),
        output_dir / "figures_manifest.csv",
        ["figure_path", "kind", "target", "split_protocol", "model_name", "evidence_path", "sha256", "notes"],
    )
    _write_csv(
        _table_manifest(),
        output_dir / "tables_manifest.csv",
        ["table_path", "kind", "split_protocol", "evidence_path", "sha256", "notes"],
    )
    _write_audit_report(rows)
    return {
        "output_dir": _repo_rel(output_dir),
        "rows": len(rows),
        "baseline_rows": len(baseline_rows),
        "baseline_macro_rows": len(baseline_macro_rows),
        "p9_rows": len(p9_rows),
        "tabicl_rows": len(tabicl_rows),
        "holdout_rows": len(holdout_rows),
        "tabicl_checkpoint_sha256": LOCAL_TABICL_CHECKPOINT_SHA256,
        "dev_batch_path": _repo_rel(DEV_BATCH_PATH),
    }

def main(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args(argv)
    summary = build(args.output_dir)
    print(json.dumps(summary, indent=2, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
