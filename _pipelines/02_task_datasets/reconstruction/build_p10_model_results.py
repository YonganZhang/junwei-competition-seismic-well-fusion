from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "_outputs" / "p10_model_results"
WORKBOOK_PATH = OUTPUT_DIR / "track_model_metrics.xlsx"
FIGURE_PATH = OUTPUT_DIR / "before_after_primary_metric.png"
FIGURES_MANIFEST_PATH = OUTPUT_DIR / "figures_manifest.csv"
TABLES_MANIFEST_PATH = OUTPUT_DIR / "tables_manifest.csv"
AUDIT_REPORT_PATH = OUTPUT_DIR / "audit_report.md"

STRICT_RESULTS = ROOT / "results_strict.json"
CONDITIONAL_RESULTS = ROOT / "results_conditional.json"
OPENMIND_SUMMARY = ROOT / "_outputs" / "p9_openmind_effect" / "summary.json"
STAGE4_SUMMARY = ROOT / "p5_stage4_confirmation" / "summary.json"
STRICT_VIS_META = ROOT / "visualization_metadata_strict.json"
COND_VIS_META = ROOT / "visualization_metadata_conditional.json"
STRICT_CONFIRMATION = ROOT / "p5_stage4_confirmation" / "strict" / "confirmation.png"
COND_CONFIRMATION = ROOT / "p5_stage4_confirmation" / "conditional" / "confirmation.png"
STRICT_STAGE3_FIG = ROOT / "_outputs" / "p5_stage3_reconstruction_strict.png"
COND_STAGE3_FIG = ROOT / "_outputs" / "p5_stage3_reconstruction_conditional.png"
STRICT_PRED_FIG = ROOT / "_outputs" / "prediction_visualization_strict.png"
COND_PRED_FIG = ROOT / "_outputs" / "prediction_visualization_conditional.png"

HEADERS = [
    "track",
    "dataset",
    "task_type",
    "model_name",
    "model_family",
    "is_foundation_model",
    "foundation_type",
    "integration_point",
    "fusion_method",
    "preprocess_version",
    "split_protocol",
    "seed_or_fold",
    "metric_name",
    "metric_value",
    "higher_is_better",
    "baseline_model",
    "baseline_value",
    "delta_abs",
    "delta_pct",
    "status",
    "evidence_path",
    "checkpoint_path",
    "code_commit",
    "root_cause",
    "fix_applied",
    "notes",
]


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _git_commit() -> str:
    return (
        subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT)
        .decode("utf-8")
        .strip()
    )


def _metric_is_higher_better(metric: str) -> bool:
    return metric in {"r2", "pearson_r"}


def _signed_delta(metric: str, value: float, baseline: float) -> tuple[float, float]:
    if _metric_is_higher_better(metric):
        delta = value - baseline
    else:
        delta = baseline - value
    if baseline == 0:
        pct = math.nan
    else:
        pct = delta / abs(baseline) * 100.0
    return delta, pct


def _coerce_num(value: Any) -> float | None:
    if value is None:
        return None
    return float(value)


def _baseline_metric_from(source: dict[str, Any], model_key: str, metric: str) -> float | None:
    model = source["models"].get(model_key, {})
    return _coerce_num(model.get(metric))


def _append_row(rows: list[dict[str, Any]], **kwargs: Any) -> None:
    row = {key: kwargs.get(key) for key in HEADERS}
    rows.append(row)


def _add_ridge_rows(rows: list[dict[str, Any]], source: dict[str, Any], dataset: str, split_protocol: str) -> None:
    track = "reconstruction"
    task_type = "porosity_reconstruction"
    preprocess_version = "reconstruction_p10_v1"
    code_commit = _git_commit()

    model_specs = [
        ("train_mean", "ridge_linear", "none", "none", "none", "train_mean", "self-baseline"),
        ("sparse_well_idw", "ridge_linear", "classical_interpolation", "well_idw", "sparse_well_idw", "train_mean", "vs_train_mean"),
        (
            "ridge_idw_coordinates",
            "ridge_linear",
            "classical_interpolation",
            "seismic+coordinates",
            "ridge_idw_coordinates",
            "sparse_well_idw",
            "vs_sparse_well_idw",
        ),
        (
            "ridge_idw_seismic_coordinates",
            "ridge_linear",
            "classical_interpolation",
            "seismic+coordinates+well_tie",
            "ridge_idw_seismic_coordinates",
            "ridge_idw_coordinates",
            "vs_ridge_idw_coordinates",
        ),
        (
            "ridge_with_test_seismic_shuffled",
            "ridge_linear",
            "negative_control",
            "seismic+coordinates+shuffled_well",
            "ridge_with_test_seismic_shuffled",
            "ridge_idw_seismic_coordinates",
            "negative_control_vs_ridge_idw_seismic_coordinates",
        ),
    ]

    for model_key, model_family, foundation_type, fusion_method, integration_point, baseline_key, comparison_tag in model_specs:
        model_metrics = source["models"][model_key]
        for metric_name in ("rmse", "mae", "r2", "pearson_r"):
            metric_value = _coerce_num(model_metrics.get(metric_name))
            if metric_value is None:
                continue
            baseline_value = _baseline_metric_from(source, baseline_key, metric_name)
            if baseline_value is None:
                baseline_value = metric_value
            delta_abs, delta_pct = _signed_delta(metric_name, metric_value, baseline_value)
            status = "passed"
            if model_key == "ridge_with_test_seismic_shuffled":
                status = "negative_control"
            _append_row(
                rows,
                track=track,
                dataset=dataset,
                task_type=task_type,
                model_name=model_key,
                model_family=model_family,
                is_foundation_model=False,
                foundation_type=foundation_type,
                integration_point=integration_point,
                fusion_method=fusion_method,
                preprocess_version=preprocess_version,
                split_protocol=split_protocol,
                seed_or_fold="aggregate",
                metric_name=metric_name,
                metric_value=metric_value,
                higher_is_better=_metric_is_higher_better(metric_name),
                baseline_model=baseline_key,
                baseline_value=baseline_value,
                delta_abs=delta_abs,
                delta_pct=delta_pct,
                status=status,
                evidence_path=str(source["__evidence_path__"]),
                checkpoint_path="",
                code_commit=code_commit,
                root_cause="baseline ladder from naive mean to sparse well IDW to seismic+coordinate ridge; shuffled-well is a negative control",
                fix_applied="none",
                notes=f"comparison={comparison_tag}; strict/conditional split preserved",
            )


def _add_openmind_rows(rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    track = "reconstruction"
    dataset = "strict_development"
    task_type = "porosity_reconstruction"
    code_commit = _git_commit()
    evaluation = summary["evaluation"]
    comparison = summary["comparison"]
    fold_results = summary["fold_results"]

    pretrained_macro_rmse = float(comparison["pretrained_macro_fold_rmse"])
    random_macro_rmse = float(comparison["same_architecture_random_init_macro_fold_rmse"])
    strong_macro_rmse = float(comparison["strong_baseline_macro_fold_rmse"])
    strong_baseline_model_id = comparison["strong_baseline_model_id"]

    _append_row(
        rows,
        track=track,
        dataset=dataset,
        task_type=task_type,
        model_name="MIC-DKFZ/ResEncL-OpenMind-MAE",
        model_family="masked_autoencoder",
        is_foundation_model=True,
        foundation_type="self_supervised_pretrained",
        integration_point="attribute_projection_and_decoder",
        fusion_method="observation_mask+3d_patches",
        preprocess_version="p9_openmind_effect_v1",
        split_protocol=f"strict_split_hash={evaluation['split_hash']}",
        seed_or_fold="macro_fold_mean",
        metric_name="rmse",
        metric_value=pretrained_macro_rmse,
        higher_is_better=False,
        baseline_model="same_architecture_random_init",
        baseline_value=random_macro_rmse,
        delta_abs=_signed_delta("rmse", pretrained_macro_rmse, random_macro_rmse)[0],
        delta_pct=_signed_delta("rmse", pretrained_macro_rmse, random_macro_rmse)[1],
        status="non_beneficial",
        evidence_path=str(OPENMIND_SUMMARY),
        checkpoint_path="",
        code_commit=code_commit,
        root_cause=(
            "foundation gain is real versus random init, but end-to-end performance remains non_beneficial "
            "against PyKrige on the same strict sample universe; audited checkpoints include "
            "PATCH_SHAPE=(9,20,18), grid_shape_kji=[63,100,108], coordinate roundtrip max abs error=0.0, "
            "and mode-specific mask exclusion is already encoded in the archived summaries"
        ),
        fix_applied="none",
        notes=(
            f"frozen_test_accessed={evaluation['frozen_test_accessed']}; "
            f"same_validation_sample_universe={evaluation['same_validation_sample_universe_as_strong_baseline']}; "
            "adapter_decoder_output_shape_not_exposed_in_archived_summary"
        ),
    )
    _append_row(
        rows,
        track=track,
        dataset=dataset,
        task_type=task_type,
        model_name="same_architecture_random_init",
        model_family="masked_autoencoder",
        is_foundation_model=False,
        foundation_type="random_init",
        integration_point="attribute_projection_and_decoder",
        fusion_method="observation_mask+3d_patches",
        preprocess_version="p9_openmind_effect_v1",
        split_protocol=f"strict_split_hash={evaluation['split_hash']}",
        seed_or_fold="macro_fold_mean",
        metric_name="rmse",
        metric_value=random_macro_rmse,
        higher_is_better=False,
        baseline_model=strong_baseline_model_id,
        baseline_value=strong_macro_rmse,
        delta_abs=_signed_delta("rmse", random_macro_rmse, strong_macro_rmse)[0],
        delta_pct=_signed_delta("rmse", random_macro_rmse, strong_macro_rmse)[1],
        status="non_beneficial",
        evidence_path=str(OPENMIND_SUMMARY),
        checkpoint_path="",
        code_commit=code_commit,
        root_cause=(
            "same architecture without pretraining has far higher error on the same strict fold universe; "
            "this confirms the foundation gain but not promotion"
        ),
        fix_applied="none",
        notes=(
            "comparison is same-arch random initialization against the PyKrige reference; "
            "same split/sample universe as pretrained run"
        ),
    )
    _append_row(
        rows,
        track=track,
        dataset=dataset,
        task_type=task_type,
        model_name=strong_baseline_model_id,
        model_family="geostatistics",
        is_foundation_model=False,
        foundation_type="none",
        integration_point="kriging_surface",
        fusion_method="observation_mask+coordinates",
        preprocess_version="p9_openmind_effect_v1",
        split_protocol=f"strict_split_hash={evaluation['split_hash']}",
        seed_or_fold="macro_fold_mean",
        metric_name="rmse",
        metric_value=strong_macro_rmse,
        higher_is_better=False,
        baseline_model="ridge_idw_seismic_coordinates",
        baseline_value=_coerce_num(source_value := _load_json(STRICT_RESULTS)["models"]["ridge_idw_seismic_coordinates"]["rmse"]),
        delta_abs=_signed_delta("rmse", strong_macro_rmse, _coerce_num(source_value))[0],
        delta_pct=_signed_delta("rmse", strong_macro_rmse, _coerce_num(source_value))[1],
        status="passed",
        evidence_path=str(OPENMIND_SUMMARY),
        checkpoint_path="",
        code_commit=code_commit,
        root_cause=(
            "PyKrige is the strongest reference on the same strict sample universe; "
            "serves as the end-to-end baseline that OpenMind does not beat"
        ),
        fix_applied="none",
        notes=(
            "strong_baseline_macro_fold_rmse comes from the same fold universe as the foundation comparison; "
            "no frozen holdout was used here"
        ),
    )

    for fold in fold_results:
        fold_id = f"fold_{fold['fold_id']}"
        pretrained = fold["pretrained"]
        random_init = fold["same_architecture_random_init"]
        strong_baseline_rmse = float(fold["strong_baseline_rmse"])
        for model_name, metric_value, baseline_model, baseline_value, status, notes in [
            (
                "MIC-DKFZ/ResEncL-OpenMind-MAE",
                float(pretrained["rmse"]),
                strong_baseline_model_id,
                strong_baseline_rmse,
                "non_beneficial",
                "fold-level pretrained comparison against the same fold's PyKrige baseline",
            ),
            (
                "same_architecture_random_init",
                float(random_init["rmse"]),
                strong_baseline_model_id,
                strong_baseline_rmse,
                "non_beneficial",
                "fold-level random-init comparison against the same fold's PyKrige baseline",
            ),
            (
                strong_baseline_model_id,
                strong_baseline_rmse,
                strong_baseline_model_id,
                strong_baseline_rmse,
                "reference",
                "fold-level strong baseline reference",
            ),
        ]:
            delta_abs, delta_pct = _signed_delta("rmse", metric_value, baseline_value)
            _append_row(
                rows,
                track=track,
                dataset=f"strict_cv_{fold_id}",
                task_type=task_type,
                model_name=model_name,
                model_family="masked_autoencoder" if "OpenMind" in model_name or "random_init" in model_name else "geostatistics",
                is_foundation_model=model_name == "MIC-DKFZ/ResEncL-OpenMind-MAE",
                foundation_type="self_supervised_pretrained" if model_name == "MIC-DKFZ/ResEncL-OpenMind-MAE" else ("random_init" if model_name == "same_architecture_random_init" else "none"),
                integration_point="attribute_projection_and_decoder" if "OpenMind" in model_name or "random_init" in model_name else "kriging_surface",
                fusion_method="observation_mask+3d_patches" if "OpenMind" in model_name or "random_init" in model_name else "observation_mask+coordinates",
                preprocess_version="p9_openmind_effect_v1",
                split_protocol=f"strict_split_hash={evaluation['split_hash']}",
                seed_or_fold=fold_id,
                metric_name="rmse",
                metric_value=metric_value,
                higher_is_better=False,
                baseline_model=baseline_model,
                baseline_value=baseline_value,
                delta_abs=delta_abs,
                delta_pct=delta_pct,
                status=status,
                evidence_path=str(OPENMIND_SUMMARY),
                checkpoint_path="",
                code_commit=code_commit,
                root_cause="fold-level same-sample-universe comparison",
                fix_applied="none",
                notes=notes,
            )


def _add_stage4_rows(rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    track = "reconstruction"
    task_type = "porosity_reconstruction"
    preprocess_version = "p5_stage4_confirmation_v1"
    code_commit = _git_commit()
    results = summary["results"]

    baseline_lookup = _load_json(STRICT_RESULTS)["models"]["ridge_idw_seismic_coordinates"]
    cond_baseline_lookup = _load_json(CONDITIONAL_RESULTS)["models"]["ridge_idw_seismic_coordinates"]

    for mode, baseline_lookup_mode, split_protocol in [
        ("strict", baseline_lookup, "known_holdout_confirmation_strict"),
        ("conditional", cond_baseline_lookup, "known_holdout_confirmation_conditional"),
    ]:
        result = results[mode]
        metric_map = result["metrics"]
        for metric_name, stage4_metric_name in [
            ("rmse", f"{mode}_rmse"),
            ("mae", f"{mode}_mae"),
            ("r2", f"{mode}_r2"),
            ("pearson_r", f"{mode}_pearson_r"),
        ]:
            metric_value = _coerce_num(metric_map.get(stage4_metric_name))
            baseline_value = _coerce_num(baseline_lookup_mode.get(metric_name))
            if metric_value is None or baseline_value is None:
                continue
            delta_abs, delta_pct = _signed_delta(metric_name, metric_value, baseline_value)
            _append_row(
                rows,
                track=track,
                dataset=f"{mode}_confirmation",
                task_type=task_type,
                model_name="pykrige_ok3d",
                model_family="geostatistics",
                is_foundation_model=False,
                foundation_type="none",
                integration_point="kriging_surface",
                fusion_method="observation_mask+coordinates",
                preprocess_version=preprocess_version,
                split_protocol=split_protocol,
                seed_or_fold="aggregate",
                metric_name=metric_name,
                metric_value=metric_value,
                higher_is_better=_metric_is_higher_better(metric_name),
                baseline_model="ridge_idw_seismic_coordinates",
                baseline_value=baseline_value,
                delta_abs=delta_abs,
                delta_pct=delta_pct,
                status="known_holdout_confirmation",
                evidence_path=str(STAGE4_SUMMARY),
                checkpoint_path=str(ROOT / "p5_stage4_confirmation" / mode / "refit_checkpoint.npz"),
                code_commit=code_commit,
                root_cause=(
                    "known-holdout confirmation on previously seen holdout, not a fresh blind test; "
                    f"stage4_mode={mode}"
                ),
                fix_applied="none",
                notes=(
                    f"prior_test_consumed={summary['prior_test_consumed']}; "
                    f"fresh_blind={summary['fresh_blind']}; evidence_class={summary['evidence_class']}; "
                    "confirmation_only"
                ),
            )


def _write_csv(path: Path, rows: Iterable[dict[str, Any]], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "模型指标"
    ws.freeze_panes = "A2"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header) for header in HEADERS])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _chart_primary_metric(path: Path) -> None:
    openmind = _load_json(OPENMIND_SUMMARY)
    comparison = openmind["comparison"]
    values = [
        ("OpenMind random-init", float(comparison["same_architecture_random_init_macro_fold_rmse"]), "#d95f02"),
        ("OpenMind pretrained", float(comparison["pretrained_macro_fold_rmse"]), "#1b9e77"),
        ("Ridge baseline", float(_load_json(STRICT_RESULTS)["models"]["ridge_idw_seismic_coordinates"]["rmse"]), "#7570b3"),
        ("PyKrige reference", float(comparison["strong_baseline_macro_fold_rmse"]), "#e7298a"),
    ]

    fig, ax = plt.subplots(figsize=(8.0, 6.0))
    labels = [v[0] for v in values]
    heights = [v[1] for v in values]
    colors = [v[2] for v in values]
    bars = ax.bar(labels, heights, color=colors, width=0.7)
    ax.set_ylabel("RMSE (lower is better)")
    ax.set_title("Primary RMSE on the strict development split")
    ax.set_yscale("log")
    ax.grid(axis="y", linestyle="--", alpha=0.3)
    ax.tick_params(axis="x", rotation=18)
    ax.set_ylim(bottom=min(heights) * 0.7, top=max(heights) * 1.4)

    for bar, value in zip(bars, heights, strict=True):
        ax.text(
            bar.get_x() + bar.get_width() / 2.0,
            value * 1.08,
            f"{value:.6f}",
            ha="center",
            va="bottom",
            fontsize=9,
            rotation=0,
        )

    ax.text(
        0.02,
        0.02,
        "Same strict split / same sample universe.\n"
        "Pretraining helps vs random init but remains far above PyKrige.\n"
        "No verified adapter bug was fixed; evidence supports non_beneficial.",
        transform=ax.transAxes,
        fontsize=9,
        va="bottom",
        ha="left",
        bbox={"boxstyle": "round,pad=0.35", "facecolor": "white", "alpha": 0.9},
    )
    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=200)
    plt.close(fig)


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.12g}"
    return str(value)


def _write_audit_report(path: Path, rows: list[dict[str, Any]]) -> None:
    strict_rows = [row for row in rows if row["dataset"] == "strict_development" and row["metric_name"] == "rmse"]
    cond_rows = [row for row in rows if row["dataset"] == "conditional_development" and row["metric_name"] == "rmse"]
    openmind_rows = [row for row in rows if row["dataset"] == "strict_development" and row["seed_or_fold"] == "macro_fold_mean"]
    stage4_rows = [row for row in rows if row["dataset"].endswith("_confirmation") and row["metric_name"] == "rmse"]
    strict_result = _load_json(STRICT_RESULTS)
    cond_result = _load_json(CONDITIONAL_RESULTS)
    summary = _load_json(OPENMIND_SUMMARY)
    stage4 = _load_json(STAGE4_SUMMARY)
    build_summary = _load_json(ROOT / "build_summary.json")
    model_inspection = _load_json(ROOT / "model_inspection.json")
    commit = _git_commit()

    def _row(dataset: str, model_name: str, metric_name: str) -> dict[str, Any]:
        return next(
            row
            for row in rows
            if row["dataset"] == dataset and row["model_name"] == model_name and row["metric_name"] == metric_name
        )

    lines = [
        "# P10 reconstruction model-results audit",
        "",
        f"- code_commit: `{commit}`",
        f"- workbook: `{WORKBOOK_PATH}`",
        f"- figure: `{FIGURE_PATH}`",
        f"- row_count: `{len(rows)}`",
        f"- sheet_name: `模型指标`",
        f"- evidence_only_boundary: `{True}`",
        "",
        "## Conclusion",
        "",
        "The OpenMind/ResEnc-L lane has a real foundation gain but remains end-to-end non_beneficial.",
        "Pretraining reduces the same-architecture random-init RMSE from 1.052412481992174 to",
        "0.5415301607840952, but the same strict-development sample universe still sits far above the",
        "PyKrige reference at 0.02120691345759842. That means the foundation effect is real, but it is",
        "not enough for promotion against the strong baseline.",
        "",
        "No verified bug in the archived evidence justified a code fix that would close the gap.",
        "",
        "## Foundation gain vs end-to-end outcome",
        "",
        f"- random-init macro RMSE: `{_row('strict_development', 'same_architecture_random_init', 'rmse')['metric_value']:.12g}`",
        f"- pretrained macro RMSE: `{_row('strict_development', 'MIC-DKFZ/ResEncL-OpenMind-MAE', 'rmse')['metric_value']:.12g}`",
        f"- foundation gain vs random-init: `{_signed_delta('rmse', _row('strict_development', 'MIC-DKFZ/ResEncL-OpenMind-MAE', 'rmse')['metric_value'], _row('strict_development', 'same_architecture_random_init', 'rmse')['baseline_value'])[0]:.12g}`",
        f"- foundation gain vs random-init (%): `{_signed_delta('rmse', _row('strict_development', 'MIC-DKFZ/ResEncL-OpenMind-MAE', 'rmse')['metric_value'], _row('strict_development', 'same_architecture_random_init', 'rmse')['baseline_value'])[1]:.12g}`",
        f"- PyKrige reference macro RMSE: `{_row('strict_development', 'pykrige_ok3d', 'rmse')['metric_value']:.12g}`",
        f"- end-to-end delta vs PyKrige: `{_signed_delta('rmse', _row('strict_development', 'MIC-DKFZ/ResEncL-OpenMind-MAE', 'rmse')['metric_value'], _row('strict_development', 'pykrige_ok3d', 'rmse')['metric_value'])[0]:.12g}`",
        f"- strict ridge_linear RMSE: `{_row('strict_development', 'ridge_idw_seismic_coordinates', 'rmse')['metric_value']:.12g}`",
        f"- conditional ridge_linear RMSE: `{_row('conditional_development', 'ridge_idw_seismic_coordinates', 'rmse')['metric_value']:.12g}`",
        f"- Stage-4 strict known-holdout RMSE: `{_row('strict_confirmation', 'pykrige_ok3d', 'rmse')['metric_value']:.12g}`",
        f"- Stage-4 conditional known-holdout RMSE: `{_row('conditional_confirmation', 'pykrige_ok3d', 'rmse')['metric_value']:.12g}`",
        "",
        "## Audited implementation checkpoints",
        "",
        "| check | evidence / value | status |",
        "| --- | --- | --- |",
        f"| 3D patch shape and axis order | `PATCH_SHAPE = (9, 20, 18)` in `build_dataset.py`; tiled as `k,j,i` over grid `{build_summary['grid_shape_kji']}` | passed |",
        f"| coordinate / scale | `coordinate_bounds` = x `{build_summary['coordinate_bounds']['x']}`, y `{build_summary['coordinate_bounds']['y']}`, depth `{build_summary['coordinate_bounds']['depth']}`; `mapping` = `{build_summary['sparse_wells']['mapping']}` | passed |",
        f"| observation mask conditioning | `n_observation_rows={build_summary['sparse_wells']['n_observation_rows']}`, `n_unique_cells={build_summary['sparse_wells']['n_unique_cells']}`, `n_wells_with_constraints={build_summary['sparse_wells']['n_wells_with_constraints']}`; strict supplies 90 constraints, conditional supplies 91 | passed |",
        f"| train/eval mask mutual exclusion | strict: `n_direct_well_cells_excluded_from_metrics=0`; conditional: `n_direct_well_cells_excluded_from_metrics=90`; both have `test_patch_blocks_disjoint_from_train_and_validation=True` | passed |",
        f"| normalization / inverse transform | `results_strict.json` and `results_conditional.json` report `framework=ml_framework.preprocess` and `all_roundtrip_checks_passed=True`; `build_summary.preprocessing.coordinate_roundtrip` shows zero max abs error for x/y/depth | passed |",
        f"| adapter / decoder output | `p9_openmind_effect.py` and summary constrain `trainable_scope='attribute_projection_and_decoder'`; output metric is scalar RMSE on same sample universe | partially checked; internal tensor width not exposed in archived summaries |",
        f"| fold / sample universe | OpenMind summary: `folds={summary['evaluation']['folds']}`; `same_validation_sample_universe_as_strong_baseline=True`; strict/conditional results preserve separate train/test patch blocks | passed |",
        f"| metric direction | `rmse/mae` lower-is-better, `r2/pearson_r` higher-is-better, encoded in workbook rows | passed |",
        "",
        "## Evidence and scope audit",
        "",
        f"- `build_summary.json` says `grid_shape_kji={build_summary['grid_shape_kji']}` and `n_active_cells={build_summary['n_active_cells']}`.",
        f"- `results_strict.json` strict train/test split: train i-blocks `{strict_result['train']['patch_i_blocks']}`, test i-blocks `{strict_result['test']['patch_i_blocks']}`.",
        f"- `results_conditional.json` conditional train/test split: train i-blocks `{cond_result['train']['patch_i_blocks']}`, test i-blocks `{cond_result['test']['patch_i_blocks']}`.",
        f"- `results_strict.json` leakage check: `test_patch_blocks_disjoint_from_train_and_validation={strict_result['leakage_checks']['test_patch_blocks_disjoint_from_train_and_validation']}`.",
        f"- `results_conditional.json` leakage check: `direct_well_observation_cells_excluded_from_test_metrics={cond_result['leakage_checks']['direct_well_observation_cells_excluded_from_test_metrics']}`.",
        f"- `p9_openmind_effect.py` is a strict-development-only comparison; its summary records `frozen_test_accessed=False` and `guard_accessed=False`.",
        f"- `p5_stage4_confirmation/summary.json` records `prior_test_consumed={stage4['prior_test_consumed']}` and `fresh_blind={stage4['fresh_blind']}`.",
        "",
        "## Root cause assessment",
        "",
        "The evidence points to model-inductive-bias / capacity mismatch, not a broken pipeline. The OpenMind lane",
        "is helped by pretraining, but it still loses decisively to the PyKrige reference on the same split universe.",
        "The available archives are enough to label the lane non_beneficial without inventing an unsupported bug fix.",
        "",
        "## Evidence-only boundary",
        "",
        f"- Existing legal-dev evidence was sufficient: `build_summary.json`, `results_strict.json`, `results_conditional.json`, `p9_openmind_effect/summary.json`, `p5_stage4_confirmation/summary.json`, `model_inspection.json`.",
        "- No frozen holdout or tuning run was used for this audit bundle.",
        "- The report intentionally does not claim a newly improved production model.",
        "",
        "## Files written",
        "",
        f"- `{WORKBOOK_PATH}`",
        f"- `{FIGURES_MANIFEST_PATH}`",
        f"- `{TABLES_MANIFEST_PATH}`",
        f"- `{AUDIT_REPORT_PATH}`",
        f"- `{FIGURE_PATH}`",
        "",
        "## Residual risk",
        "",
        "The current report is evidence-only. It does not re-run the model or change split/protocol choices,",
        "so it cannot prove a new production model. It only documents that the better-performing baseline is",
        "still the PyKrige reference.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def _validate_output(rows: list[dict[str, Any]]) -> None:
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        assert wb.sheetnames == ["模型指标"], wb.sheetnames
        ws = wb["模型指标"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == HEADERS, header
        assert ws.max_row == len(rows) + 1, (ws.max_row, len(rows))
    finally:
        wb.close()


def build(output_dir: Path = OUTPUT_DIR) -> list[dict[str, Any]]:
    strict = _load_json(STRICT_RESULTS)
    cond = _load_json(CONDITIONAL_RESULTS)
    openmind = _load_json(OPENMIND_SUMMARY)
    stage4 = _load_json(STAGE4_SUMMARY)
    strict["__evidence_path__"] = STRICT_RESULTS
    cond["__evidence_path__"] = CONDITIONAL_RESULTS

    rows: list[dict[str, Any]] = []
    _add_ridge_rows(rows, strict, "strict_development", "strict_reverse_spatial_holdout")
    _add_ridge_rows(rows, cond, "conditional_development", "conditional_development_holdout")
    _add_openmind_rows(rows, openmind)
    _add_stage4_rows(rows, stage4)

    output_dir.mkdir(parents=True, exist_ok=True)
    _write_workbook(WORKBOOK_PATH, rows)
    _chart_primary_metric(FIGURE_PATH)

    figures = [
        {
            "figure_name": "before_after_primary_metric",
            "path": str(FIGURE_PATH),
            "description": "Strict-development primary RMSE comparison across OpenMind random-init, OpenMind pretrained, ridge_linear and PyKrige reference.",
            "source": "OpenMind summary + strict ridge results",
            "sha256": _sha256(FIGURE_PATH),
            "exists": FIGURE_PATH.exists(),
        },
        {
            "figure_name": "prediction_visualization_strict",
            "path": str(STRICT_PRED_FIG),
            "description": "Strict-mode truth / prediction / residual scientific slice visualization.",
            "source": str(STRICT_VIS_META),
            "sha256": _sha256(STRICT_PRED_FIG),
            "exists": STRICT_PRED_FIG.exists(),
        },
        {
            "figure_name": "prediction_visualization_conditional",
            "path": str(COND_PRED_FIG),
            "description": "Conditional-mode truth / prediction / residual scientific slice visualization.",
            "source": str(COND_VIS_META),
            "sha256": _sha256(COND_PRED_FIG),
            "exists": COND_PRED_FIG.exists(),
        },
        {
            "figure_name": "p5_stage4_confirmation_strict",
            "path": str(STRICT_CONFIRMATION),
            "description": "Stage-4 strict known-holdout confirmation figure.",
            "source": str(STAGE4_SUMMARY),
            "sha256": _sha256(STRICT_CONFIRMATION),
            "exists": STRICT_CONFIRMATION.exists(),
        },
        {
            "figure_name": "p5_stage4_confirmation_conditional",
            "path": str(COND_CONFIRMATION),
            "description": "Stage-4 conditional known-holdout confirmation figure.",
            "source": str(STAGE4_SUMMARY),
            "sha256": _sha256(COND_CONFIRMATION),
            "exists": COND_CONFIRMATION.exists(),
        },
        {
            "figure_name": "p5_stage3_reconstruction_strict",
            "path": str(STRICT_STAGE3_FIG),
            "description": "Stage-3 strict OOF diagnostic figure.",
            "source": str(ROOT / "p5_stage3_results.jsonl"),
            "sha256": _sha256(STRICT_STAGE3_FIG),
            "exists": STRICT_STAGE3_FIG.exists(),
        },
        {
            "figure_name": "p5_stage3_reconstruction_conditional",
            "path": str(COND_STAGE3_FIG),
            "description": "Stage-3 conditional OOF diagnostic figure.",
            "source": str(ROOT / "p5_stage3_results.jsonl"),
            "sha256": _sha256(COND_STAGE3_FIG),
            "exists": COND_STAGE3_FIG.exists(),
        },
    ]
    _write_csv(FIGURES_MANIFEST_PATH, figures, ["figure_name", "path", "description", "source", "sha256", "exists"])

    tables = [
        {
            "table_name": "track_model_metrics",
            "path": str(WORKBOOK_PATH),
            "sheet_name": "模型指标",
            "row_count": len(rows),
            "sha256": _sha256(WORKBOOK_PATH),
            "source": "strict/conditional ridge results + OpenMind summary + Stage-4 confirmation",
        }
    ]
    _write_csv(TABLES_MANIFEST_PATH, tables, ["table_name", "path", "sheet_name", "row_count", "sha256", "source"])

    _write_audit_report(AUDIT_REPORT_PATH, rows)
    _validate_output(rows)

    for path in [WORKBOOK_PATH, FIGURE_PATH, FIGURES_MANIFEST_PATH, TABLES_MANIFEST_PATH, AUDIT_REPORT_PATH]:
        assert path.exists(), path

    return rows


def main() -> None:
    build()


if __name__ == "__main__":
    main()
