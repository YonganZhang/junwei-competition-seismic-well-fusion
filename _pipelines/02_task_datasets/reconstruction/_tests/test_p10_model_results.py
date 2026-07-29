from __future__ import annotations

import csv
import importlib.util
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location(
    "build_p10_model_results", ROOT / "build_p10_model_results.py"
)
assert SPEC and SPEC.loader
p10 = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(p10)


class TestP10ModelResults(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.rows = p10.build()

    def test_workbook_has_single_model_metrics_sheet(self) -> None:
        wb = load_workbook(p10.WORKBOOK_PATH, read_only=True, data_only=True)
        try:
            self.assertEqual(wb.sheetnames, ["model metrics"])
            ws = wb["model metrics"]
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(header, p10.HEADERS)
            self.assertEqual(ws.max_row, len(self.rows) + 1)
        finally:
            wb.close()

    def test_manifests_and_evidence_paths_exist(self) -> None:
        self.assertTrue(p10.FIGURES_MANIFEST_PATH.exists())
        self.assertTrue(p10.TABLES_MANIFEST_PATH.exists())
        self.assertTrue(p10.AUDIT_REPORT_PATH.exists())
        self.assertTrue(p10.FIGURE_PATH.exists())
        self.assertGreater(p10.FIGURE_PATH.stat().st_size, 0)

        with p10.FIGURES_MANIFEST_PATH.open(newline="", encoding="utf-8") as fh:
            figure_rows = list(csv.DictReader(fh))
        with p10.TABLES_MANIFEST_PATH.open(newline="", encoding="utf-8") as fh:
            table_rows = list(csv.DictReader(fh))

        self.assertGreaterEqual(len(figure_rows), 1)
        self.assertEqual(len(table_rows), 1)
        self.assertEqual(int(table_rows[0]["row_count"]), len(self.rows))

        for row in self.rows:
            evidence_path = row["evidence_path"]
            if evidence_path:
                self.assertTrue(Path(evidence_path).exists(), evidence_path)
            checkpoint_path = row["checkpoint_path"]
            if checkpoint_path:
                self.assertTrue(Path(checkpoint_path).exists(), checkpoint_path)

    def test_primary_metric_rows_cover_expected_sources(self) -> None:
        names = {(row["dataset"], row["model_name"], row["metric_name"]) for row in self.rows}
        self.assertIn(("strict_development", "ridge_idw_seismic_coordinates", "rmse"), names)
        self.assertIn(("conditional_development", "ridge_idw_seismic_coordinates", "rmse"), names)
        self.assertIn(("strict_development", "train_mean", "rmse"), names)
        self.assertIn(("strict_development", "MIC-DKFZ/ResEncL-OpenMind-MAE", "rmse"), names)
        self.assertIn(("strict_development", "pykrige_ok3d", "rmse"), names)
        self.assertIn(("strict_confirmation", "pykrige_ok3d", "rmse"), names)
        self.assertIn(("conditional_confirmation", "pykrige_ok3d", "rmse"), names)


if __name__ == "__main__":
    unittest.main()
