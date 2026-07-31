"""Shared, dependency-light readers for the five root-owned sweetspot targets.

The builders read the released Volve archives without extracting them into the
repository.  Petrophysical labels are joined to raw/input curves on measured
depth.  Production examples are constructed causally: every feature comes from
the history ending before the prediction cutoff, while labels come from the
future horizon.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping
from zipfile import ZipFile

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from ..build_dataset import PROJECT_ROOT, resolve_source_root


ROOT_SEED = 2693
PETRO_ARCHIVE = Path("_sandbox/volve_data/Volve_Well_logs.zip")
PRODUCTION_ARCHIVE = Path("_sandbox/volve_data/Volve_Production_data.zip")
PRODUCTION_MEMBER = "Production_data/Volve production data.xlsx"
PETRO_PREFIX = "Well_logs/05.PETROPHYSICAL INTERPRETATION/"
RAW_LOG_FEATURES = (
    "GR", "RHOB", "NPHI", "RT", "DT", "DTS", "PEF", "DRHO", "CALI", "BS",
    "RD", "RM", "RACEHM", "RACELM", "RPCEHM", "RPCELM",
)
LABEL_ONLY_FIELDS = (
    "PHIF", "PHIE", "KLOGH", "KLOGV", "SW", "VSH", "SAND_FLAG", "PERF_FLAG",
)
PRODUCER_GROUPS = (
    "NO 15/9-F-1 C", "NO 15/9-F-11 H", "NO 15/9-F-12 H", "NO 15/9-F-14 H",
    "NO 15/9-F-15 D",
)


@dataclass(frozen=True)
class LasTable:
    member: str
    curves: Mapping[str, np.ndarray]
    null_value: float


def _source_root(explicit: Path | None = None) -> Path:
    return resolve_source_root(PROJECT_ROOT, explicit)


def _curve_names_and_null(text: str) -> tuple[list[str], float]:
    curves: list[str] = []
    in_curve = False
    null_value = -999.25
    for raw in text.splitlines():
        line = raw.strip()
        upper = line.upper()
        if upper.startswith("NULL."):
            match = re.search(r"NULL\.\S*\s+([-+0-9.Ee]+)", line, flags=re.IGNORECASE)
            if match:
                null_value = float(match.group(1))
        if upper.startswith("~CURVE"):
            in_curve = True
            continue
        if in_curve and line.startswith("~"):
            break
        if not in_curve or not line or line.startswith("#"):
            continue
        mnemonic = line.split(".", 1)[0].strip().split()[0]
        if mnemonic:
            curves.append(mnemonic.upper())
    return curves, null_value


def parse_las_bytes(payload: bytes, member: str) -> LasTable:
    """Parse an unwrapped LAS 2.0 table and replace its declared null value."""
    text = payload.decode("latin-1", errors="replace")
    names, null_value = _curve_names_and_null(text)
    if not names:
        raise ValueError(f"{member}: no LAS curves found")
    data_lines: list[str] = []
    in_data = False
    for raw in text.splitlines():
        line = raw.strip()
        if line.upper().startswith("~A"):
            in_data = True
            continue
        if in_data and line and not line.startswith("#"):
            data_lines.append(line)
    if not data_lines:
        raise ValueError(f"{member}: empty LAS data section")
    rows = np.fromstring(" ".join(data_lines), sep=" ", dtype=np.float64)
    if rows.size % len(names):
        raise ValueError(f"{member}: {rows.size} values cannot form {len(names)} columns")
    matrix = rows.reshape(-1, len(names))
    matrix[np.isclose(matrix, null_value)] = np.nan
    return LasTable(member, {name: matrix[:, i] for i, name in enumerate(names)}, null_value)


def well_family(well: str) -> str:
    normalized = well.upper().replace("_", "/").replace(" ", "")
    if "15/9-19" in normalized:
        return "15/9-19"
    match = re.search(r"15/9-F-(\d+)", normalized)
    if match:
        return f"15/9-F-{int(match.group(1))}"
    raise ValueError(f"cannot derive mother-well family from {well!r}")


def _well_folder(member: str) -> str:
    return member[len(PETRO_PREFIX):].split("/", 1)[0]


def _member_crc_evidence(archive: ZipFile, members: Iterable[str]) -> list[dict[str, Any]]:
    evidence = []
    for member in sorted(set(members)):
        info = archive.getinfo(member)
        evidence.append({
            "member": member,
            "crc32": f"{info.CRC:08x}",
            "size_bytes": info.file_size,
        })
    return evidence


def load_petrophysical_tables(source_root: Path | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Return depth-aligned raw-log inputs and interpreted label curves.

    CPI-only 15/9-19 files are retained in the availability evidence but are
    excluded from supervised rows because they have no paired raw-input LAS in
    this archive.  This prevents interpreted label curves from silently becoming
    model inputs.
    """
    root = _source_root(source_root)
    archive_path = root / PETRO_ARCHIVE
    records: list[dict[str, Any]] = []
    used_members: list[str] = []
    cpi_only: list[str] = []
    with ZipFile(archive_path) as archive:
        names = [
            name for name in archive.namelist()
            if name.startswith(PETRO_PREFIX) and name.lower().endswith(".las")
        ]
        outputs = [
            name for name in names
            if "COMPUTED_OUTPUT" in name.upper() or "/CPI/" in name.upper()
        ]
        by_folder: dict[str, list[str]] = {}
        for name in names:
            by_folder.setdefault(_well_folder(name), []).append(name)
        for output_member in sorted(outputs):
            output = parse_las_bytes(archive.read(output_member), output_member)
            if not ({"PHIF", "KLOGH"} <= set(output.curves)):
                continue
            folder = _well_folder(output_member)
            input_candidates = [
                name for name in by_folder.get(folder, [])
                if "COMPUTED_INPUT" in name.upper()
            ]
            if not input_candidates:
                cpi_only.append(output_member)
                continue
            input_member = sorted(input_candidates)[0]
            inputs = parse_las_bytes(archive.read(input_member), input_member)
            if "DEPTH" not in output.curves or "DEPTH" not in inputs.curves:
                raise ValueError(f"{folder}: DEPTH is required for leakage-safe join")
            input_index = {round(float(depth), 4): i for i, depth in enumerate(inputs.curves["DEPTH"]) if np.isfinite(depth)}
            output_indices: list[int] = []
            input_indices: list[int] = []
            for oi, depth in enumerate(output.curves["DEPTH"]):
                if not np.isfinite(depth):
                    continue
                ii = input_index.get(round(float(depth), 4))
                if ii is not None:
                    output_indices.append(oi)
                    input_indices.append(ii)
            if not output_indices:
                raise ValueError(f"{folder}: raw/input and output LAS have no common depths")
            features = {
                field: np.asarray(inputs.curves.get(field, np.full(len(input_indices), np.nan)))[input_indices]
                if field in inputs.curves else np.full(len(input_indices), np.nan)
                for field in RAW_LOG_FEATURES
            }
            labels = {
                field: np.asarray(output.curves[field])[output_indices]
                for field in LABEL_ONLY_FIELDS if field in output.curves
            }
            # F-12 carries SAND_FLAG in the computed-input file.  It may only be
            # used as a label and is never copied into the feature dictionary.
            if "SAND_FLAG" not in labels and "SAND_FLAG" in inputs.curves:
                labels["SAND_FLAG"] = np.asarray(inputs.curves["SAND_FLAG"])[input_indices]
            records.append({
                "wellbore": folder.replace("_", "/"),
                "well_family": well_family(folder),
                "depth_m": np.asarray(output.curves["DEPTH"])[output_indices],
                "features": features,
                "labels": labels,
                "input_member": input_member,
                "label_member": output_member,
            })
            used_members.extend((input_member, output_member))
        evidence = {
            "archive": PETRO_ARCHIVE.as_posix(),
            "archive_size_bytes": archive_path.stat().st_size,
            "paired_wellbores": len(records),
            "paired_families": sorted({record["well_family"] for record in records}),
            "cpi_only_excluded": sorted(cpi_only),
            "members": _member_crc_evidence(archive, used_members + cpi_only),
        }
    return records, evidence


def flatten_petrophysical_target(
    tables: Iterable[Mapping[str, Any]],
    *,
    target_name: str,
    target_fn: Any,
) -> dict[str, Any]:
    """Flatten aligned well tables while preserving family and missing masks."""
    sample_ids: list[str] = []
    groups: list[str] = []
    depths: list[float] = []
    targets: list[float] = []
    features: list[list[float]] = []
    wellbores: list[str] = []
    for table in tables:
        label = np.asarray(target_fn(table["labels"]), dtype=float)
        depth = np.asarray(table["depth_m"], dtype=float)
        matrix = np.column_stack([np.asarray(table["features"][field], dtype=float) for field in RAW_LOG_FEATURES])
        valid = np.isfinite(label) & np.isfinite(depth) & np.isfinite(matrix).any(axis=1)
        for index in np.flatnonzero(valid):
            well_token = re.sub(r"[^A-Za-z0-9]+", "-", str(table["wellbore"])).strip("-")
            sample_ids.append(f"{target_name}:{well_token}:{depth[index]:.4f}")
            groups.append(str(table["well_family"]))
            depths.append(float(depth[index]))
            targets.append(float(label[index]))
            features.append(matrix[index].tolist())
            wellbores.append(str(table["wellbore"]))
    if not sample_ids:
        raise RuntimeError(f"no valid samples for {target_name}")
    return {
        "sample_ids": sample_ids,
        "groups": groups,
        "wellbores": wellbores,
        "depth_m": depths,
        "feature_names": list(RAW_LOG_FEATURES),
        "features": features,
        "target": targets,
    }


def load_daily_production(source_root: Path | None = None) -> tuple[pd.DataFrame, dict[str, Any]]:
    root = _source_root(source_root)
    archive_path = root / PRODUCTION_ARCHIVE
    with ZipFile(archive_path) as archive:
        payload = archive.read(PRODUCTION_MEMBER)
        member_evidence = _member_crc_evidence(archive, [PRODUCTION_MEMBER])[0]
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook["Daily Production Data"]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        frame = pd.DataFrame.from_records(rows, columns=headers)
    finally:
        workbook.close()
    frame["DATEPRD"] = pd.to_datetime(frame["DATEPRD"], errors="coerce")
    numeric = [
        "ON_STREAM_HRS", "AVG_DOWNHOLE_PRESSURE", "AVG_DOWNHOLE_TEMPERATURE",
        "AVG_DP_TUBING", "AVG_ANNULUS_PRESS", "AVG_CHOKE_SIZE_P", "AVG_WHP_P",
        "AVG_WHT_P", "DP_CHOKE_SIZE", "BORE_OIL_VOL", "BORE_GAS_VOL",
        "BORE_WAT_VOL", "BORE_WI_VOL",
    ]
    for column in numeric:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame = frame[frame["WELL_BORE_CODE"].isin(PRODUCER_GROUPS)].copy()
    frame.sort_values(["WELL_BORE_CODE", "DATEPRD"], inplace=True)
    evidence = {
        "archive": PRODUCTION_ARCHIVE.as_posix(),
        "archive_size_bytes": archive_path.stat().st_size,
        "member": member_evidence,
        "rows_selected": int(len(frame)),
        "well_groups": sorted(frame["WELL_BORE_CODE"].dropna().unique().tolist()),
        "date_min": frame["DATEPRD"].min().date().isoformat(),
        "date_max": frame["DATEPRD"].max().date().isoformat(),
    }
    return frame, evidence


def _history_features(history: pd.DataFrame, *, history_days: int) -> tuple[list[str], list[float]]:
    columns = (
        "BORE_OIL_VOL", "BORE_GAS_VOL", "BORE_WAT_VOL", "ON_STREAM_HRS",
        "AVG_DOWNHOLE_PRESSURE", "AVG_CHOKE_SIZE_P", "AVG_WHP_P",
    )
    names: list[str] = []
    values: list[float] = []
    for column in columns:
        series = history[column].astype(float)
        for suffix, value in (
            ("mean", series.mean()), ("std", series.std(ddof=0)), ("last", series.iloc[-1]),
        ):
            names.append(f"history_{history_days}d_{column.lower()}_{suffix}")
            values.append(float(value) if np.isfinite(value) else np.nan)
    return names, values


def build_productivity_dataset(source_root: Path | None = None) -> tuple[dict[str, Any], dict[str, Any]]:
    """Build weekly-cutoff examples for future 30-day mean oil production."""
    frame, evidence = load_daily_production(source_root)
    sample_ids: list[str] = []
    groups: list[str] = []
    cutoffs: list[str] = []
    targets: list[float] = []
    features: list[list[float]] = []
    feature_names: list[str] | None = None
    for group, well in frame.groupby("WELL_BORE_CODE", sort=True):
        well = well.sort_values("DATEPRD").reset_index(drop=True)
        for index in range(30, len(well) - 29, 7):
            history = well.iloc[index - 30:index]
            future = well.iloc[index:index + 30]
            if (history["DATEPRD"].iloc[-1] - history["DATEPRD"].iloc[0]).days > 45:
                continue
            if (future["DATEPRD"].iloc[-1] - future["DATEPRD"].iloc[0]).days > 45:
                continue
            target = future["BORE_OIL_VOL"].mean()
            if not np.isfinite(target) or future["BORE_OIL_VOL"].notna().sum() < 21:
                continue
            names, row = _history_features(history, history_days=30)
            if not np.isfinite(row).any():
                continue
            cutoff = well.loc[index, "DATEPRD"]
            sample_ids.append(f"productivity:{group}:{cutoff.date().isoformat()}")
            groups.append(str(group))
            cutoffs.append(cutoff.date().isoformat())
            targets.append(float(target))
            features.append(row)
            feature_names = names
    if not sample_ids:
        raise RuntimeError("no causal productivity examples were constructed")
    dataset = {
        "sample_ids": sample_ids, "groups": groups, "cutoff_dates": cutoffs,
        "feature_names": feature_names, "features": features,
        "target_future_30d_mean_oil_sm3_day": targets,
    }
    evidence["constructed_samples"] = len(sample_ids)
    evidence["samples_by_group"] = {group: groups.count(group) for group in sorted(set(groups))}
    return dataset, evidence


def _first_sustained_positive(values: np.ndarray, days: int = 7) -> int | None:
    positive = np.isfinite(values) & (values > 0.0)
    if len(positive) < days:
        return None
    run = np.convolve(positive.astype(int), np.ones(days, dtype=int), mode="valid")
    positions = np.flatnonzero(run == days)
    return int(positions[0]) if positions.size else None


def build_water_breakthrough_dataset(source_root: Path | None = None) -> tuple[dict[str, Any], dict[str, Any]]:
    """Build pre-event 30-day water-risk examples with explicit censoring rules."""
    frame, evidence = load_daily_production(source_root)
    sample_ids: list[str] = []
    groups: list[str] = []
    cutoffs: list[str] = []
    labels: list[int] = []
    features: list[list[float]] = []
    feature_names: list[str] | None = None
    events: dict[str, Any] = {}
    exclusions: dict[str, str] = {}
    for group, well in frame.groupby("WELL_BORE_CODE", sort=True):
        well = well.sort_values("DATEPRD").reset_index(drop=True)
        event_index = _first_sustained_positive(well["BORE_WAT_VOL"].to_numpy(float), days=7)
        history_days = 7
        if event_index is not None and event_index < history_days:
            exclusions[str(group)] = "left-truncated: sustained water begins before 7-day history exists"
            continue
        event_date = None if event_index is None else well.loc[event_index, "DATEPRD"]
        events[str(group)] = None if event_date is None else event_date.date().isoformat()
        last_index = len(well) if event_index is None else event_index
        for index in range(history_days, last_index, 7):
            history = well.iloc[index - history_days:index]
            if (history["DATEPRD"].iloc[-1] - history["DATEPRD"].iloc[0]).days > 45:
                continue
            names, row = _history_features(history, history_days=history_days)
            if not np.isfinite(row).any():
                continue
            label = int(event_index is not None and 1 <= event_index - index <= 30)
            cutoff = well.loc[index, "DATEPRD"]
            sample_ids.append(f"water-risk:{group}:{cutoff.date().isoformat()}")
            groups.append(str(group))
            cutoffs.append(cutoff.date().isoformat())
            labels.append(label)
            features.append(row)
            feature_names = names
    if not sample_ids or len(set(labels)) < 2:
        raise RuntimeError("water-risk builder needs both positive and negative pre-event samples")
    dataset = {
        "sample_ids": sample_ids, "groups": groups, "cutoff_dates": cutoffs,
        "feature_names": feature_names, "features": features,
        "event_within_30d": labels,
    }
    evidence.update({
        "event_definition": "first run of 7 consecutive reported days with BORE_WAT_VOL > 0",
        "history_days": 7,
        "events": events,
        "excluded_groups": exclusions,
        "constructed_samples": len(sample_ids),
        "positive_samples": int(sum(labels)),
        "samples_by_group": {group: groups.count(group) for group in sorted(set(groups))},
    })
    return dataset, evidence


def stable_payload_hash(payload: Mapping[str, Any]) -> str:
    import json

    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()
