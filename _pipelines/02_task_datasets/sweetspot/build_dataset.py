"""Sweetspot track data-readiness audit and label-contract gate.

This file intentionally has no dataset-building mode. It never imports the shared
dataset writer and cannot create train/test HDF5 files. Until an approved label
contract passes every gate, the only supported operations are:

* ``audit``: inspect real Layer1, LAS and production fields and write reports.
* ``validate-only``: run the same audit, then validate an approved label spec.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZipFile

import h5py
import numpy as np
import yaml
from jsonschema import Draft202012Validator
from openpyxl import load_workbook


TRACK_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = TRACK_DIR.parents[2]
DEFAULT_SCHEMA_PATH = TRACK_DIR / "label_spec.schema.v1.json"
DEFAULT_REPORT_DIR = TRACK_DIR / "audit"
FORBIDDEN_DATASET_RELATIVE = Path("_data/processed/sweetspot")
UNRESOLVED_MARKERS = ("<required", "todo", "tbd", "待定", "待批准")


def resolve_source_root(project_root: Path, explicit_source_root: Path | None = None) -> Path:
    """Find the shared main checkout that owns the ignored raw data.

    A worktree does not necessarily contain ``_sandbox`` or ignored HDF5 files.
    The main checkout is derived from Git's ``commondir`` file without mutating
    Git state. Callers may override this with ``--source-root``.
    """

    candidates: list[Path] = []
    if explicit_source_root is not None:
        candidates.append(explicit_source_root.resolve())

    git_marker = project_root / ".git"
    if git_marker.is_file():
        line = git_marker.read_text(encoding="utf-8").strip()
        if line.startswith("gitdir:"):
            git_dir = Path(line.split(":", 1)[1].strip()).resolve()
            common_file = git_dir / "commondir"
            common_dir = (
                (git_dir / common_file.read_text(encoding="utf-8").strip()).resolve()
                if common_file.exists()
                else git_dir
            )
            candidates.append(common_dir.parent)

    # Prefer the shared main checkout over a worktree. ``_sandbox`` may be a
    # symlink inside the worktree while ignored Layer1 artifacts (notably the
    # cleaned well-log HDF5) exist only in the main checkout.
    candidates.append(project_root.resolve())

    for candidate in candidates:
        if (candidate / "_sandbox/volve_data").is_dir():
            return candidate
    checked = ", ".join(str(path) for path in candidates)
    raise FileNotFoundError(f"找不到共享 Volve 数据根目录；已检查: {checked}")


def _safe_relative(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return path.name


def _json_scalar(value: Any) -> Any:
    if isinstance(value, np.generic):
        return value.item()
    return value


def _array_stats(array: np.ndarray) -> dict[str, Any]:
    total = int(array.size)
    result: dict[str, Any] = {
        "shape": list(array.shape),
        "dtype": str(array.dtype),
        "n_values": total,
    }
    flat = array.reshape(-1)
    if array.dtype.kind in "fiu":
        finite = np.isfinite(flat)
        present = int(finite.sum())
        result.update(
            {
                "n_present": present,
                "n_missing": total - present,
                "coverage": round(present / total, 6) if total else None,
            }
        )
        if present:
            valid = flat[finite]
            result["min"] = _json_scalar(valid.min())
            result["max"] = _json_scalar(valid.max())
    else:
        present = sum(value is not None and str(value) != "" for value in flat)
        result.update(
            {
                "n_present": int(present),
                "n_missing": total - int(present),
                "coverage": round(present / total, 6) if total else None,
            }
        )
    return result


def _audit_npz(path: Path, source_id: str) -> tuple[dict[str, Any], list[str]]:
    if not path.exists():
        return {"present": False, "path": path.name}, []
    fields: dict[str, Any] = {}
    with np.load(path, allow_pickle=True) as archive:
        for key in archive.files:
            fields[key] = _array_stats(np.asarray(archive[key]))
    return {
        "present": True,
        "path": path.name,
        "size_bytes": path.stat().st_size,
        "source_id": source_id,
        "fields": fields,
    }, sorted(fields)


def _audit_clean_well_logs(
    path: Path, source_root: Path
) -> tuple[dict[str, Any], dict[str, list[str]]]:
    catalog = {
        "layer1.well_logs_clean": [],
        "layer1.well_logs_clean.clean": [],
        "layer1.well_logs_clean.norm": [],
    }
    if not path.exists():
        return {"present": False, "path": path.name}, catalog

    tracks: dict[str, Any] = {}
    presence: dict[str, Counter[str]] = {
        "clean": Counter(),
        "norm": Counter(),
    }
    with h5py.File(path, "r") as handle:
        for track_id in sorted(handle.keys()):
            group = handle[track_id]
            track: dict[str, Any] = {"attributes": {}, "datasets": {}}
            for key, value in group.attrs.items():
                if isinstance(value, np.ndarray):
                    value = value.tolist()
                elif isinstance(value, np.generic):
                    value = value.item()
                if key.endswith("_path") and isinstance(value, str):
                    value = _safe_relative(Path(value), source_root)
                track["attributes"][key] = value
            if "depth_grid_m" in group:
                track["datasets"]["depth_grid_m"] = _array_stats(group["depth_grid_m"][()])
            for family in ("clean", "norm"):
                if family not in group:
                    continue
                for curve in sorted(group[family].keys()):
                    presence[family][curve] += 1
                    track["datasets"][f"{family}/{curve}"] = _array_stats(
                        group[family][curve][()]
                    )
            tracks[track_id] = track

    n_tracks = len(tracks)
    catalog["layer1.well_logs_clean"] = ["depth_grid_m"]
    for family in ("clean", "norm"):
        catalog[f"layer1.well_logs_clean.{family}"] = sorted(presence[family])
    return {
        "present": True,
        "path": path.name,
        "size_bytes": path.stat().st_size,
        "n_tracks": n_tracks,
        "field_presence_across_tracks": {
            family: {
                curve: {
                    "tracks_present": count,
                    "tracks_total": n_tracks,
                    "coverage": round(count / n_tracks, 6) if n_tracks else None,
                }
                for curve, count in sorted(presence[family].items())
            }
            for family in ("clean", "norm")
        },
        "tracks": tracks,
    }, catalog


def _parse_las_curve_fields(path: Path) -> list[str]:
    fields: list[str] = []
    in_curve = False
    with path.open("r", encoding="latin-1", errors="replace") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            lowered = line.lower()
            if lowered.startswith("~curve"):
                in_curve = True
                continue
            if in_curve and line.startswith("~"):
                break
            if not in_curve or not line or line.startswith("#"):
                continue
            mnemonic = line.split(".", 1)[0].strip().split()[0]
            if mnemonic:
                fields.append(mnemonic)
    return fields


def _audit_las(source_root: Path) -> tuple[dict[str, Any], list[str]]:
    base = source_root / "_sandbox/volve_data/_extracted_welllogs/Well_logs/06.LFP"
    paths = sorted(base.glob("*/*.las")) + sorted(base.glob("*/*.LAS"))
    per_file: dict[str, Any] = {}
    presence: Counter[str] = Counter()
    for path in paths:
        fields = _parse_las_curve_fields(path)
        for field in set(fields):
            presence[field] += 1
        per_file[path.name] = {
            "track": path.parent.name,
            "n_curve_fields": len(fields),
            "fields": fields,
        }
    total = len(paths)
    return {
        "present": bool(paths),
        "path": "_sandbox/volve_data/_extracted_welllogs/Well_logs/06.LFP/*/*.las",
        "n_files": total,
        "coverage_kind": "field presence across LAS tracks; value missingness uses cleaned Layer1 HDF5",
        "field_presence_across_tracks": {
            field: {
                "tracks_present": count,
                "tracks_total": total,
                "coverage": round(count / total, 6) if total else None,
            }
            for field, count in sorted(presence.items())
        },
        "files": per_file,
    }, sorted(presence)


def _audit_production(source_root: Path) -> tuple[dict[str, Any], dict[str, list[str]]]:
    zip_path = source_root / "_sandbox/volve_data/Volve_Production_data.zip"
    catalog = {"production.daily": [], "production.monthly": []}
    if not zip_path.exists():
        return {"present": False, "path": zip_path.name}, catalog

    with ZipFile(zip_path) as archive:
        workbook_bytes = archive.read("Production_data/Volve production data.xlsx")
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheets: dict[str, Any] = {}
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = [str(value).strip() if value is not None else f"column_{i + 1}" for i, value in enumerate(header_row)]
            non_null = [0] * len(headers)
            n_rows = 0
            for row in rows:
                if not any(value is not None for value in row):
                    continue
                n_rows += 1
                for i, value in enumerate(row[: len(headers)]):
                    if value is not None and str(value).strip().upper() != "NULL":
                        non_null[i] += 1
            columns = {
                header: {
                    "n_present": non_null[i],
                    "n_missing": n_rows - non_null[i],
                    "coverage": round(non_null[i] / n_rows, 6) if n_rows else None,
                }
                for i, header in enumerate(headers)
            }
            sheets[sheet.title] = {"n_data_rows": n_rows, "columns": columns}
            if sheet.title == "Daily Production Data":
                catalog["production.daily"] = headers
            elif sheet.title == "Monthly Production Data":
                catalog["production.monthly"] = headers
    finally:
        workbook.close()
    return {
        "present": True,
        "path": "_sandbox/volve_data/Volve_Production_data.zip::Production_data/Volve production data.xlsx",
        "size_bytes": zip_path.stat().st_size,
        "sheets": sheets,
    }, catalog


def _fraction_in_bounds(values: np.ndarray, lower: float, upper: float) -> float | None:
    finite = values[np.isfinite(values)]
    if not finite.size:
        return None
    return round(float(((finite >= lower) & (finite <= upper)).mean()), 6)


def _coordinate_alignment(layer1_dir: Path) -> dict[str, Any]:
    seismic_path = layer1_dir / "seismic_index.npz"
    if not seismic_path.exists():
        return {"status": "blocked", "reason": "seismic_index.npz missing"}
    with np.load(seismic_path) as seismic:
        bounds = {
            "inline": [int(seismic["il_min"]), int(seismic["il_max"])],
            "crossline": [int(seismic["xl_min"]), int(seismic["xl_max"])],
            "twt_ms": [float(seismic["samples_ms"].min()), float(seismic["samples_ms"].max())],
        }

    datasets: dict[str, Any] = {}
    for name, filename in (
        ("fault_points", "fault_points.npz"),
        ("horizon_bcu_points", "horizon_bcu_points.npz"),
    ):
        path = layer1_dir / filename
        if not path.exists():
            datasets[name] = {"present": False}
            continue
        with np.load(path, allow_pickle=True) as archive:
            datasets[name] = {
                "present": True,
                "inline_in_grid_fraction": _fraction_in_bounds(
                    archive["inline"], *bounds["inline"]
                ),
                "crossline_in_grid_fraction": _fraction_in_bounds(
                    archive["crossline"], *bounds["crossline"]
                ),
                "twt_in_sample_range_fraction": _fraction_in_bounds(
                    archive["twt_ms"], *bounds["twt_ms"]
                ),
            }

    tie_path = layer1_dir / "well_tie_weak.npz"
    tracks: dict[str, Any] = {}
    if tie_path.exists():
        with np.load(tie_path) as archive:
            prefixes = sorted({key.split("__", 1)[0] for key in archive.files if "__" in key})
            for prefix in prefixes:
                tracks[prefix] = {
                    "inline_in_grid_fraction": _fraction_in_bounds(
                        archive[f"{prefix}__inline"], *bounds["inline"]
                    ),
                    "crossline_in_grid_fraction": _fraction_in_bounds(
                        archive[f"{prefix}__crossline"], *bounds["crossline"]
                    ),
                    "twt_in_sample_range_fraction": _fraction_in_bounds(
                        archive[f"{prefix}__twt_est_ms"], *bounds["twt_ms"]
                    ),
                }
    datasets["well_tie_weak"] = {"present": tie_path.exists(), "tracks": tracks}
    return {
        "status": "audited",
        "seismic_grid_bounds": bounds,
        "datasets": datasets,
        "warning": "well_tie_weak is an interpolation/extrapolation approximation, not exact well tie truth",
    }


def audit_sources(project_root: Path, source_root: Path) -> dict[str, Any]:
    """Inspect real source fields without constructing or inferring labels."""

    layer1_dir = source_root / "_pipelines/01_common_preprocess/outputs"
    artifacts: dict[str, Any] = {}
    catalog: dict[str, list[str]] = {}

    for source_id, filename in (
        ("layer1.seismic_index", "seismic_index.npz"),
        ("layer1.fault_points", "fault_points.npz"),
        ("layer1.horizon_bcu_points", "horizon_bcu_points.npz"),
    ):
        detail, fields = _audit_npz(layer1_dir / filename, source_id)
        detail["path"] = f"_pipelines/01_common_preprocess/outputs/{filename}"
        artifacts[source_id] = detail
        catalog[source_id] = fields

    tie_detail, tie_keys = _audit_npz(
        layer1_dir / "well_tie_weak.npz", "layer1.well_tie_weak"
    )
    tie_detail["path"] = "_pipelines/01_common_preprocess/outputs/well_tie_weak.npz"
    tie_suffixes = sorted({key.split("__", 1)[1] for key in tie_keys if "__" in key})
    artifacts["layer1.well_tie_weak"] = tie_detail
    catalog["layer1.well_tie_weak"] = tie_suffixes

    clean_detail, clean_catalog = _audit_clean_well_logs(
        layer1_dir / "well_logs_clean.h5", source_root
    )
    clean_detail["path"] = "_pipelines/01_common_preprocess/outputs/well_logs_clean.h5"
    artifacts["layer1.well_logs_clean"] = clean_detail
    catalog.update(clean_catalog)

    las_detail, las_fields = _audit_las(source_root)
    artifacts["las"] = las_detail
    catalog["las"] = las_fields

    production_detail, production_catalog = _audit_production(source_root)
    artifacts["production"] = production_detail
    catalog.update(production_catalog)

    catalog = {key: sorted(set(values)) for key, values in sorted(catalog.items())}
    direct_label_names = sorted(
        field
        for fields in catalog.values()
        for field in fields
        if "sweetspot" in field.lower() or "sweet_spot" in field.lower()
    )
    return {
        "audit_schema_version": "sweetspot-data-availability/v1",
        "mode": "field_inventory_only",
        "source_root_kind": "git-shared-main" if source_root.resolve() != project_root.resolve() else "current-checkout",
        "artifacts": artifacts,
        "field_catalog": catalog,
        "coordinate_alignment": _coordinate_alignment(layer1_dir),
        "label_readiness": {
            "sweetspot_truth_found": bool(direct_label_names),
            "direct_label_fields": direct_label_names,
            "candidate_evidence_only": [
                "fault and BCU interpretation points",
                "three-track weak well tie",
                "LFP petrophysical interpretation curves",
                "daily and monthly production tables",
            ],
            "decision_owner": "Junwei / designated domain expert",
            "hard_blocker": "No approved sweetspot label contract; evidence fields must not be combined into labels.",
        },
        "outputs_created": {
            "data_availability_report": True,
            "train_h5": False,
            "test_h5": False,
            "labels": False,
            "model": False,
            "checkpoint": False,
            "metrics": False,
        },
    }


def _coverage_summary(inventory: dict[str, Any]) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    clean = inventory["artifacts"]["layer1.well_logs_clean"]
    if clean.get("present"):
        for curve, info in clean["field_presence_across_tracks"].get("clean", {}).items():
            rows.append((f"Layer1 clean/{curve}", f"{info['tracks_present']}/{info['tracks_total']} tracks"))
    las = inventory["artifacts"]["las"]
    for curve in ("LFP_GR", "LFP_PHIE", "LFP_VSH", "LFP_OIL", "LFP_GAS"):
        info = las.get("field_presence_across_tracks", {}).get(curve)
        if info:
            rows.append((f"LAS {curve}", f"{info['tracks_present']}/{info['tracks_total']} tracks"))
    production = inventory["artifacts"]["production"]
    if production.get("present"):
        daily = production["sheets"].get("Daily Production Data", {})
        for field in ("WELL_BORE_CODE", "BORE_OIL_VOL", "BORE_GAS_VOL", "BORE_WAT_VOL"):
            info = daily.get("columns", {}).get(field)
            if info:
                rows.append((f"Production daily/{field}", f"{info['coverage']:.1%} non-null"))
    return rows


def render_readiness_report(inventory: dict[str, Any]) -> str:
    readiness = inventory["label_readiness"]
    lines = [
        "# Sweetspot data readiness audit",
        "",
        "> Validate-only evidence report. It does not define or generate labels.",
        "",
        "## Decision boundary",
        "",
        f"- Sweetspot truth found: **{str(readiness['sweetspot_truth_found']).lower()}**",
        f"- Decision owner: {readiness['decision_owner']}",
        f"- Hard blocker: {readiness['hard_blocker']}",
        "",
        "## Real source availability",
        "",
        "| Source | Present | Audited fields |",
        "|---|---:|---:|",
    ]
    catalog = inventory["field_catalog"]
    source_to_artifact = {
        "layer1.seismic_index": "layer1.seismic_index",
        "layer1.fault_points": "layer1.fault_points",
        "layer1.horizon_bcu_points": "layer1.horizon_bcu_points",
        "layer1.well_tie_weak": "layer1.well_tie_weak",
        "layer1.well_logs_clean.clean": "layer1.well_logs_clean",
        "las": "las",
        "production.daily": "production",
        "production.monthly": "production",
    }
    for source in source_to_artifact:
        artifact = inventory["artifacts"][source_to_artifact[source]]
        lines.append(f"| `{source}` | {artifact.get('present', False)} | {len(catalog.get(source, []))} |")

    lines.extend(["", "## Coverage and missingness", "", "| Field | Coverage |", "|---|---:|"])
    for field, coverage in _coverage_summary(inventory):
        lines.append(f"| `{field}` | {coverage} |")

    alignment = inventory["coordinate_alignment"]
    lines.extend(["", "## Coordinate alignment", ""])
    if alignment.get("status") != "audited":
        lines.append(f"- Blocked: {alignment.get('reason')}")
    else:
        for name, detail in alignment["datasets"].items():
            if name == "well_tie_weak":
                for track, track_detail in detail.get("tracks", {}).items():
                    lines.append(
                        f"- `{name}/{track}`: inline={track_detail['inline_in_grid_fraction']}, "
                        f"crossline={track_detail['crossline_in_grid_fraction']}, "
                        f"TWT={track_detail['twt_in_sample_range_fraction']} within seismic bounds."
                    )
            elif detail.get("present"):
                lines.append(
                    f"- `{name}`: inline={detail['inline_in_grid_fraction']}, "
                    f"crossline={detail['crossline_in_grid_fraction']}, "
                    f"TWT={detail['twt_in_sample_range_fraction']} within seismic bounds."
                )
        lines.append(f"- Warning: {alignment['warning']}")

    lines.extend(
        [
            "",
            "## Explicitly not produced",
            "",
            "- No sweetspot label or proxy label.",
            "- No `_data/processed/sweetspot` directory or train/test HDF5.",
            "- No model, checkpoint, training run, metric, or synthetic dataset.",
            "",
        ]
    )
    return "\n".join(lines)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def write_audit_reports(inventory: dict[str, Any], report_dir: Path) -> tuple[Path, Path]:
    report_dir.mkdir(parents=True, exist_ok=True)
    json_path = report_dir / "data_availability.json"
    markdown_path = report_dir / "data_readiness.md"
    _write_json(json_path, inventory)
    markdown_path.write_text(render_readiness_report(inventory), encoding="utf-8")
    return json_path, markdown_path


def load_spec(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"label spec 不存在: {path}")
    with path.open("r", encoding="utf-8") as handle:
        value = yaml.safe_load(handle)
    if not isinstance(value, dict):
        raise ValueError("label spec 顶层必须是 mapping/object")
    return value


def _path_string(path: Iterable[Any]) -> str:
    parts = [str(part) for part in path]
    return ".".join(parts) if parts else "<root>"


def _is_unresolved(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    return not text or any(marker in text for marker in UNRESOLVED_MARKERS)


def _collect_unresolved(value: Any, path: tuple[Any, ...] = ()) -> list[str]:
    unresolved: list[str] = []
    if isinstance(value, dict):
        for key, child in value.items():
            unresolved.extend(_collect_unresolved(child, (*path, key)))
    elif isinstance(value, list):
        for i, child in enumerate(value):
            unresolved.extend(_collect_unresolved(child, (*path, i)))
    elif isinstance(value, str) and _is_unresolved(value):
        unresolved.append(_path_string(path))
    return unresolved


def _field_pair(entry: dict[str, Any]) -> tuple[str, str]:
    return str(entry.get("source", "")), str(entry.get("field", ""))


def validate_label_spec(
    spec: dict[str, Any] | None,
    inventory: dict[str, Any],
    schema_path: Path = DEFAULT_SCHEMA_PATH,
) -> list[str]:
    """Return every contract error. An empty list is the only pass state."""

    if spec is None:
        return ["缺少 label spec；validate-only 禁止继续"]

    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    errors = [
        f"schema {_path_string(error.absolute_path)}: {error.message}"
        for error in sorted(
            Draft202012Validator(schema).iter_errors(spec),
            key=lambda item: list(item.absolute_path),
        )
    ]

    approval = spec.get("approval") if isinstance(spec.get("approval"), dict) else {}
    if approval.get("approved") is not True:
        errors.append("approval.approved=false；必须由军伟/指定领域专家批准")
    if spec.get("status") != "approved":
        errors.append("status 必须为 approved")
    for field in ("approved_by", "approved_role", "approved_at", "decision_record"):
        if _is_unresolved(approval.get(field)):
            errors.append(f"approval.{field} 未完整定义")

    unresolved = _collect_unresolved(spec)
    if unresolved:
        errors.append("合同仍含未决定占位符: " + ", ".join(unresolved))

    construction = (
        spec.get("label_construction")
        if isinstance(spec.get("label_construction"), dict)
        else {}
    )
    fit_domain = construction.get("fit_domain") if isinstance(construction.get("fit_domain"), dict) else {}
    formula_text = str(construction.get("formula", ""))
    if fit_domain.get("uses_test_statistics") is not False or fit_domain.get("statistics_scope") == "test":
        errors.append("禁止使用 test 统计拟合标签公式/阈值/权重")
    if re.search(r"test[_ -]?(mean|std|quantile|median|stat)|测试集.*统计", formula_text, re.I):
        errors.append("公式文本疑似引用 test 统计")
    if not construction.get("thresholds") and _is_unresolved(
        construction.get("thresholds_not_applicable_reason")
    ):
        errors.append("thresholds 为空时必须给出明确的 not-applicable 理由")
    if not construction.get("weights") and _is_unresolved(
        construction.get("weights_not_applicable_reason")
    ):
        errors.append("weights 为空时必须给出明确的 not-applicable 理由")

    class_rules = spec.get("class_rules") if isinstance(spec.get("class_rules"), dict) else {}
    if _is_unresolved(class_rules.get("negative")):
        errors.append("负样本规则未定义；不得把“非正样本”默认当负样本")
    if _is_unresolved(class_rules.get("positive")) or _is_unresolved(class_rules.get("unlabeled")):
        errors.append("正样本或未标注样本规则未完整定义")

    spatial = spec.get("spatial_scale") if isinstance(spec.get("spatial_scale"), dict) else {}
    for field in ("support", "coordinate_system", "vertical_domain", "resolution", "alignment_tolerance"):
        if _is_unresolved(spatial.get(field)):
            errors.append(f"空间尺度 spatial_scale.{field} 未定义")

    split = spec.get("split_strategy") if isinstance(spec.get("split_strategy"), dict) else {}
    split_fields = ("strategy", "group_key", "train_rule", "validation_rule", "test_rule")
    if any(_is_unresolved(split.get(field)) for field in split_fields):
        errors.append("井/空间 split 规则不完整")
    if split.get("fit_statistics_scope") != "train_only":
        errors.append("split.fit_statistics_scope 必须为 train_only")
    if not split.get("leakage_guards"):
        errors.append("split 必须定义至少一条 leakage guard")

    catalog = inventory.get("field_catalog", {})
    allowed_entries = spec.get("allowed_source_fields", [])
    formula_entries = construction.get("formula_field_refs", [])
    inference_entries = spec.get("inference_allowed_inputs", [])
    allowed_pairs = {
        _field_pair(entry) for entry in allowed_entries if isinstance(entry, dict)
    }
    for section, entries in (
        ("allowed_source_fields", allowed_entries),
        ("label_construction.formula_field_refs", formula_entries),
        ("inference_allowed_inputs", inference_entries),
    ):
        if not isinstance(entries, list):
            continue
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            source, field = _field_pair(entry)
            if source not in catalog or field not in catalog.get(source, []):
                errors.append(
                    f"{section}: 未知或不存在的真实字段 {source}.{field}"
                )
            if section != "allowed_source_fields" and (source, field) not in allowed_pairs:
                errors.append(f"{section}: {source}.{field} 未列入 allowed_source_fields")

    if inventory.get("label_readiness", {}).get("sweetspot_truth_found"):
        errors.append("审计异常：发现疑似 sweetspot 直接标签，必须人工核验后才能继续")
    return list(dict.fromkeys(errors))


def run_validate_only(
    spec: dict[str, Any] | None,
    inventory: dict[str, Any],
    project_root: Path,
    schema_path: Path = DEFAULT_SCHEMA_PATH,
    validation_report_path: Path | None = None,
) -> dict[str, Any]:
    """Validate the contract and prove that no dataset path was created."""

    forbidden_dir = project_root / FORBIDDEN_DATASET_RELATIVE
    existed_before = forbidden_dir.exists()
    errors = validate_label_spec(spec, inventory, schema_path)
    result = {
        "validation_schema_version": "sweetspot-contract-validation/v1",
        "valid": not errors,
        "spec_version": spec.get("spec_version") if isinstance(spec, dict) else None,
        "errors": errors,
        "dataset_write_attempted": False,
        "forbidden_dataset_dir_existed_before": existed_before,
        "forbidden_dataset_dir_exists_after": forbidden_dir.exists(),
        "outputs_created": {
            "train_h5": False,
            "test_h5": False,
            "labels": False,
            "model": False,
            "checkpoint": False,
            "metrics": False,
        },
    }
    if not existed_before and forbidden_dir.exists():
        raise RuntimeError("fail-closed 违约：validate-only 意外创建了 sweetspot 数据目录")
    if validation_report_path is not None:
        _write_json(validation_report_path, result)
    return result


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Sweetspot validate-only field audit and label-contract gate"
    )
    parser.add_argument("--mode", choices=("audit", "validate-only"), required=True)
    parser.add_argument("--spec", type=Path, help="Approved label spec YAML; required by validate-only")
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT)
    parser.add_argument("--source-root", type=Path)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    project_root = args.project_root.resolve()
    source_root = resolve_source_root(project_root, args.source_root)
    inventory = audit_sources(project_root, source_root)
    json_path, markdown_path = write_audit_reports(inventory, args.report_dir)

    if args.mode == "audit":
        print(
            json.dumps(
                {
                    "mode": "audit",
                    "data_availability": str(json_path),
                    "data_readiness": str(markdown_path),
                    "sweetspot_truth_found": inventory["label_readiness"]["sweetspot_truth_found"],
                    "labels_created": False,
                    "dataset_created": False,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    spec: dict[str, Any] | None = None
    load_error: str | None = None
    if args.spec is None:
        load_error = "缺少 label spec；validate-only 禁止继续"
    else:
        try:
            spec = load_spec(args.spec)
        except (FileNotFoundError, ValueError, yaml.YAMLError) as error:
            load_error = str(error)

    validation_path = args.report_dir / "contract_validation.json"
    result = run_validate_only(
        spec,
        inventory,
        project_root=project_root,
        schema_path=DEFAULT_SCHEMA_PATH,
        validation_report_path=validation_path,
    )
    if load_error and load_error not in result["errors"]:
        result["errors"].insert(0, load_error)
        result["valid"] = False
        _write_json(validation_path, result)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["valid"] else 2


if __name__ == "__main__":
    sys.exit(main())
