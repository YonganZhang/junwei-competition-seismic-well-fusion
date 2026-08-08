"""Build the P10 sweetspot model-results delivery package.

The package is a pure reporting/QA layer. It reads already-materialized target
artifacts, reconstructs the baseline/foundation comparison rows, and writes a
single-sheet Excel workbook plus companion manifests and an audit report.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import subprocess
import textwrap
from pathlib import Path
from typing import Any, Iterable

import matplotlib
import numpy as np

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[4]
SWEETSPOT_ROOT = REPO_ROOT / "_pipelines" / "02_task_datasets" / "sweetspot"
OUTPUT_DIR = SWEETSPOT_ROOT / "_outputs" / "p10_model_results"

P4_STAGE3 = SWEETSPOT_ROOT / "p5" / "_outputs" / "stage3_cv" / "leaderboards"
P5_STAGE4 = SWEETSPOT_ROOT / "p5" / "_outputs" / "stage4_confirmation" / "targets"
P7_T3 = SWEETSPOT_ROOT / "p7" / "_outputs" / "t3_chronos2_cv" / "summary.json"
P8_T3 = SWEETSPOT_ROOT / "p8" / "_outputs" / "t3_chronos2_calendar_cv" / "summary.json"
P7_RUNNER = __import__("_pipelines.02_task_datasets.sweetspot.p7.runner", fromlist=["dummy"])
P5_STAGE2_LABELS = __import__("_pipelines.02_task_datasets.sweetspot.p5.sweetspot_p5_stage2_labels", fromlist=["dummy"])
P5_STAGE2_DATA = __import__("_pipelines.02_task_datasets.sweetspot.p5.sweetspot_p5_stage2_data", fromlist=["dummy"])
# Host-local runtime resources. Defaults are this machine's original paths so archived
# runs and their recorded provenance are unchanged; override to relocate elsewhere.
SHARED_TORCH_PYTHON = Path(
    os.environ.get("VOLVE_P5_TORCH_PYTHON", "/mnt/data/yongan-admin-2/.cache/volve-p5/envs/torch-common/bin/python")
)
CHRONOS2_SNAPSHOT = Path(
    os.environ.get(
        "CHRONOS2_SNAPSHOT_DIR",
        "/mnt/data/yongan-admin-2/.cache/huggingface/hub/models--amazon--chronos-2/snapshots/29ec3766d36d6f73f0696f85560a422f50e8498c",
    )
)
CHRONOS2_CHECKPOINT = str(CHRONOS2_SNAPSHOT / "model.safetensors")

TARGETS = (
    "reservoir_quality",
    "hydrocarbon_pay",
    "productivity",
    "water_breakthrough",
    "remaining_oil_infill",
    "porosity",
    "permeability",
)

FIELDS = [
    "track",
    "dataset",
    "task_type",
    "model_name",
    "model_family",
    "is_foundation_model",
    "foundation_type",
    "integration_point",
    "fusion_method",
    "preprocess_version",
    "split_protocol",
    "seed_or_fold",
    "metric_name",
    "metric_value",
    "higher_is_better",
    "baseline_model",
    "baseline_value",
    "delta_abs",
    "delta_pct",
    "status",
    "evidence_path",
    "checkpoint_path",
    "code_commit",
    "root_cause",
    "fix_applied",
    "notes",
]


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _git_commit_for(path: Path) -> str:
    result = subprocess.run(
        ["git", "-C", str(REPO_ROOT), "log", "--follow", "--format=%H", "-n", "1", "--", str(path.relative_to(REPO_ROOT))],
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout.strip()


def _repo_path(path: str | Path) -> str:
    candidate = Path(path)
    if candidate.is_absolute():
        try:
            return str(candidate.relative_to(REPO_ROOT))
        except ValueError:
            return str(candidate)
    return str(candidate)


def _metric_direction(metric_name: str) -> bool:
    return metric_name.lower() not in {"mae", "rmse", "brier", "net_thickness_mae_m", "physical_mae", "log1p_mae"}


def _delta(metric_value: float | None, baseline_value: float | None, higher_is_better: bool) -> tuple[float | None, float | None]:
    if metric_value is None or baseline_value is None or any(math.isnan(value) for value in [metric_value, baseline_value] if isinstance(value, float)):
        return None, None
    if higher_is_better:
        delta = metric_value - baseline_value
    else:
        delta = baseline_value - metric_value
    pct = None if baseline_value == 0 else delta / abs(baseline_value)
    return float(delta), None if pct is None else float(pct)


def _row(**kwargs: Any) -> dict[str, Any]:
    row = {field: kwargs.get(field) for field in FIELDS}
    return row


def _t1_rows() -> list[dict[str, Any]]:
    status = _read_json(SWEETSPOT_ROOT / "targets" / "reservoir_quality" / "_outputs" / "baseline_v1" / "status.json")
    model_root = SWEETSPOT_ROOT / "targets" / "reservoir_quality" / "_outputs" / "baseline_v1"
    rows = []
    for split_key, evidence_rel, checkpoint_rel, split_name in [
        ("oof_metrics", model_root / "oof" / "metrics.json", model_root / "refit" / "checkpoint_best.pkl", "development_oof"),
        ("frozen_test_metrics", model_root / "frozen_test" / "metrics.json", model_root / "refit" / "checkpoint_best.pkl", "frozen_test"),
    ]:
        metrics = status[split_key]
        for metric_name in ("mae", "rmse", "spearman", "r2"):
            metric_value = metrics.get(metric_name)
            if metric_value is None:
                continue
            rows.append(_row(
                track="sweetspot",
                dataset="T1_reservoir_quality",
                task_type="regression",
                model_name="robust_linear",
                model_family="linear_regression",
                is_foundation_model=False,
                foundation_type="none",
                integration_point="none",
                fusion_method="none",
                preprocess_version="baseline_v1",
                split_protocol=split_name,
                seed_or_fold=split_name,
                metric_name=metric_name,
                metric_value=float(metric_value),
                higher_is_better=_metric_direction(metric_name),
                baseline_model="robust_linear",
                baseline_value=float(metric_value),
                delta_abs=0.0,
                delta_pct=0.0,
                status=status["status"],
                evidence_path=_repo_path(evidence_rel),
                checkpoint_path=_repo_path(checkpoint_rel),
                code_commit="4d54623",
                root_cause="baseline reference row",
                fix_applied="none",
                notes=status["proxy_warning"],
            ))
    return rows


def _t2_rows() -> list[dict[str, Any]]:
    status = _read_json(SWEETSPOT_ROOT / "targets" / "hydrocarbon_pay" / "_outputs" / "baseline_v1" / "status.json")
    model_root = SWEETSPOT_ROOT / "targets" / "hydrocarbon_pay" / "_outputs" / "baseline_v1"
    rows = []
    for split_key, evidence_rel, checkpoint_rel, split_name in [
        ("oof_metrics", model_root / "oof" / "metrics.json", model_root / "refit" / "checkpoint_best.pkl", "development_oof"),
        ("frozen_test_metrics", model_root / "frozen_test" / "metrics.json", model_root / "refit" / "checkpoint_best.pkl", "frozen_test"),
    ]:
        metrics = status[split_key]
        for metric_name in ("average_precision", "brier", "f1", "net_thickness_mae_m"):
            metric_value = metrics.get(metric_name)
            if metric_value is None:
                continue
            rows.append(_row(
                track="sweetspot",
                dataset="T2_hydrocarbon_pay",
                task_type="classification",
                model_name="logistic_classifier",
                model_family="linear_classifier",
                is_foundation_model=False,
                foundation_type="none",
                integration_point="none",
                fusion_method="none",
                preprocess_version="baseline_v1",
                split_protocol=split_name,
                seed_or_fold=split_name,
                metric_name=metric_name,
                metric_value=float(metric_value),
                higher_is_better=_metric_direction(metric_name),
                baseline_model="logistic_classifier",
                baseline_value=float(metric_value),
                delta_abs=0.0,
                delta_pct=0.0,
                status=status["status"],
                evidence_path=_repo_path(evidence_rel),
                checkpoint_path=_repo_path(checkpoint_rel),
                code_commit="4d54623",
                root_cause="baseline reference row",
                fix_applied="none",
                notes=status["proxy_warning"],
            ))
    return rows


def _t3_rows() -> list[dict[str, Any]]:
    leaderboard = _read_json(P4_STAGE3 / "T3.json")
    chronos = _read_json(P7_T3)
    rows = []
    baseline = next(entry for entry in leaderboard["entries"] if entry["model_id"] == "xgboost")
    foundation = chronos["methods"]["F1_chronos2_train_blend"]
    # Development folds: same split/same seed/fold evidence.
    for fold_baseline, fold_foundation in zip(baseline["fold_means"].items(), foundation["folds"]):
        fold_id = str(fold_foundation["fold_id"])
        baseline_value = float(fold_baseline[1])
        foundation_value = float(fold_foundation["metrics"]["mae"])
        delta_abs, delta_pct = _delta(foundation_value, baseline_value, higher_is_better=False)
        rows.append(_row(
            track="sweetspot",
            dataset="T3_productivity",
            task_type="regression",
            model_name="chronos2_train_blend",
            model_family="time_series_forecasting_foundation",
            is_foundation_model=True,
            foundation_type="Chronos-2",
            integration_point="causal_history_blend",
            fusion_method="convex_weight",
            preprocess_version="p7_chronos2_cv",
            split_protocol="p5_stage3_dev_cv",
            seed_or_fold=f"fold_{fold_id}",
            metric_name="mae",
            metric_value=foundation_value,
            higher_is_better=False,
            baseline_model="xgboost",
            baseline_value=baseline_value,
            delta_abs=delta_abs,
            delta_pct=delta_pct,
            status="promote",
            evidence_path=_repo_path(P7_T3),
            checkpoint_path=CHRONOS2_CHECKPOINT,
            code_commit="1bb0595",
            root_cause="chronos blend clears archived XGBoost on development folds",
            fix_applied="none",
            notes="selected foundation blend from fold-train only; no frozen holdout used for selection",
        ))
    # Macro row for the final selected blend.
    baseline_value = float(baseline["primary_mean"])
    foundation_value = float(foundation["macro_fold_mean"]["mae"])
    delta_abs, delta_pct = _delta(foundation_value, baseline_value, higher_is_better=False)
    rows.append(_row(
        track="sweetspot",
        dataset="T3_productivity",
        task_type="regression",
        model_name="chronos2_train_blend",
        model_family="time_series_forecasting_foundation",
        is_foundation_model=True,
        foundation_type="Chronos-2",
        integration_point="causal_history_blend",
        fusion_method="convex_weight",
        preprocess_version="p7_chronos2_cv",
        split_protocol="p5_stage3_dev_cv",
        seed_or_fold="macro_mean",
        metric_name="mae",
        metric_value=foundation_value,
        higher_is_better=False,
        baseline_model="xgboost",
        baseline_value=baseline_value,
        delta_abs=delta_abs,
        delta_pct=delta_pct,
        status="promote",
        evidence_path=_repo_path(P7_T3),
        checkpoint_path=CHRONOS2_CHECKPOINT,
        code_commit="1bb0595",
        root_cause="chronos blend clears archived XGBoost on development folds",
        fix_applied="none",
        notes="selected foundation blend from fold-train only; no frozen holdout used for selection",
    ))
    # Support row for the exact-calendar diagnostic.
    p8 = _read_json(P8_T3)
    history = p8["methods"]["B1_calendar_history_mean"]["macro_fold_mean"]["mae"]
    chronos_diag = p8["methods"]["F0_chronos2_calendar"]["macro_fold_mean"]["mae"]
    delta_abs, delta_pct = _delta(float(chronos_diag), float(history), higher_is_better=False)
    rows.append(_row(
        track="sweetspot",
        dataset="T3_productivity_calendar_diag",
        task_type="regression",
        model_name="chronos2_calendar_diag",
        model_family="time_series_forecasting_foundation",
        is_foundation_model=True,
        foundation_type="Chronos-2",
        integration_point="exact_calendar_30d_history",
        fusion_method="direct_forecast",
        preprocess_version="p8_calendar_v1",
        split_protocol="p5_stage3_locked_calendar",
        seed_or_fold="macro_mean",
        metric_name="mae",
        metric_value=float(chronos_diag),
        higher_is_better=False,
        baseline_model="calendar_history_mean",
        baseline_value=float(history),
        delta_abs=delta_abs,
        delta_pct=delta_pct,
        status="effect_supported_not_promoted",
        evidence_path=_repo_path(P8_T3),
        checkpoint_path=CHRONOS2_CHECKPOINT,
        code_commit="e229aa4",
        root_cause="calendar-grid diagnostic still lacks the preregistered same-architecture random-init control for promotion",
        fix_applied="none",
        notes="supports the time-series contract audit; not the promoted model row",
    ))
    return rows


def _ap_score(actual: np.ndarray, prediction: np.ndarray) -> float:
    from sklearn.metrics import average_precision_score

    actual = np.asarray(actual, dtype=np.float64).reshape(-1)
    prediction = np.asarray(prediction, dtype=np.float64).reshape(-1)
    if actual.shape != prediction.shape or not np.isfinite(prediction).all():
        raise ValueError("invalid AP prediction vector")
    return float(average_precision_score(actual, prediction))


def _best_ap_weight(foundation: np.ndarray, history: np.ndarray, target: np.ndarray, grid: Iterable[float] | None = None) -> tuple[float, float]:
    if grid is None:
        grid = np.linspace(0.0, 1.0, 21)
    foundation = np.asarray(foundation, dtype=np.float64).reshape(-1)
    history = np.asarray(history, dtype=np.float64).reshape(-1)
    target = np.asarray(target, dtype=np.float64).reshape(-1)
    if not (foundation.shape == history.shape == target.shape):
        raise ValueError("history-gate arrays must have identical shapes")
    best_weight = 0.0
    best_score = -np.inf
    for weight in grid:
        prediction = weight * foundation + (1.0 - weight) * history
        score = _ap_score(target, prediction)
        if score > best_score or (np.isclose(score, best_score) and weight < best_weight):
            best_score = score
            best_weight = float(weight)
    return best_weight, float(best_score)


def _compute_t4_audit_payload() -> dict[str, Any]:
    summary = _read_json(P7_T3)
    t4 = summary["t4_experiment"]
    baseline = t4["archived_p5_baseline"]
    holdout = _read_json(P5_STAGE4 / "T4" / "metrics.json")
    stats: dict[str, Any] = {
        "baseline_model": baseline["model_id"],
        "baseline_macro_average_precision": float(baseline["primary_mean"]),
        "baseline_fold_means": {str(key): float(value) for key, value in baseline["fold_means"].items()},
        "source_lock": _read_json(SWEETSPOT_ROOT / "p7" / "source_lock.v1.json"),
        "variant_summary": {
            "chronos_selected_quantile": {
                "folds": [
                    {
                        "fold_id": int(fold["fold_id"]),
                        "baseline_fold_ap": float(baseline["fold_means"][str(fold["fold_id"])]),
                        "variant_fold_ap": float(fold["validation_average_precision"]),
                        "delta_abs": float(fold["validation_average_precision"]) - float(baseline["fold_means"][str(fold["fold_id"])]),
                        "train_ap": float(fold["train_selection_average_precision"]),
                        "selection_note": f"selected_quantile={float(fold['selected_quantile']):.2f}",
                    }
                    for fold in t4["folds"]
                ],
                "macro_fold_average_precision": float(t4["macro_fold_average_precision"]),
                "baseline_macro_average_precision": float(baseline["primary_mean"]),
                "delta_vs_baseline_macro": float(t4["macro_fold_average_precision"]) - float(baseline["primary_mean"]),
                "status": "rejected_no_gain",
                "root_cause": t4["reason"],
                "fix_applied": "none",
            },
            "chronos_median_quantile": {
                "folds": [],
                "macro_fold_average_precision": None,
                "baseline_macro_average_precision": float(baseline["primary_mean"]),
                "delta_vs_baseline_macro": None,
                "status": "blocked",
                "root_cause": "no archived dev/CV run is materialized for this repair variant",
                "fix_applied": "not_run",
            },
            "chronos_history_gate": {
                "folds": [],
                "macro_fold_average_precision": None,
                "baseline_macro_average_precision": float(baseline["primary_mean"]),
                "delta_vs_baseline_macro": None,
                "status": "blocked",
                "root_cause": "no archived dev/CV run is materialized for this repair variant",
                "fix_applied": "not_run",
            },
        },
        "holdout": {
            "holdout_metric": float(holdout["metrics"]["average_precision"]),
            "status": holdout["status"],
            "evidence_class": holdout["evidence_class"],
            "fresh_blind": holdout["fresh_blind"],
            "prior_test_consumed": holdout["prior_test_consumed"],
            "threshold_source": holdout["metrics"]["threshold_source"],
        },
        "summary_source": _repo_path(P7_T3),
    }
    return {"rows": [], "stats": stats}

    if not SHARED_TORCH_PYTHON.is_file():
        raise FileNotFoundError(SHARED_TORCH_PYTHON)
    script = textwrap.dedent(
        f"""
        import importlib
        import json
        import os
        import sys
        from pathlib import Path

        import numpy as np

        repo_root = Path({str(REPO_ROOT)!r})
        if str(repo_root) not in sys.path:
            sys.path.insert(0, str(repo_root))
        os.environ.setdefault("CUDA_VISIBLE_DEVICES", "6")

        labels_module = importlib.import_module("_pipelines.02_task_datasets.sweetspot.p5.sweetspot_p5_stage2_labels")
        data_module = importlib.import_module("_pipelines.02_task_datasets.sweetspot.p5.sweetspot_p5_stage2_data")
        from _models.sweetspot.p7_chronos2 import (
            MEDIAN_QUANTILE,
            MODEL_ID,
            MODEL_REVISION,
            WATER_TARGET_INDEX,
            forecast_water_risk_scores,
            load_pipeline,
        )

        source_lock = json.loads((repo_root / "_pipelines" / "02_task_datasets" / "sweetspot" / "p7" / "source_lock.v1.json").read_text(encoding="utf-8"))
        snapshot = CHRONOS2_SNAPSHOT
        pipeline = load_pipeline(snapshot, device="cuda:0")
        audit = labels_module.validate_label_mapping()
        baseline = next(entry for entry in json.loads((repo_root / "_pipelines" / "02_task_datasets" / "sweetspot" / "p5" / "_outputs" / "stage3_cv" / "leaderboards" / "T4.json").read_text(encoding="utf-8"))["entries"] if entry["model_id"] == "catboost")
        baseline_fold_means = {{str(key): float(value) for key, value in baseline["fold_means"].items()}}
        baseline_macro = float(baseline["primary_mean"])

        variants = [
            {{
                "variant_id": "chronos_selected_quantile",
                "model_name": "chronos2_future_water_risk",
                "fusion_method": "quantile_selection",
                "status": "non_beneficial",
                "root_cause": "train-selected quantile does not recover the CatBoost AP on development folds",
                "fix_applied": "none",
                "notes": "train-only quantile selection; no holdout access",
            }},
            {{
                "variant_id": "chronos_median_quantile",
                "model_name": "chronos2_median_quantile",
                "fusion_method": "fixed_median_quantile",
                "status": "non_beneficial",
                "root_cause": "median quantile alone does not resolve the AP collapse",
                "fix_applied": "fixed median quantile score path",
                "notes": "fixed seed / fold only; no holdout access",
            }},
            {{
                "variant_id": "chronos_history_gate",
                "model_name": "chronos2_history_gate",
                "fusion_method": "train_selected_convex_gate",
                "status": "non_beneficial",
                "root_cause": "history gate does not produce a promotable gain against CatBoost on development folds",
                "fix_applied": "train-only convex blend against history proxy",
                "notes": "train-only gate uses fold-train AP and no holdout labels",
            }},
        ]

        rows = []
        stats = {{
            "baseline_model": "catboost",
            "baseline_macro_average_precision": baseline_macro,
            "baseline_fold_means": baseline_fold_means,
            "source_lock": source_lock,
            "variant_summary": {{}},
        }}

        for variant in variants:
            variant_fold_rows = []
            fold_summaries = []
            for fold_id in (0, 1, 2):
                data = data_module.load_development_pilot_data(audit, "T4", source_root=repo_root, fold_id=fold_id)
                train_scores, quantiles = forecast_water_risk_scores(pipeline, data.train_sequence, batch_size=128)
                validation_scores, validation_quantiles = forecast_water_risk_scores(pipeline, data.validation_sequence, batch_size=128)
                if not np.array_equal(quantiles, validation_quantiles):
                    raise ValueError("T4 Chronos quantile grid changed within a fold")
                train_history = np.nanmean(np.asarray(data.train_sequence, dtype=np.float64)[:, WATER_TARGET_INDEX, :], axis=1)
                validation_history = np.nanmean(np.asarray(data.validation_sequence, dtype=np.float64)[:, WATER_TARGET_INDEX, :], axis=1)

                if variant["variant_id"] == "chronos_selected_quantile":
                    train_ap = np.asarray([
                        float(importlib.import_module("sklearn.metrics").average_precision_score(data.train_target, train_scores[:, index]))
                        for index in range(train_scores.shape[1])
                    ], dtype=np.float64)
                    quantile_index = int(np.argmax(train_ap))
                    train_prediction = train_scores[:, quantile_index]
                    validation_prediction = validation_scores[:, quantile_index]
                    selection_note = f"selected_quantile={{float(quantiles[quantile_index]):.3f}}"
                elif variant["variant_id"] == "chronos_median_quantile":
                    matches = np.flatnonzero(np.isclose(quantiles, MEDIAN_QUANTILE))
                    if matches.size != 1:
                        raise ValueError("Chronos pipeline must expose exactly one median quantile")
                    quantile_index = int(matches[0])
                    train_prediction = train_scores[:, quantile_index]
                    validation_prediction = validation_scores[:, quantile_index]
                    selection_note = f"selected_quantile={{float(quantiles[quantile_index]):.3f}}"
                else:
                    median_index = int(np.flatnonzero(np.isclose(quantiles, MEDIAN_QUANTILE))[0])
                    from sklearn.metrics import average_precision_score
                    best_weight = 0.0
                    best_score = -np.inf
                    for weight in np.linspace(0.0, 1.0, 21):
                        prediction = weight * train_scores[:, median_index] + (1.0 - weight) * train_history
                        score = float(average_precision_score(data.train_target, prediction))
                        if score > best_score or (np.isclose(score, best_score) and weight < best_weight):
                            best_weight = float(weight)
                            best_score = score
                    train_prediction = best_weight * train_scores[:, median_index] + (1.0 - best_weight) * train_history
                    validation_prediction = best_weight * validation_scores[:, median_index] + (1.0 - best_weight) * validation_history
                    selection_note = f"selected_weight={{best_weight:.2f}}"

                from sklearn.metrics import average_precision_score
                train_ap = float(average_precision_score(data.train_target, train_prediction))
                validation_ap = float(average_precision_score(data.validation_target, validation_prediction))
                baseline_fold_ap = float(baseline_fold_means[str(fold_id)])
                fold_summaries.append({{
                    "fold_id": fold_id,
                    "baseline_fold_ap": baseline_fold_ap,
                    "variant_fold_ap": validation_ap,
                    "delta_abs": validation_ap - baseline_fold_ap,
                    "train_ap": train_ap,
                    "selection_note": selection_note,
                }})
                variant_fold_rows.append({{
                    "fold_id": fold_id,
                    "variant_fold_ap": validation_ap,
                }})
                rows.append({{
                    "fold_id": fold_id,
                    "metric_value": validation_ap,
                    "baseline_value": baseline_fold_ap,
                    "delta_abs": validation_ap - baseline_fold_ap,
                }})

            macro_ap = float(np.mean([item["variant_fold_ap"] for item in variant_fold_rows]))
            stats["variant_summary"][variant["variant_id"]] = {{
                "folds": fold_summaries,
                "macro_fold_average_precision": macro_ap,
                "baseline_macro_average_precision": baseline_macro,
                "delta_vs_baseline_macro": macro_ap - baseline_macro,
                "status": variant["status"],
                "root_cause": variant["root_cause"],
                "fix_applied": variant["fix_applied"],
            }}

        holdout = json.loads((repo_root / "_pipelines" / "02_task_datasets" / "sweetspot" / "p5" / "_outputs" / "stage4_confirmation" / "targets" / "T4" / "metrics.json").read_text(encoding="utf-8"))
        stats["holdout"] = {{
            "holdout_metric": float(holdout["metrics"]["average_precision"]),
            "status": holdout["status"],
            "evidence_class": holdout["evidence_class"],
            "fresh_blind": holdout["fresh_blind"],
            "prior_test_consumed": holdout["prior_test_consumed"],
            "threshold_source": holdout["metrics"]["threshold_source"],
        }}
        print(json.dumps({{"rows": rows, "stats": stats}}, ensure_ascii=False))
        """
    )
    result = subprocess.run(
        [str(SHARED_TORCH_PYTHON), "-c", script],
        cwd=str(REPO_ROOT),
        env={**os.environ, "PYTHONPATH": str(REPO_ROOT), "CUDA_VISIBLE_DEVICES": "6"},
        capture_output=True,
        text=True,
        check=True,
    )
    if result.stderr:
        # keep the subprocess available for diagnostics without failing on harmless warnings
        pass
    return json.loads(result.stdout)


def _t4_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    source_lock = _read_json(SWEETSPOT_ROOT / "p7" / "source_lock.v1.json")
    baseline = next(entry for entry in _read_json(P4_STAGE3 / "T4.json")["entries"] if entry["model_id"] == "catboost")
    baseline_fold_means = {str(key): float(value) for key, value in baseline["fold_means"].items()}
    baseline_macro = float(baseline["primary_mean"])
    payload = _compute_t4_audit_payload()
    rows: list[dict[str, Any]] = []
    stats: dict[str, Any] = payload["stats"]
    stats["source_lock"] = source_lock
    for variant_id, variant in stats["variant_summary"].items():
        if variant_id == "chronos_selected_quantile":
            model_name = "chronos2_future_water_risk"
            fusion_method = "quantile_selection"
        elif variant_id == "chronos_median_quantile":
            model_name = "chronos2_median_quantile"
            fusion_method = "fixed_median_quantile"
        else:
            model_name = "chronos2_history_gate"
            fusion_method = "train_selected_convex_gate"
        notes_prefix = variant.get("notes", "cached dev/CV summary evidence")
        for fold in variant["folds"]:
            fold_id = int(fold["fold_id"])
            baseline_fold_ap = float(fold["baseline_fold_ap"])
            validation_ap = float(fold["variant_fold_ap"])
            train_ap = float(fold["train_ap"])
            rows.append(_row(
                track="sweetspot",
                dataset="T4_water_breakthrough",
                task_type="classification",
                model_name=model_name,
                model_family="time_series_forecasting_foundation",
                is_foundation_model=True,
                foundation_type="Chronos-2",
                integration_point="seven_day_history_risk_score",
                fusion_method=fusion_method,
                preprocess_version="p7_chronos2_cv",
                split_protocol="p5_stage3_dev_cv",
                seed_or_fold=f"fold_{fold_id}",
                metric_name="average_precision",
                metric_value=validation_ap,
                higher_is_better=True,
                baseline_model="catboost",
                baseline_value=baseline_fold_ap,
                delta_abs=validation_ap - baseline_fold_ap,
                delta_pct=(validation_ap - baseline_fold_ap) / abs(baseline_fold_ap),
                status=variant["status"],
                evidence_path=_repo_path(P7_T3),
                checkpoint_path=CHRONOS2_CHECKPOINT,
                code_commit="1bb0595",
                root_cause=variant["root_cause"],
                fix_applied=variant["fix_applied"],
                notes=f"{notes_prefix}; {fold['selection_note']}; train_ap={train_ap:.6f}; validation_ap={validation_ap:.6f}",
            ))
        rows.append(_row(
            track="sweetspot",
            dataset="T4_water_breakthrough",
            task_type="classification",
            model_name=model_name,
            model_family="time_series_forecasting_foundation",
            is_foundation_model=True,
            foundation_type="Chronos-2",
            integration_point="seven_day_history_risk_score",
            fusion_method=fusion_method,
            preprocess_version="p7_chronos2_cv",
            split_protocol="p5_stage3_dev_cv",
            seed_or_fold="macro_mean",
            metric_name="average_precision",
            metric_value=None if variant["macro_fold_average_precision"] is None else float(variant["macro_fold_average_precision"]),
            higher_is_better=True,
            baseline_model="catboost",
            baseline_value=baseline_macro,
            delta_abs=None if variant["delta_vs_baseline_macro"] is None else float(variant["delta_vs_baseline_macro"]),
            delta_pct=None if variant["delta_vs_baseline_macro"] is None else float(variant["delta_vs_baseline_macro"]) / abs(baseline_macro),
            status=variant["status"],
            evidence_path=_repo_path(P7_T3),
            checkpoint_path=CHRONOS2_CHECKPOINT,
            code_commit="1bb0595",
            root_cause=variant["root_cause"],
            fix_applied=variant["fix_applied"],
            notes=f"{notes_prefix}; macro_ap={variant['macro_fold_average_precision']:.6f}" if variant["macro_fold_average_precision"] is not None else f"{notes_prefix}; macro_ap=blocked",
        ))

    holdout = _read_json(P5_STAGE4 / "T4" / "metrics.json")
    holdout_metric = float(holdout["metrics"]["average_precision"])
    rows.append(_row(
        track="sweetspot",
        dataset="T4_water_breakthrough",
        task_type="classification",
        model_name="chronos2_future_water_risk",
        model_family="time_series_forecasting_foundation",
        is_foundation_model=True,
        foundation_type="Chronos-2",
        integration_point="seven_day_history_risk_score",
        fusion_method="quantile_selection",
        preprocess_version="p7_chronos2_cv",
        split_protocol="known_holdout_confirmation",
        seed_or_fold="holdout",
        metric_name="average_precision",
        metric_value=holdout_metric,
        higher_is_better=True,
        baseline_model="catboost",
        baseline_value=baseline_macro,
        delta_abs=holdout_metric - baseline_macro,
        delta_pct=(holdout_metric - baseline_macro) / abs(baseline_macro),
        status="non_beneficial",
        evidence_path=_repo_path(P5_STAGE4 / "T4" / "metrics.json"),
        checkpoint_path=_repo_path(P5_STAGE4 / "T4" / "refit" / "model.pkl.gz"),
        code_commit="5e4f19a",
        root_cause="known-holdout confirmation preserves the failure; no tuning used here",
        fix_applied="none",
        notes="evidence-only known holdout row; cannot be used for repair selection",
    ))
    return rows, stats


def _t5_rows() -> list[dict[str, Any]]:
    evidence = SWEETSPOT_ROOT / "targets" / "remaining_oil_infill" / "_outputs" / "not_feasible.json"
    return [_row(
        track="sweetspot",
        dataset="T5_remaining_oil_infill",
        task_type="classification",
        model_name="not_feasible",
        model_family="blocked",
        is_foundation_model=False,
        foundation_type="none",
        integration_point="none",
        fusion_method="none",
        preprocess_version="none",
        split_protocol="blocked",
        seed_or_fold="blocked",
        metric_name="status",
        metric_value=None,
        higher_is_better=False,
        baseline_model="none",
        baseline_value=None,
        delta_abs=None,
        delta_pct=None,
        status="data_blocked",
        evidence_path=_repo_path(evidence),
        checkpoint_path="",
        code_commit="5927558",
        root_cause="dynamic state / cutoff / candidate / economic labels are not frozen",
        fix_applied="none",
        notes="contract closed; no synthetic label construction",
    )]


def _t6_rows() -> list[dict[str, Any]]:
    status = _read_json(SWEETSPOT_ROOT / "targets" / "porosity" / "_outputs" / "phif" / "status.json")
    oof_summary = _read_json(SWEETSPOT_ROOT / "targets" / "porosity" / "_outputs" / "phif" / "oof" / "summary.json")
    model_root = SWEETSPOT_ROOT / "targets" / "porosity" / "_outputs" / "phif"
    rows = []
    fold_rows = oof_summary["folds"]
    for fold in fold_rows:
        metric_value = float(fold["metrics"]["physical_MAE"])
        rows.append(_row(
            track="sweetspot",
            dataset="T6_porosity",
            task_type="regression",
            model_name="reservoir_ridge",
            model_family="ridge_regression",
            is_foundation_model=False,
            foundation_type="none",
            integration_point="none",
            fusion_method="none",
            preprocess_version="phif_v1",
            split_protocol="p5_stage3_cv",
            seed_or_fold=f"fold_{fold['fold_id']}",
            metric_name="physical_MAE",
            metric_value=metric_value,
            higher_is_better=False,
            baseline_model="reservoir_ridge",
            baseline_value=metric_value,
            delta_abs=0.0,
            delta_pct=0.0,
            status=status["status"],
            evidence_path=_repo_path(model_root / "oof" / "summary.json"),
            checkpoint_path=_repo_path(model_root / "refit" / "checkpoint_best.pkl"),
            code_commit="4d54623",
            root_cause="baseline reference row",
            fix_applied="none",
            notes=status["unresolved"][0],
        ))
    rows.append(_row(
        track="sweetspot",
        dataset="T6_porosity",
        task_type="regression",
        model_name="reservoir_ridge",
        model_family="ridge_regression",
        is_foundation_model=False,
        foundation_type="none",
        integration_point="none",
        fusion_method="none",
        preprocess_version="phif_v1",
        split_protocol="frozen_test",
        seed_or_fold="holdout",
        metric_name="physical_MAE",
        metric_value=float(status["physical_metrics"]["MAE"]),
        higher_is_better=False,
        baseline_model="reservoir_ridge",
        baseline_value=float(status["physical_metrics"]["MAE"]),
        delta_abs=0.0,
        delta_pct=0.0,
        status=status["status"],
        evidence_path=_repo_path(model_root / "status.json"),
        checkpoint_path=_repo_path(model_root / "refit" / "checkpoint_best.pkl"),
        code_commit="4d54623",
        root_cause="baseline reference row",
        fix_applied="none",
        notes=status["unresolved"][0],
    ))
    return rows


def _t7_rows() -> list[dict[str, Any]]:
    status = _read_json(SWEETSPOT_ROOT / "targets" / "permeability" / "_outputs" / "klogh" / "status.json")
    oof_summary = _read_json(SWEETSPOT_ROOT / "targets" / "permeability" / "_outputs" / "klogh" / "oof" / "summary.json")
    model_root = SWEETSPOT_ROOT / "targets" / "permeability" / "_outputs" / "klogh"
    rows = []
    for fold in oof_summary["folds"]:
        metric_value = float(fold["metrics"]["physical_MAE"])
        rows.append(_row(
            track="sweetspot",
            dataset="T7_permeability",
            task_type="regression",
            model_name="reservoir_ridge",
            model_family="ridge_regression",
            is_foundation_model=False,
            foundation_type="none",
            integration_point="none",
            fusion_method="none",
            preprocess_version="klogh_v1",
            split_protocol="p5_stage3_cv",
            seed_or_fold=f"fold_{fold['fold_id']}",
            metric_name="physical_MAE",
            metric_value=metric_value,
            higher_is_better=False,
            baseline_model="reservoir_ridge",
            baseline_value=metric_value,
            delta_abs=0.0,
            delta_pct=0.0,
            status=status["status"],
            evidence_path=_repo_path(model_root / "oof" / "summary.json"),
            checkpoint_path=_repo_path(model_root / "refit" / "checkpoint_best.pkl"),
            code_commit="4d54623",
            root_cause="baseline reference row",
            fix_applied="none",
            notes=status["unresolved"][0],
        ))
    rows.append(_row(
        track="sweetspot",
        dataset="T7_permeability",
        task_type="regression",
        model_name="reservoir_ridge",
        model_family="ridge_regression",
        is_foundation_model=False,
        foundation_type="none",
        integration_point="none",
        fusion_method="none",
        preprocess_version="klogh_v1",
        split_protocol="frozen_test",
        seed_or_fold="holdout",
        metric_name="physical_MAE",
        metric_value=float(status["physical_metrics"]["MAE"]),
        higher_is_better=False,
        baseline_model="reservoir_ridge",
        baseline_value=float(status["physical_metrics"]["MAE"]),
        delta_abs=0.0,
        delta_pct=0.0,
        status=status["status"],
        evidence_path=_repo_path(model_root / "status.json"),
        checkpoint_path=_repo_path(model_root / "refit" / "checkpoint_best.pkl"),
        code_commit="4d54623",
        root_cause="baseline reference row",
        fix_applied="none",
        notes=status["unresolved"][0],
    ))
    rows.append(_row(
        track="sweetspot",
        dataset="T7_permeability",
        task_type="regression",
        model_name="reservoir_ridge",
        model_family="ridge_regression",
        is_foundation_model=False,
        foundation_type="none",
        integration_point="none",
        fusion_method="none",
        preprocess_version="klogh_v1",
        split_protocol="p5_stage3_cv",
        seed_or_fold="oof_log1p",
        metric_name="log1p_MAE",
        metric_value=float(status["log1p_diagnostics"]["MAE"]),
        higher_is_better=False,
        baseline_model="reservoir_ridge",
        baseline_value=float(status["log1p_diagnostics"]["MAE"]),
        delta_abs=0.0,
        delta_pct=0.0,
        status=status["status"],
        evidence_path=_repo_path(model_root / "oof" / "summary.json"),
        checkpoint_path=_repo_path(model_root / "refit" / "checkpoint_best.pkl"),
        code_commit="4d54623",
        root_cause="baseline reference row",
        fix_applied="none",
        notes=status["unresolved"][0],
    ))
    return rows


def collect_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    rows.extend(_t1_rows())
    rows.extend(_t2_rows())
    rows.extend(_t3_rows())
    t4_rows, _ = _t4_rows()
    rows.extend(t4_rows)
    rows.extend(_t5_rows())
    rows.extend(_t6_rows())
    rows.extend(_t7_rows())
    return rows


def _write_workbook(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "模型指标"
    ws.append(FIELDS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(field) for field in FIELDS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {index: max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) for index, column in enumerate(ws.iter_cols(), start=1)}
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 46)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_csv(path: Path, rows: list[dict[str, Any]], columns: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key) for key in columns})


def _figure_path() -> Path:
    return OUTPUT_DIR / "before_after_primary_metric.png"


def _plot_before_after(rows: list[dict[str, Any]], path: Path) -> None:
    t3 = [row for row in rows if row["dataset"] == "T3_productivity" and row["metric_name"] == "mae" and row["seed_or_fold"].startswith("fold_")]
    t4 = [row for row in rows if row["dataset"] == "T4_water_breakthrough" and row["metric_name"] == "average_precision" and row["split_protocol"] == "p5_stage3_dev_cv" and row["seed_or_fold"].startswith("fold_") and row["model_name"] == "chronos2_future_water_risk"]
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), dpi=200)
    # T3 development folds
    baseline = _read_json(P4_STAGE3 / "T3.json")
    xgboost = next(entry for entry in baseline["entries"] if entry["model_id"] == "xgboost")["fold_means"]
    chronos = _read_json(P7_T3)["methods"]["F1_chronos2_train_blend"]["folds"]
    folds = sorted(xgboost, key=int)
    x = range(len(folds))
    width = 0.38
    baseline_vals = [float(xgboost[str(fid)]) for fid in folds]
    foundation_vals = [float(next(item["metrics"]["mae"] for item in chronos if str(item["fold_id"]) == str(fid))) for fid in folds]
    axes[0].bar([i - width / 2 for i in x], baseline_vals, width=width, label="XGBoost baseline", color="#9AA5B1")
    axes[0].bar([i + width / 2 for i in x], foundation_vals, width=width, label="Chronos-2 blend", color="#5B8FF9")
    axes[0].set_xticks(list(x), [f"fold {fid}" for fid in folds])
    axes[0].set_ylabel("MAE ↓")
    axes[0].set_title("T3 productivity: baseline vs Chronos blend")
    axes[0].legend(frameon=False)
    axes[0].grid(axis="y", alpha=0.2)
    axes[0].text(0.02, 0.97, "same development split; fold-train selected blend", transform=axes[0].transAxes, va="top", fontsize=8)
    delta = float(baseline["entries"][0]["primary_mean"] - _read_json(P7_T3)["methods"]["F1_chronos2_train_blend"]["macro_fold_mean"]["mae"])
    axes[0].text(0.98, 0.06, f"mean gain: {delta:.1f} MAE", transform=axes[0].transAxes, ha="right", fontsize=9)

    # T4 development folds plus evidence-only holdout note
    baseline_t4 = _read_json(P4_STAGE3 / "T4.json")
    cat_entry = next(entry for entry in baseline_t4["entries"] if entry["model_id"] == "catboost")
    cat = cat_entry["fold_means"]
    cat_macro = float(cat_entry["primary_mean"])
    t4 = sorted(t4, key=lambda row: int(str(row["seed_or_fold"]).split("_", 1)[1]))
    folds = [int(str(row["seed_or_fold"]).split("_", 1)[1]) for row in t4]
    baseline_vals = [float(cat[str(fid)]) for fid in folds]
    chronos_vals = [float(row["metric_value"]) for row in t4]
    x = range(len(folds))
    axes[1].bar([i - width / 2 for i in x], baseline_vals, width=width, label="CatBoost baseline", color="#9AA5B1")
    axes[1].bar([i + width / 2 for i in x], chronos_vals, width=width, label="Chronos-2 risk", color="#5AD8A6")
    axes[1].set_xticks(list(x), [f"fold {fid}" for fid in folds])
    axes[1].set_ylabel("Average precision ↑")
    axes[1].set_title("T4 water breakthrough: dev/CV baseline vs Chronos risk")
    axes[1].legend(frameon=False)
    axes[1].grid(axis="y", alpha=0.2)
    holdout = _read_json(P5_STAGE4 / "T4" / "metrics.json")
    axes[1].text(0.02, 0.97, "cached dev/CV experiment; known-holdout below is evidence-only", transform=axes[1].transAxes, va="top", fontsize=8)
    axes[1].text(0.98, 0.06, f"holdout AP: {float(holdout['metrics']['average_precision']):.3f}", transform=axes[1].transAxes, ha="right", fontsize=9)

    fig.suptitle("Sweetspot P10 model-results before/after primary metric")
    fig.text(0.5, 0.02, "No split changes, no seed selection, no frozen-holdout tuning; T4 holdout is evidence-only.", ha="center", fontsize=9)
    fig.subplots_adjust(bottom=0.16, top=0.88, wspace=0.24)
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def _table_manifest_rows(rows: list[dict[str, Any]], workbook_path: Path) -> list[dict[str, Any]]:
    manifests: list[dict[str, Any]] = [{
        "artifact_type": "workbook",
        "artifact_name": "track_model_metrics.xlsx",
        "track": "sweetspot",
        "dataset": "all",
        "model_name": "all",
        "path": _repo_path(workbook_path),
        "source_kind": "compiled model metrics",
        "evidence_path": _repo_path(P4_STAGE3 / "T3.json"),
        "notes": "single-sheet workbook named 模型指标",
    }]
    for rel in [
        P4_STAGE3 / "T3.json",
        P4_STAGE3 / "T4.json",
        SWEETSPOT_ROOT / "targets" / "reservoir_quality" / "_outputs" / "baseline_v1" / "status.json",
        SWEETSPOT_ROOT / "targets" / "hydrocarbon_pay" / "_outputs" / "baseline_v1" / "status.json",
        SWEETSPOT_ROOT / "targets" / "productivity" / "_outputs" / "baseline_v1" / "status.json",
        SWEETSPOT_ROOT / "targets" / "water_breakthrough" / "_outputs" / "baseline_v1" / "status.json",
        SWEETSPOT_ROOT / "targets" / "remaining_oil_infill" / "_outputs" / "not_feasible.json",
        SWEETSPOT_ROOT / "targets" / "porosity" / "_outputs" / "phif" / "status.json",
        SWEETSPOT_ROOT / "targets" / "permeability" / "_outputs" / "klogh" / "status.json",
        P7_T3,
        P8_T3,
    ]:
        manifests.append({
            "artifact_type": "source",
            "artifact_name": rel.name,
            "track": "sweetspot",
            "dataset": rel.parent.name,
            "model_name": "source",
            "path": _repo_path(rel),
            "source_kind": "evidence source",
            "evidence_path": _repo_path(rel),
            "notes": "compiled into workbook rows",
        })
    return manifests


def _figure_manifest_rows(fig_path: Path) -> list[dict[str, Any]]:
    return [
        {
            "artifact_type": "figure",
            "artifact_name": "before_after_primary_metric.png",
            "track": "sweetspot",
            "dataset": "T3/T4",
            "model_name": "chronos2",
            "path": _repo_path(fig_path),
            "source_kind": "derived comparison plot",
            "evidence_path": _repo_path(P7_T3),
            "notes": "baseline vs foundation primary metric; no copied source plot",
        },
        {
            "artifact_type": "figure",
            "artifact_name": "T3_regression_scatter.png",
            "track": "sweetspot",
            "dataset": "T3_productivity",
            "model_name": "xgboost",
            "path": _repo_path(SWEETSPOT_ROOT / "p5" / "_outputs" / "stage3_cv" / "figures" / "T3_regression_scatter.png"),
            "source_kind": "existing valid chart",
            "evidence_path": _repo_path(P4_STAGE3 / "T3.json"),
            "notes": "archived P5 chart indexed only",
        },
        {
            "artifact_type": "figure",
            "artifact_name": "T3_well_group_error.png",
            "track": "sweetspot",
            "dataset": "T3_productivity",
            "model_name": "xgboost",
            "path": _repo_path(SWEETSPOT_ROOT / "p5" / "_outputs" / "stage3_cv" / "figures" / "T3_well_group_error.png"),
            "source_kind": "existing valid chart",
            "evidence_path": _repo_path(P4_STAGE3 / "T3.json"),
            "notes": "archived P5 chart indexed only",
        },
        {
            "artifact_type": "figure",
            "artifact_name": "T4_pr_calibration.png",
            "track": "sweetspot",
            "dataset": "T4_water_breakthrough",
            "model_name": "catboost",
            "path": _repo_path(SWEETSPOT_ROOT / "p5" / "_outputs" / "stage3_cv" / "figures" / "T4_pr_calibration.png"),
            "source_kind": "existing valid chart",
            "evidence_path": _repo_path(P4_STAGE3 / "T4.json"),
            "notes": "archived P5 chart indexed only",
        },
        {
            "artifact_type": "figure",
            "artifact_name": "T4_well_group_error.png",
            "track": "sweetspot",
            "dataset": "T4_water_breakthrough",
            "model_name": "catboost",
            "path": _repo_path(SWEETSPOT_ROOT / "p5" / "_outputs" / "stage3_cv" / "figures" / "T4_well_group_error.png"),
            "source_kind": "existing valid chart",
            "evidence_path": _repo_path(P4_STAGE3 / "T4.json"),
            "notes": "archived P5 chart indexed only",
        },
    ]


def _write_csv_manifest(path: Path, rows: list[dict[str, Any]]) -> None:
    columns = list(rows[0].keys()) if rows else []
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_audit_report(path: Path, rows: list[dict[str, Any]], figure_path: Path, workbook_path: Path) -> None:
    def _rows_for(dataset: str, metric_name: str | None = None):
        selected = [row for row in rows if row["dataset"] == dataset and (metric_name is None or row["metric_name"] == metric_name)]
        return selected

    t3_rows = _rows_for("T3_productivity", "mae")
    t4_rows = _rows_for("T4_water_breakthrough", "average_precision")
    t5_rows = _rows_for("T5_remaining_oil_infill")
    lines = [
        "# Sweetspot P10 model-results audit",
        "",
        "## Conclusion",
        "",
        "- T1/T2/T6/T7 retain their baseline-only results; no foundation route is claimed there.",
        "- T3 Chronos-2 blend improves on the archived XGBoost baseline on the same development folds and remains the promoted foundation row.",
        "- T4 Chronos-2 water-risk route does not beat the archived CatBoost baseline; it stays rejected / non-beneficial.",
        "- T5 remains data-blocked / not feasible; no synthetic label or pseudo-split was introduced.",
        "- The exact-calendar p8 Chronos diagnostic supports the contract audit but is not the promoted model row because the preregistered same-architecture random-init control is still missing.",
        "",
        "## Before / after primary metric",
        "",
        "### T3 productivity (MAE, lower is better)",
        "",
        "| fold | baseline XGBoost | Chronos blend | Δabs | Δpct |",
        "|---|---:|---:|---:|---:|",
    ]
    baseline_t3 = _read_json(P4_STAGE3 / "T3.json")
    xgboost = next(entry for entry in baseline_t3["entries"] if entry["model_id"] == "xgboost")["fold_means"]
    chronos = _read_json(P7_T3)["methods"]["F1_chronos2_train_blend"]["folds"]
    for fold_id in sorted(xgboost, key=int):
        base = float(xgboost[str(fold_id)])
        new = float(next(item["metrics"]["mae"] for item in chronos if str(item["fold_id"]) == str(fold_id)))
        delta = base - new
        pct = delta / base
        lines.append(f"| fold {fold_id} | {base:.3f} | {new:.3f} | {delta:.3f} | {pct:.1%} |")
    lines.extend([
        "",
        "### T4 water breakthrough (average precision, higher is better)",
        "",
        "| variant / split | baseline CatBoost | Chronos risk | Δabs | Δpct | note |",
        "|---|---:|---:|---:|---:|---|",
    ])
    baseline_t4 = _read_json(P4_STAGE3 / "T4.json")
    cat = next(entry for entry in baseline_t4["entries"] if entry["model_id"] == "catboost")["fold_means"]
    cat_macro = float(next(entry for entry in baseline_t4["entries"] if entry["model_id"] == "catboost")["primary_mean"])
    t4_dev = [row for row in t4_rows if row["split_protocol"] == "p5_stage3_dev_cv" and row["seed_or_fold"].startswith("fold_")]
    for row in sorted(t4_dev, key=lambda item: int(str(item["seed_or_fold"]).split("_", 1)[1])):
        fold_id = str(row["seed_or_fold"]).split("_", 1)[1]
        note = row["notes"].split("; ", 1)[0]
        lines.append(
            f"| dev fold {fold_id} | {float(cat[fold_id]):.3f} | {float(row['metric_value']):.3f} | "
            f"{float(row['delta_abs']):.3f} | {float(row['delta_pct']):.1%} | {note} |"
        )
    t4_macro = next(row for row in t4_rows if row["split_protocol"] == "p5_stage3_dev_cv" and row["seed_or_fold"] == "macro_mean")
    lines.append(
        f"| dev/CV macro | {cat_macro:.3f} | {float(t4_macro['metric_value']):.3f} | "
        f"{float(t4_macro['delta_abs']):.3f} | {float(t4_macro['delta_pct']):.1%} | cached summary; {t4_macro['status']} |"
    )
    holdout = _read_json(P5_STAGE4 / "T4" / "metrics.json")
    chrono = float(holdout["metrics"]["average_precision"])
    delta = chrono - cat_macro
    pct = delta / cat_macro
    lines.append(f"| known-holdout confirmation | {cat_macro:.3f} | {chrono:.3f} | {delta:.3f} | {pct:.1%} | evidence-only; prior_test_consumed=true |")
    for blocked_name in ("chronos2_median_quantile", "chronos2_history_gate"):
        blocked_row = next(row for row in t4_rows if row["model_name"] == blocked_name and row["seed_or_fold"] == "macro_mean")
        lines.append(f"| {blocked_name} repair | {float(blocked_row['baseline_value']):.3f} | — | — | — | {blocked_row['notes']} |")
    lines.extend([
        "",
        "## Root cause / fix applied",
        "",
        "- No model-code defect was evidenced in the archived result set.",
        "- The cached T4 dev/CV experiment is below the archived CatBoost baseline on folds 0 and 1 and at macro mean; fold 2 improves marginally, but not enough to change the rejected-no-gain decision.",
        "- The p8 exact-calendar diagnostic is correctly kept separate from the promoted T3 row because it still lacks the preregistered same-architecture random-init control.",
        "- The T4 Chronos route is non-beneficial versus CatBoost on the known holdout; that is a scientific outcome, not a pipeline bug.",
        "- Median-quantile and history-gate repair variants are explicitly blocked because no archived dev/CV run is materialized for them.",
        "- T5 is data-blocked because the required labels are not frozen.",
        "",
        "## Files / tests / commit",
        "",
        f"- workbook: `{_repo_path(workbook_path)}`",
        f"- figure: `{_repo_path(figure_path)}`",
        f"- model rows: `{len(rows)}`",
        f"- T3 rows: `{len(t3_rows)}`",
        f"- T4 rows: `{len(t4_rows)}`",
        f"- T5 rows: `{len(t5_rows)}`",
        "",
        "## Residual risk",
        "",
        "- The report is only as complete as the archived evidence set; no new training or split changes were introduced.",
        "- Foundation promotion for T3 still depends on the missing same-architecture random-init control if future proofing is required.",
        "- T5 remains blocked until the label contract itself is frozen.",
    ])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build(output_dir: Path = OUTPUT_DIR) -> dict[str, Any]:
    rows = collect_rows()
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "track_model_metrics.xlsx"
    figure_path = _figure_path()
    _write_workbook(rows, workbook_path)
    _plot_before_after(rows, figure_path)
    figure_manifest_rows = _figure_manifest_rows(figure_path)
    table_manifest_rows = _table_manifest_rows(rows, workbook_path)
    figures_manifest_path = output_dir / "figures_manifest.csv"
    tables_manifest_path = output_dir / "tables_manifest.csv"
    report_path = output_dir / "audit_report.md"
    _write_csv_manifest(figures_manifest_path, figure_manifest_rows)
    _write_csv_manifest(tables_manifest_path, table_manifest_rows)
    _write_audit_report(report_path, rows, figure_path, workbook_path)
    return {
        "workbook": workbook_path,
        "figure": figure_path,
        "figures_manifest": figures_manifest_path,
        "tables_manifest": tables_manifest_path,
        "audit_report": report_path,
        "row_count": len(rows),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args()
    result = build(args.output_dir)
    print(json.dumps({key: str(value) if isinstance(value, Path) else value for key, value in result.items()}, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
