from __future__ import annotations

import csv
import importlib
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


results = importlib.import_module("_pipelines.02_task_datasets.sweetspot.p10.results")


class SweetspotP10ResultsTests(unittest.TestCase):
    def test_build_outputs_and_excel_roundtrip(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = results.build(Path(tmp))
            workbook = Path(out["workbook"])
            figure = Path(out["figure"])
            figures_manifest = Path(out["figures_manifest"])
            tables_manifest = Path(out["tables_manifest"])
            audit_report = Path(out["audit_report"])
            self.assertTrue(workbook.is_file())
            self.assertTrue(figure.is_file())
            self.assertTrue(figures_manifest.is_file())
            self.assertTrue(tables_manifest.is_file())
            self.assertTrue(audit_report.is_file())
            wb = load_workbook(workbook, read_only=True)
            self.assertEqual(wb.sheetnames, ["模型指标"])
            ws = wb["模型指标"]
            self.assertGreater(ws.max_row, 10)
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.assertIn("evidence_path", header)
            self.assertIn("checkpoint_path", header)
            self.assertEqual(header[0], "track")
            with Image.open(figure) as image:
                self.assertGreaterEqual(image.size[0], 1200)
                self.assertGreaterEqual(image.size[1], 500)
            with figures_manifest.open(encoding="utf-8", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertTrue(any(row["artifact_name"] == "before_after_primary_metric.png" for row in rows))
            with tables_manifest.open(encoding="utf-8", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertTrue(any(row["artifact_name"] == "track_model_metrics.xlsx" for row in rows))
            for row in rows:
                path = Path(results.REPO_ROOT) / row["path"]
                self.assertTrue(path.exists(), path)

    def test_row_contract_includes_blocked_and_foundation_rows(self) -> None:
        rows = results.collect_rows()
        statuses = {row["status"] for row in rows}
        self.assertIn("promote", statuses)
        self.assertIn("rejected_no_gain", statuses)
        self.assertIn("data_blocked", statuses)
        self.assertTrue(any(row["dataset"] == "T3_productivity_calendar_diag" for row in rows))
        self.assertTrue(any(row["dataset"] == "T5_remaining_oil_infill" for row in rows))
        self.assertTrue(any(row["dataset"] == "T7_permeability" and row["metric_name"] == "log1p_MAE" for row in rows))
        for row in rows:
            self.assertIn(row["track"], {"sweetspot"})
            self.assertIn("evidence_path", row)
            self.assertTrue(row["evidence_path"] == "" or (Path(results.REPO_ROOT) / row["evidence_path"]).exists())


if __name__ == "__main__":
    unittest.main()

