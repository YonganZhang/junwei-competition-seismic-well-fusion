"""Build the six-track model-results summary bundle.

This script only reads the already-published track-level artifacts and writes a
portable summary bundle under ``_outputs/model_results_summary``.
"""
from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, NamedTuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parents[1]
OUTPUT_DIR = HERE

CANONICAL_TRACKS = ("fault", "facies", "property", "lithofacies", "sweetspot", "reconstruction")
TRACK_ALIASES = {
    "reservoir": "property",
    "property": "property",
    "fault": "fault",
    "facies": "facies",
    "lithofacies": "lithofacies",
    "sweetspot": "sweetspot",
    "reconstruction": "reconstruction",
}
EXPECTED_SHEET = "模型指标"
EXPECTED_COLS = 26
WORKBOOK_HEADERS: list[str] | None = None


class TrackSource(NamedTuple):
    source_key: str
    canonical_track: str
    workbook: Path
    figures_manifest: Path
    tables_manifest: Path
    source_root: Path


TRACK_SOURCES = (
    TrackSource(
        source_key="fault",
        canonical_track="fault",
        workbook=PROJECT_ROOT / "_pipelines/02_task_datasets/fault/_outputs/p10_model_results/track_model_metrics.xlsx",
        figures_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/fault/_outputs/p10_model_results/figures_manifest.csv",
        tables_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/fault/_outputs/p10_model_results/tables_manifest.csv",
        source_root=PROJECT_ROOT / "_pipelines/02_task_datasets/fault",
    ),
    TrackSource(
        source_key="facies",
        canonical_track="facies",
        workbook=PROJECT_ROOT / "_pipelines/02_task_datasets/facies/_outputs/p10_model_results/track_model_metrics.xlsx",
        figures_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/facies/_outputs/p10_model_results/figures_manifest.csv",
        tables_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/facies/_outputs/p10_model_results/tables_manifest.csv",
        source_root=PROJECT_ROOT / "_pipelines/02_task_datasets/facies",
    ),
    TrackSource(
        source_key="reservoir",
        canonical_track="property",
        workbook=PROJECT_ROOT / "_pipelines/02_task_datasets/reservoir/_outputs/p10_model_results/track_model_metrics.xlsx",
        figures_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/reservoir/_outputs/p10_model_results/figures_manifest.csv",
        tables_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/reservoir/_outputs/p10_model_results/tables_manifest.csv",
        source_root=PROJECT_ROOT / "_pipelines/02_task_datasets/reservoir",
    ),
    TrackSource(
        source_key="lithofacies",
        canonical_track="lithofacies",
        workbook=PROJECT_ROOT / "_pipelines/02_task_datasets/lithofacies/_outputs/p10_model_results/track_model_metrics.xlsx",
        figures_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/lithofacies/_outputs/p10_model_results/figures_manifest.csv",
        tables_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/lithofacies/_outputs/p10_model_results/tables_manifest.csv",
        source_root=PROJECT_ROOT / "_pipelines/02_task_datasets/lithofacies",
    ),
    TrackSource(
        source_key="sweetspot",
        canonical_track="sweetspot",
        workbook=PROJECT_ROOT / "_pipelines/02_task_datasets/sweetspot/_outputs/p10_model_results/track_model_metrics.xlsx",
        figures_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/sweetspot/_outputs/p10_model_results/figures_manifest.csv",
        tables_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/sweetspot/_outputs/p10_model_results/tables_manifest.csv",
        source_root=PROJECT_ROOT / "_pipelines/02_task_datasets/sweetspot",
    ),
    TrackSource(
        source_key="reconstruction",
        canonical_track="reconstruction",
        workbook=PROJECT_ROOT / "_pipelines/02_task_datasets/reconstruction/_outputs/p10_model_results/track_model_metrics.xlsx",
        figures_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/reconstruction/_outputs/p10_model_results/figures_manifest.csv",
        tables_manifest=PROJECT_ROOT / "_pipelines/02_task_datasets/reconstruction/_outputs/p10_model_results/tables_manifest.csv",
        source_root=PROJECT_ROOT / "_pipelines/02_task_datasets/reconstruction",
    ),
)


def _repo_rel(path: Path | str) -> str:
    text = str(path)
    if text.startswith(str(PROJECT_ROOT)):
        try:
            return Path(text).resolve().relative_to(PROJECT_ROOT.resolve()).as_posix()
        except ValueError:
            pass
    marker = "_pipelines/"
    if marker in text:
        return text[text.index(marker) :]
    return Path(text).as_posix()


def _resolve_source_path(raw: str | None, source_root: Path) -> str:
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""
    if text.startswith(str(PROJECT_ROOT)):
        return _repo_rel(text)
    if text.startswith("/"):
        if "_pipelines/" in text:
            return text[text.index("_pipelines/") :]
        return text
    if text.startswith("_pipelines/"):
        return text
    return _repo_rel(source_root / text)


def _looks_like_path(text: str | None) -> bool:
    if text is None:
        return False
    value = str(text).strip()
    if not value:
        return False
    if value.startswith("/") or value.startswith("_"):
        return True
    if "_outputs/" in value or "_pipelines/" in value:
        return True
    return value.endswith((".json", ".jsonl", ".csv", ".xlsx", ".png", ".md"))


def _normalize_workbook_path(raw: Any, source_root: Path) -> Any:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text in {"artifact_unavailable", "none", "null"}:
        return raw
    if text.startswith("/"):
        return _repo_rel(text)
    if text.startswith("_pipelines/"):
        return text
    if text.startswith("_") or "/" in text:
        return _resolve_source_path(text, source_root)
    return raw


def _json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_workbook(source: TrackSource) -> tuple[list[str], list[dict[str, Any]]]:
    if not source.workbook.is_file():
        raise FileNotFoundError(source.workbook)
    wb = load_workbook(source.workbook, read_only=True)
    if wb.sheetnames != [EXPECTED_SHEET]:
        raise ValueError(f"{source.workbook}: expected exactly one sheet named {EXPECTED_SHEET}, got {wb.sheetnames}")
    ws = wb[EXPECTED_SHEET]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if len(header) != EXPECTED_COLS:
        raise ValueError(f"{source.workbook}: expected {EXPECTED_COLS} columns, got {len(header)}")
    global WORKBOOK_HEADERS
    if WORKBOOK_HEADERS is None:
        WORKBOOK_HEADERS = header
    elif header != WORKBOOK_HEADERS:
        raise ValueError(f"{source.workbook}: workbook header mismatch")
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header[i]: values[i] for i in range(len(header))}
        track_value = str(row.get("track", "")).strip()
        canonical = TRACK_ALIASES.get(track_value, track_value)
        if canonical not in CANONICAL_TRACKS:
            raise ValueError(f"{source.workbook}: unknown track label {track_value!r}")
        row["track"] = canonical
        for field in ("evidence_path", "checkpoint_path"):
            if field in row:
                row[field] = _normalize_workbook_path(row[field], source.source_root)
        rows.append(row)
    return header, rows


def _manifest_rows(path: Path, source: TrackSource, manifest_kind: str) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for raw in reader:
            path_value = raw.get("path") or raw.get("figure_path") or raw.get("table_path") or raw.get("figure") or raw.get("table")
            source_value = raw.get("source")
            evidence_value = raw.get("evidence_path") or raw.get("source_evidence")
            if not evidence_value and _looks_like_path(source_value):
                evidence_value = source_value
            source_kind = raw.get("source_kind") or (source_value if source_value and not _looks_like_path(source_value) else manifest_kind)
            artifact_name = raw.get("artifact_name") or raw.get("name") or raw.get("figure_name") or raw.get("table_name") or Path(str(path_value or "")).name
            row = {
                "track": source.canonical_track,
                "artifact_type": raw.get("artifact_type") or raw.get("kind") or manifest_kind.rstrip("s"),
                "artifact_name": artifact_name,
                "kind": raw.get("kind") or raw.get("artifact_type") or manifest_kind.rstrip("s"),
                "dataset": raw.get("dataset") or "",
                "model_name": raw.get("model_name") or "",
                "target": raw.get("target") or "",
                "split_protocol": raw.get("split_protocol") or raw.get("split") or "",
                "path": _resolve_source_path(path_value, source.source_root),
                "status": raw.get("status") or ("generated" if "generated" in str(source_kind) else "indexed"),
                "source_kind": source_kind,
                "evidence_path": _resolve_source_path(evidence_value, source.source_root) if evidence_value else _resolve_source_path(path_value, source.source_root),
                "sha256": raw.get("sha256") or "",
                "row_count": raw.get("row_count") or raw.get("rows") or "",
                "sheet_name": raw.get("sheet_name") or "",
                "notes": raw.get("notes") or raw.get("description") or "",
            }
            rows.append(row)
    return rows


def _validate_path(path_text: str) -> bool:
    if not path_text:
        return False
    path = PROJECT_ROOT / path_text if not Path(path_text).is_absolute() else Path(path_text)
    return path.is_file()


def _write_csv(rows: list[dict[str, Any]], path: Path, fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def _write_workbook(rows: list[dict[str, Any]], path: Path) -> None:
    headers = WORKBOOK_HEADERS
    if headers is None:
        raise RuntimeError("workbook headers were not initialized")
    wb = Workbook()
    ws = wb.active
    ws.title = EXPECTED_SHEET
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column[: min(len(column), 40)]]
        width = min(max([len(v) for v in values] + [len(str(column[0].value))]) + 2, 42)
        ws.column_dimensions[column[0].column_letter].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    reopened = load_workbook(path, read_only=True)
    if reopened.sheetnames != [EXPECTED_SHEET]:
        raise RuntimeError(f"{path}: workbook must contain exactly one sheet named {EXPECTED_SHEET}")


def _collect_source_data() -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    workbook_rows: list[dict[str, Any]] = []
    figure_rows: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []
    validation: dict[str, Any] = {
        "schema_version": "model-results-summary/v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "repo_root": _repo_rel(PROJECT_ROOT),
        "sources": {},
        "track_alias_map": TRACK_ALIASES,
        "canonical_tracks": list(CANONICAL_TRACKS),
    }

    combined_track_counts: Counter[str] = Counter()
    combined_status_counts: Counter[str] = Counter()
    track_seen = set()
    workbook_headers = None

    for source in TRACK_SOURCES:
        header, rows = _read_workbook(source)
        workbook_headers = header
        if not rows:
            raise ValueError(f"{source.workbook}: workbook is empty")
        source_track_values = sorted({str(row["track"]) for row in rows})
        if source.canonical_track not in source_track_values:
            raise ValueError(f"{source.workbook}: normalized track {source.canonical_track!r} missing")
        track_seen.update(source_track_values)
        track_counter = Counter(str(row["track"]) for row in rows)
        status_counter = Counter(str(row.get("status", "")) for row in rows)
        combined_track_counts.update(track_counter)
        combined_status_counts.update(status_counter)
        workbook_rows.extend(rows)

        fig_rows = _manifest_rows(source.figures_manifest, source, "figures")
        tab_rows = _manifest_rows(source.tables_manifest, source, "tables")
        figure_rows.extend(fig_rows)
        table_rows.extend(tab_rows)

        # Include the source workbook itself in the table index.
        table_rows.append(
            {
                "track": source.canonical_track,
                "artifact_type": "workbook",
                "artifact_name": "track_model_metrics.xlsx",
                "kind": "workbook",
                "dataset": "",
                "model_name": "",
                "target": "",
                "split_protocol": "",
                "path": _repo_rel(source.workbook),
                "status": "generated",
                "source_kind": "source_workbook",
                "evidence_path": _repo_rel(source.workbook),
                "sha256": _sha256(source.workbook),
                "row_count": str(len(rows)),
                "sheet_name": EXPECTED_SHEET,
                "notes": "source workbook indexed for global summary",
            }
        )
        validation["sources"][source.canonical_track] = {
            "source_key": source.source_key,
            "workbook": {
                "path": _repo_rel(source.workbook),
                "sheet_name": EXPECTED_SHEET,
                "header_count": len(header),
                "row_count": len(rows),
                "tracks": source_track_values,
                "status_counts": dict(status_counter),
            },
            "figures_manifest": {
                "path": _repo_rel(source.figures_manifest),
                "row_count": len(fig_rows),
            },
            "tables_manifest": {
                "path": _repo_rel(source.tables_manifest),
                "row_count": len(tab_rows),
            },
        }

    combined_track_counts = Counter({k: combined_track_counts[k] for k in sorted(combined_track_counts)})
    combined_status_counts = Counter({k: combined_status_counts[k] for k in sorted(combined_status_counts)})
    validation["combined"] = {
        "workbook": {
            "sheet_name": EXPECTED_SHEET,
            "row_count": len(workbook_rows),
            "column_count": len(workbook_headers or []),
            "path": _repo_rel(OUTPUT_DIR / "six_track_model_metrics.xlsx"),
        },
        "track_counts": dict(combined_track_counts),
        "status_counts": dict(combined_status_counts),
        "tracks_present": sorted(track_seen),
        "track_count": len(track_seen),
        "manifest_rows": {
            "figures": len(figure_rows),
            "tables": len(table_rows),
        },
        "evidence_missing": [],
        "evidence_checked": 0,
    }

    return workbook_rows, figure_rows, table_rows, validation


def _add_generated_artifacts(table_rows: list[dict[str, Any]]) -> None:
    generated = [
        ("workbook", "six_track_model_metrics.xlsx", OUTPUT_DIR / "six_track_model_metrics.xlsx", "generated_summary", "single-sheet consolidated workbook"),
        ("json", "field_validation_report.json", OUTPUT_DIR / "field_validation_report.json", "validation_report", "machine-readable validation report"),
        ("markdown", "human_summary.md", OUTPUT_DIR / "human_summary.md", "human_summary", "reader-facing summary of the six-track bundle"),
    ]
    for artifact_type, artifact_name, path, source_kind, notes in generated:
        table_rows.append(
            {
                "track": "summary",
                "artifact_type": artifact_type,
                "artifact_name": artifact_name,
                "kind": artifact_type,
                "dataset": "",
                "model_name": "",
                "target": "",
                "split_protocol": "",
                "path": _repo_rel(path),
                "status": "generated",
                "source_kind": source_kind,
                "evidence_path": _repo_rel(path),
                "sha256": _sha256(path) if path.is_file() else "",
                "row_count": "",
                "sheet_name": EXPECTED_SHEET if artifact_name.endswith(".xlsx") else "",
                "notes": notes,
            }
        )


def _validate_outputs(workbook_rows: list[dict[str, Any]], figure_rows: list[dict[str, Any]], table_rows: list[dict[str, Any]]) -> dict[str, Any]:
    workbook_path = OUTPUT_DIR / "six_track_model_metrics.xlsx"
    wb = load_workbook(workbook_path, read_only=True)
    if wb.sheetnames != [EXPECTED_SHEET]:
        raise RuntimeError("combined workbook must contain exactly one sheet named 模型指标")
    ws = wb[EXPECTED_SHEET]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if len(header) != EXPECTED_COLS:
        raise RuntimeError(f"combined workbook must have exactly {EXPECTED_COLS} columns")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    row_count = len(rows)
    if row_count != len(workbook_rows):
        raise RuntimeError(f"combined workbook row count mismatch: {row_count} vs {len(workbook_rows)}")
    evidence_idx = header.index("evidence_path")
    missing = sorted({str(row[evidence_idx]) for row in rows if row[evidence_idx] and not _validate_path(str(row[evidence_idx]))})
    tracks = sorted({str(row[header.index("track")]) for row in rows})
    status_counts = Counter(str(row[header.index("status")]) for row in rows)
    if set(tracks) != set(CANONICAL_TRACKS):
        raise RuntimeError(f"combined workbook tracks mismatch: {tracks}")
    if missing:
        raise RuntimeError(f"combined workbook has missing evidence paths: {missing[:5]}")
    validation = {
        "schema_version": "model-results-summary/v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook": {
            "path": _repo_rel(workbook_path),
            "sheet_name": EXPECTED_SHEET,
            "row_count": row_count,
            "column_count": len(header),
            "tracks": tracks,
            "track_count": len(tracks),
            "status_counts": dict(status_counts),
        },
        "figures_manifest": {
            "path": _repo_rel(OUTPUT_DIR / "all_figures_manifest.csv"),
            "row_count": len(figure_rows),
        },
        "tables_manifest": {
            "path": _repo_rel(OUTPUT_DIR / "all_tables_manifest.csv"),
            "row_count": len(table_rows),
        },
        "evidence_missing": missing,
        "evidence_checked": row_count,
        "track_alias_map": TRACK_ALIASES,
    }
    return validation


def _validate_manifest_evidence(rows: Iterable[dict[str, Any]]) -> list[str]:
    missing: set[str] = set()
    for row in rows:
        for key in ("path", "evidence_path"):
            value = str(row.get(key, "")).strip()
            if not value:
                continue
            if not _validate_path(value):
                missing.add(value)
    return sorted(missing)


def _write_human_summary(validation: dict[str, Any]) -> str:
    lines = [
        "# 六赛道模型总表汇总",
        "",
        f"- 统一工作簿：`{validation['workbook']['path']}`",
        f"- 单 Sheet：`{validation['workbook']['sheet_name']}`",
        f"- 行数：{validation['workbook']['row_count']}",
        f"- 列数：{validation['workbook']['column_count']}",
        f"- 六赛道：{', '.join(validation['workbook']['tracks'])}",
        "",
        "结论：",
        "- fault：data_blocked，当前没有可直接交付的有效 fault 开发分数。",
        "- facies：SAM2 预训练路径优于随机初始化，但仍低于强基线；修复后的门控残差也没有把它推过基线。",
        "- property：TabICL 的三目标开发集结果是有效的，PHIF / KLOGH / SW 的 RMSE 分别约改善 4.9%、13.6%、25.2%，但仍按 effect_supported_not_promoted 和重现阻断分开保留。",
        "- lithofacies：MOMENT 小幅优于随机初始化，但仍低于 XGBoost 基线。",
        "- sweetspot：T3 有提升，T4 总体不利，T5 blocked。",
        "- reconstruction：预训练比随机初始化约好 48.5%，但仍低于 PyKrige。",
        "",
        "校验：",
        f"- 所有非空 evidence_path 已通过存在性检查；总表包含 {validation['workbook']['track_count']} 个赛道。",
        f"- 组合后的状态分布：{validation['workbook']['status_counts']}",
        "",
        "说明：",
        "- 这里只合并已发布的赛道产物，不改科学数值，不读 holdout，不重训。",
    ]
    text = "\n".join(lines) + "\n"
    (OUTPUT_DIR / "human_summary.md").write_text(text, encoding="utf-8")
    return text


def build(output_dir: Path = OUTPUT_DIR) -> dict[str, Any]:
    global OUTPUT_DIR
    OUTPUT_DIR = output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_rows, figure_rows, table_rows, initial_validation = _collect_source_data()
    _write_workbook(workbook_rows, output_dir / "six_track_model_metrics.xlsx")
    # Generate the machine-readable validation artifact before indexing the generated files.
    validation = _validate_outputs(workbook_rows, figure_rows, table_rows)
    (output_dir / "field_validation_report.json").write_text(json.dumps(validation, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    human_summary_text = _write_human_summary(validation)
    _add_generated_artifacts(table_rows)

    figure_fieldnames = [
        "track",
        "artifact_type",
        "artifact_name",
        "kind",
        "dataset",
        "model_name",
        "target",
        "split_protocol",
        "path",
        "status",
        "source_kind",
        "evidence_path",
        "sha256",
        "row_count",
        "sheet_name",
        "notes",
    ]
    table_fieldnames = figure_fieldnames

    _write_csv(figure_rows, output_dir / "all_figures_manifest.csv", figure_fieldnames)
    _write_csv(table_rows, output_dir / "all_tables_manifest.csv", table_fieldnames)

    # Recompute after generated artifacts are indexed and cross-check evidence.
    final_validation = _validate_outputs(workbook_rows, figure_rows, table_rows)
    final_validation["manifest_evidence_missing"] = _validate_manifest_evidence(figure_rows + table_rows)
    final_validation["manifest_evidence_checked"] = len(figure_rows) + len(table_rows)
    if final_validation["manifest_evidence_missing"]:
        raise RuntimeError(
            "manifest evidence missing: " + ", ".join(final_validation["manifest_evidence_missing"][:10])
        )
    final_validation["generated_files"] = {
        "workbook": _repo_rel(output_dir / "six_track_model_metrics.xlsx"),
        "figures_manifest": _repo_rel(output_dir / "all_figures_manifest.csv"),
        "tables_manifest": _repo_rel(output_dir / "all_tables_manifest.csv"),
        "field_validation_report": _repo_rel(output_dir / "field_validation_report.json"),
        "human_summary": _repo_rel(output_dir / "human_summary.md"),
    }
    final_validation["initial_validation"] = initial_validation
    (output_dir / "field_validation_report.json").write_text(json.dumps(final_validation, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    # Re-open the workbook once more to ensure the final on-disk file is still valid.
    reopened = load_workbook(output_dir / "six_track_model_metrics.xlsx", read_only=True)
    if reopened.sheetnames != [EXPECTED_SHEET]:
        raise RuntimeError("final workbook validation failed")
    return {
        "output_dir": _repo_rel(output_dir),
        "rows": len(workbook_rows),
        "figures": len(figure_rows),
        "tables": len(table_rows),
        "tracks": list(CANONICAL_TRACKS),
        "human_summary": _repo_rel(output_dir / "human_summary.md"),
        "workbook": _repo_rel(output_dir / "six_track_model_metrics.xlsx"),
        "validation": _repo_rel(output_dir / "field_validation_report.json"),
    }


def main(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args(argv)
    summary = build(args.output_dir)
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
