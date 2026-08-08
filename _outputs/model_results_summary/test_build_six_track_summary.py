from __future__ import annotations

import csv
import importlib.util
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent


def _load_module():
    path = HERE / "build_six_track_summary.py"
    spec = importlib.util.spec_from_file_location("build_six_track_summary", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def test_build_six_track_summary(tmp_path: Path) -> None:
    mod = _load_module()
    summary = mod.build(tmp_path)
    workbook_path = tmp_path / "six_track_model_metrics.xlsx"
    figures_manifest_path = tmp_path / "all_figures_manifest.csv"
    tables_manifest_path = tmp_path / "all_tables_manifest.csv"
    validation_path = tmp_path / "field_validation_report.json"
    human_summary_path = tmp_path / "human_summary.md"

    assert workbook_path.is_file()
    assert figures_manifest_path.is_file()
    assert tables_manifest_path.is_file()
    assert validation_path.is_file()
    assert human_summary_path.is_file()

    wb = load_workbook(workbook_path, read_only=True)
    assert wb.sheetnames == ["模型指标"]
    ws = wb["模型指标"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert len(header) == 26
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1456

    track_idx = header.index("track")
    status_idx = header.index("status")
    evidence_idx = header.index("evidence_path")
    assert set(row[track_idx] for row in rows) == {"fault", "facies", "property", "lithofacies", "sweetspot", "reconstruction"}
    assert Counter(row[status_idx] for row in rows)["data_blocked"] > 0
    assert Counter(row[status_idx] for row in rows)["reference_only"] > 0
    assert Counter(row[status_idx] for row in rows)["effect_supported_not_promoted"] > 0
    assert all((not row[evidence_idx]) or Path(str(row[evidence_idx])).is_file() for row in rows)

    figures = list(csv.DictReader(figures_manifest_path.open(encoding="utf-8")))
    tables = list(csv.DictReader(tables_manifest_path.open(encoding="utf-8")))
    assert figures and tables
    assert all((not row.get("evidence_path")) or Path(row["evidence_path"]).is_file() for row in figures)
    assert all((not row.get("evidence_path")) or Path(row["evidence_path"]).is_file() for row in tables)
    assert any(row["path"].endswith("six_track_model_metrics.xlsx") for row in tables)
    assert any(row["path"].endswith("human_summary.md") for row in tables)
    assert "property" in {row[track_idx] for row in rows}
    assert summary["rows"] == 1456
    assert summary["tracks"] == ["fault", "facies", "property", "lithofacies", "sweetspot", "reconstruction"]
